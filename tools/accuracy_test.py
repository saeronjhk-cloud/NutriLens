#!/usr/bin/env python3
"""
NutriLens 정확도 테스트 도구 v2
────────────────────────────────
두 가지 방식으로 정확도를 측정합니다:

  1. DB 매칭 테스트 (--db)
     → 흔한 음식 이름 100개로 DB에서 영양정보를 찾을 수 있는지 확인
     → API 비용 없음, 즉시 실행

  2. 실제 사진 테스트 (--photo)
     → .tmp/test_images/ 폴더에 음식 사진을 넣으면 AI 인식 테스트
     → 파일명이 정답 (예: "김치.jpg", "비빔밥.jpg")

실행법:
  python tools/accuracy_test.py --db       # DB 매칭 테스트
  python tools/accuracy_test.py --photo    # 사진 인식 테스트
  python tools/accuracy_test.py --all      # 둘 다
"""

import os
import sys
import json
import time
import base64
import urllib.request
import urllib.error
from pathlib import Path
from datetime import datetime

# ── 경로 설정 ──
PROJECT_DIR = Path(__file__).parent.parent
TOOLS_DIR = Path(__file__).parent
TEST_DIR = PROJECT_DIR / ".tmp" / "test_images"
RESULT_DIR = PROJECT_DIR / ".tmp"
sys.path.insert(0, str(TOOLS_DIR))

# ── .env 로드 ──
def load_env():
    env_paths = [PROJECT_DIR / '.env', Path.cwd() / '.env']
    for p in env_paths:
        if p.exists():
            with open(p) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, val = line.split('=', 1)
                        os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))
            break

load_env()


# ══════════════════════════════════════════════════════════
#  1. DB 매칭 테스트
# ══════════════════════════════════════════════════════════

# 실제 한국인이 자주 먹는 음식 100개 (한식 + 다이어트 + 간식 + 외식)
COMMON_FOODS = [
    # 한식 (50개)
    "김치", "비빔밥", "불고기", "김밥", "떡볶이",
    "된장찌개", "삼겹살", "잡채", "순두부찌개", "갈비탕",
    "냉면", "김치찌개", "제육볶음", "파전", "감자탕",
    "칼국수", "삼계탕", "족발", "라면", "닭갈비",
    "볶음밥", "만두", "육회", "떡국", "순대",
    "치킨", "해물탕", "오징어볶음", "떡", "김치볶음밥",
    "갈비찜", "보쌈", "콩나물국", "미역국", "어묵",
    "김치전", "잔치국수", "비빔냉면", "소불고기", "돼지갈비",
    "멸치볶음", "두부조림", "감자조림", "계란찜", "깍두기",
    "시금치나물", "콩나물무침", "오이소박이", "무생채", "열무김치",
    # 다이어트/피트니스 (20개)
    "닭가슴살", "샐러드", "고구마", "계란", "현미밥",
    "오트밀", "연어", "아보카도", "두부", "바나나",
    "그릭요거트", "브로콜리", "퀴노아", "단호박", "토마토",
    "블루베리", "아몬드", "땅콩버터", "치아시드", "단백질보충제",
    # 프랜차이즈/외식 (15개)
    "햄버거", "피자", "파스타", "초밥", "돈까스",
    "짜장면", "짬뽕", "탕수육", "양장피", "마라탕",
    "쌀국수", "카레", "우동", "스테이크", "리조또",
    # 간식/음료 (15개)
    "아메리카노", "카페라떼", "녹차", "콜라", "사이다",
    "떡케이크", "호떡", "붕어빵", "아이스크림", "초콜릿",
    "과일주스", "스무디", "쿠키", "마카롱", "크로와상",
]


def run_db_test():
    """DB 매칭 정확도 테스트 — 흔한 음식이 DB에 있는지 확인"""
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl 설치 중...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    db_path = PROJECT_DIR / "NutriLens_음식DB.xlsx"
    if not db_path.exists():
        print(f"  DB 파일 없음: {db_path}")
        return None

    print()
    print("=" * 60)
    print("  NutriLens DB 매칭 테스트")
    print("=" * 60)
    print(f"  DB 파일: {db_path.name}")
    print(f"  테스트 음식: {len(COMMON_FOODS)}종")
    print()

    # DB 로드 — 음식 이름 목록 추출
    print("  DB 로딩 중...", end=" ", flush=True)
    wb = openpyxl.load_workbook(db_path, read_only=True, data_only=True)
    ws = wb.active

    db_names = set()
    db_names_lower = {}  # 소문자 → 원본 매핑
    rows_checked = 0
    name_col = None

    # 헤더에서 음식명 컬럼 찾기
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
        for cell in row:
            val = str(cell.value or "").strip()
            if val in ("식품명", "음식명", "food_name", "FOOD_NM_KR", "name"):
                name_col = cell.column
                break

    if name_col is None:
        name_col = 2  # 기본값: B열

    for row in ws.iter_rows(min_row=2, min_col=name_col, max_col=name_col, values_only=True):
        val = row[0]
        if val:
            name = str(val).strip()
            if name:
                db_names.add(name)
                db_names_lower[name.lower().replace(" ", "")] = name
        rows_checked += 1

    wb.close()
    print(f"OK ({len(db_names):,}종)")
    print()

    # 매칭 테스트
    results = []
    found = 0
    partial = 0
    not_found = 0

    for food_name in COMMON_FOODS:
        norm = food_name.lower().replace(" ", "")

        # 1) 정확 매칭
        if food_name in db_names or norm in db_names_lower:
            matched = db_names_lower.get(norm, food_name)
            results.append({"food": food_name, "status": "EXACT", "matched": matched})
            found += 1
            print(f"  ✓ {food_name} → 정확 일치 ({matched})")
            continue

        # 2) 부분 매칭 (DB에 포함된 이름)
        partial_matches = []
        for db_norm, db_orig in db_names_lower.items():
            if norm in db_norm or db_norm in norm:
                partial_matches.append(db_orig)
            elif len(norm) >= 2 and (db_norm.startswith(norm[:2]) and db_norm.endswith(norm[-1])):
                partial_matches.append(db_orig)

        if partial_matches:
            best = partial_matches[0]
            results.append({"food": food_name, "status": "PARTIAL", "matched": best, "candidates": len(partial_matches)})
            partial += 1
            print(f"  △ {food_name} → 부분 일치 ({best}, +{len(partial_matches)-1}개)")
            continue

        # 3) 매칭 실패
        results.append({"food": food_name, "status": "MISS", "matched": None})
        not_found += 1
        print(f"  ✗ {food_name} → DB에 없음")

    total = len(COMMON_FOODS)
    exact_pct = found / total * 100
    coverage_pct = (found + partial) / total * 100

    print()
    print("=" * 60)
    print(f"  DB 매칭 결과")
    print("=" * 60)
    print(f"  총 테스트: {total}종")
    print(f"  정확 일치: {found}종 ({exact_pct:.1f}%)")
    print(f"  부분 일치: {partial}종")
    print(f"  미발견:    {not_found}종")
    print(f"  커버리지:  {coverage_pct:.1f}% (정확+부분)")
    print("=" * 60)

    # 미발견 목록
    missed = [r["food"] for r in results if r["status"] == "MISS"]
    if missed:
        print(f"\n  DB에 없는 음식:")
        for m in missed:
            print(f"    - {m}")

    # 리포트 저장
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    report = {
        "test_type": "db_matching",
        "date": datetime.now().isoformat(),
        "db_size": len(db_names),
        "test_count": total,
        "exact_match": found,
        "partial_match": partial,
        "not_found": not_found,
        "exact_pct": round(exact_pct, 1),
        "coverage_pct": round(coverage_pct, 1),
        "missed_foods": missed,
        "details": results,
    }

    json_path = RESULT_DIR / "db_test_results.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"\n  결과 저장: {json_path}")

    return coverage_pct


# ══════════════════════════════════════════════════════════
#  2. 실제 사진 AI 인식 테스트
# ══════════════════════════════════════════════════════════

def call_openai_vision(base64_image, media_type, api_key):
    """GPT-4o Vision API 호출"""
    from food_analyzer import SYSTEM_PROMPT

    url = "https://api.openai.com/v1/chat/completions"
    payload = {
        "model": "gpt-4o",
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "이 사진에 있는 음식을 분석해주세요."},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{media_type};base64,{base64_image}",
                            "detail": "low"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 1000,
        "temperature": 0.1,
    }

    data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=data, method='POST')
    req.add_header("Authorization", f"Bearer {api_key}")
    req.add_header("Content-Type", "application/json")

    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace')[:200]
        return None, f"API 에러 ({e.code}): {body}"
    except Exception as e:
        return None, f"요청 실패: {e}"

    content = result["choices"][0]["message"]["content"]

    if "```json" in content:
        content = content.split("```json")[1].split("```")[0]
    elif "```" in content:
        content = content.split("```")[1].split("```")[0]

    try:
        parsed = json.loads(content.strip())
        return parsed, None
    except json.JSONDecodeError:
        return None, f"파싱 실패: {content[:200]}"


def normalize(name):
    name = name.strip().lower().replace(" ", "").replace("_", "")
    return name


def is_match(expected, ai_foods):
    expected_norm = normalize(expected)
    for food in ai_foods:
        ai_name = normalize(food.get("name_ko", "") or food.get("name", ""))
        if expected_norm == ai_name:
            return True, "정확 일치", ai_name
        if expected_norm in ai_name or ai_name in expected_norm:
            return True, "부분 일치", ai_name
        if len(expected_norm) >= 2 and len(ai_name) >= 2:
            if expected_norm[:2] == ai_name[:2] and expected_norm[-2:] == ai_name[-2:]:
                return True, "유사 일치", ai_name
    ai_names = [normalize(f.get("name_ko", "") or f.get("name", "")) for f in ai_foods]
    return False, "불일치", ", ".join(ai_names) if ai_names else "인식 없음"


def run_photo_test():
    """로컬 사진으로 AI 인식 정확도 테스트"""
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        print("  OPENAI_API_KEY가 설정되지 않았습니다.")
        return None

    if not TEST_DIR.exists():
        TEST_DIR.mkdir(parents=True, exist_ok=True)

    # jpg, jpeg, png 모두 지원
    images = sorted(
        [f for f in TEST_DIR.iterdir() if f.suffix.lower() in ('.jpg', '.jpeg', '.png')]
    )

    if not images:
        print()
        print("=" * 60)
        print("  사진 테스트 — 이미지 없음")
        print("=" * 60)
        print()
        print("  테스트 방법:")
        print(f"  1. {TEST_DIR} 폴더에 음식 사진을 넣으세요")
        print("  2. 파일명 = 정답 음식명 (예: 김치.jpg, 비빔밥.png)")
        print("  3. 다시 python tools/accuracy_test.py --photo 실행")
        print()
        print("  팁: 핸드폰으로 음식 사진 10~20장만 찍으면 충분합니다!")
        return None

    print()
    print("=" * 60)
    print("  NutriLens AI 사진 인식 테스트")
    print("=" * 60)
    print(f"  테스트 이미지: {len(images)}장")
    print(f"  모델: GPT-4o Vision")
    cost_est = len(images) * 0.01
    print(f"  예상 비용: ~${cost_est:.2f}")
    print()

    results = []
    correct = 0
    errors = 0

    for i, img_path in enumerate(images, 1):
        # 파일명에서 정답 추출
        expected_name = img_path.stem
        # "01_김치" 형태면 숫자 제거
        if "_" in expected_name and expected_name.split("_")[0].isdigit():
            expected_name = expected_name.split("_", 1)[1]

        print(f"  [{i:02d}/{len(images)}] {expected_name}...", end=" ", flush=True)

        # 이미지 → base64
        with open(img_path, 'rb') as f:
            image_data = f.read()
        base64_image = base64.b64encode(image_data).decode('utf-8')

        ext = img_path.suffix.lower()
        media_type = "image/png" if ext == ".png" else "image/jpeg"

        analysis, err = call_openai_vision(base64_image, media_type, api_key)

        if err:
            print(f"에러: {err}")
            results.append({
                "expected": expected_name, "result": "ERROR",
                "match_type": "에러", "ai_answer": str(err),
            })
            errors += 1
            time.sleep(1)
            continue

        foods = analysis.get("foods", [])
        matched, match_type, ai_answer = is_match(expected_name, foods)

        if matched:
            correct += 1
            print(f"✓ {match_type} (AI: {ai_answer})")
        else:
            print(f"✗ 불일치 (AI: {ai_answer})")

        results.append({
            "expected": expected_name,
            "result": "CORRECT" if matched else "WRONG",
            "match_type": match_type,
            "ai_answer": ai_answer,
            "confidence": foods[0].get("confidence", 0) if foods else 0,
        })
        time.sleep(1)

    total = len(results)
    accuracy = (correct / total * 100) if total > 0 else 0

    print()
    print("=" * 60)
    print(f"  사진 인식 테스트 결과")
    print("=" * 60)
    print(f"  총 테스트: {total}장")
    print(f"  정답: {correct}장")
    print(f"  오답: {total - correct - errors}장")
    print(f"  에러: {errors}장")
    print(f"  정확도: {accuracy:.1f}%")
    print("=" * 60)

    # 리포트 저장
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    report = {
        "test_type": "photo_recognition",
        "date": datetime.now().isoformat(),
        "model": "gpt-4o",
        "total": total,
        "correct": correct,
        "errors": errors,
        "accuracy_pct": round(accuracy, 1),
        "details": results,
    }

    json_path = RESULT_DIR / "photo_test_results.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"\n  결과 저장: {json_path}")

    # 텍스트 리포트
    report_path = RESULT_DIR / "accuracy_report.txt"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("NutriLens AI 정확도 테스트 리포트\n")
        f.write(f"테스트 일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
        f.write(f"모델: GPT-4o Vision\n")
        f.write("=" * 60 + "\n\n")
        f.write(f"총 테스트: {total}장\n")
        f.write(f"정답: {correct}장 / 오답: {total-correct-errors}장 / 에러: {errors}장\n")
        f.write(f"정확도: {accuracy:.1f}%\n\n")
        f.write("-" * 60 + "\n")
        for r in results:
            status = "✓" if r["result"] == "CORRECT" else "✗"
            f.write(f"{status} {r['expected']} → {r['ai_answer']} ({r['match_type']})\n")
    print(f"  텍스트 리포트: {report_path}")

    return accuracy


# ══════════════════════════════════════════════════════════
#  메인
# ══════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(description="NutriLens 정확도 테스트 v2")
    parser.add_argument("--db", action="store_true", help="DB 매칭 테스트 (API 비용 없음)")
    parser.add_argument("--photo", action="store_true", help="사진 인식 테스트 (GPT-4o)")
    parser.add_argument("--all", action="store_true", help="DB + 사진 둘 다")
    args = parser.parse_args()

    if not args.db and not args.photo and not args.all:
        print()
        print("NutriLens 정확도 테스트 v2")
        print()
        print("사용법:")
        print("  python tools/accuracy_test.py --db      # DB 매칭 (무료, 즉시)")
        print("  python tools/accuracy_test.py --photo   # 사진 인식 (~$0.01/장)")
        print("  python tools/accuracy_test.py --all     # 둘 다")
        print()
        print("사진 테스트 준비:")
        print(f"  1. {TEST_DIR} 폴더에 음식 사진 넣기")
        print("  2. 파일명 = 정답 (예: 김치.jpg, 삼겹살.png)")
        print("  3. --photo 실행")
        sys.exit(0)

    if args.db or args.all:
        db_coverage = run_db_test()
        if db_coverage is not None:
            if db_coverage >= 90:
                print(f"\n  ✓ DB 커버리지 {db_coverage:.1f}% — 우수!")
            elif db_coverage >= 70:
                print(f"\n  △ DB 커버리지 {db_coverage:.1f}% — 양호")
            else:
                print(f"\n  ✗ DB 커버리지 {db_coverage:.1f}% — 개선 필요")

    if args.photo or args.all:
        accuracy = run_photo_test()
        if accuracy is not None:
            if accuracy >= 80:
                print(f"\n  ✓ 인식 정확도 {accuracy:.1f}% — 양호!")
            elif accuracy >= 60:
                print(f"\n  △ 인식 정확도 {accuracy:.1f}% — 개선 필요")
            else:
                print(f"\n  ✗ 인식 정확도 {accuracy:.1f}% — 심각한 개선 필요")


if __name__ == "__main__":
    main()
