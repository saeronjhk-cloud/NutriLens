#!/usr/bin/env python3
"""
NutriLens DB 통합 도구
─────────────────────
여러 소스의 영양 데이터를 기존 DB에 합칩니다.

소스:
  1. 가공식품 CSV (공공데이터포털)
  2. SR Legacy JSON (USDA)

실행법:
  python tools/merge_databases.py

결과: NutriLens_음식DB.xlsx 에 신규 데이터 추가
"""

import os
import sys
import csv
import json
import argparse
from pathlib import Path
from collections import Counter

PROJECT_DIR = Path(__file__).parent.parent

# ── 파일 경로 ──
DB_PATH = PROJECT_DIR / "NutriLens_음식DB.xlsx"
PROCESSED_CSV = PROJECT_DIR / "전국통합식품영양성분정보_가공식품_표준데이터.csv"
SR_LEGACY_JSON = PROJECT_DIR / "FoodData_Central_sr_legacy_food_json_2018-04.json"

# 업로드 폴더 대체 경로
UPLOADS_DIR = Path("/sessions/tender-friendly-feynman/mnt/uploads")
PROCESSED_CSV_ALT = UPLOADS_DIR / "전국통합식품영양성분정보_가공식품_표준데이터.csv"
SR_LEGACY_JSON_ALT = UPLOADS_DIR / "FoodData_Central_sr_legacy_food_json_2018-04.json"


def safe_float(val):
    if val is None or val == "" or val == "-" or val == "N/A":
        return 0.0
    try:
        return round(float(str(val).replace(",", "")), 1)
    except (ValueError, TypeError):
        return 0.0


def load_existing_db():
    """기존 엑셀 DB 로드"""
    if not DB_PATH.exists():
        print(f"  기존 DB 없음: {DB_PATH}")
        return [], set()

    try:
        import openpyxl
    except ImportError:
        print("openpyxl 필요: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(DB_PATH, read_only=True, data_only=True)
    ws = wb['음식DB_전체']
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    existing = []
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            food = dict(zip(headers, row))
            existing.append(food)
            name = str(food.get("name_ko", "")).strip()
            if name:
                names.add(name)
    wb.close()
    return existing, names


def classify_food(name, category_name="", maker=""):
    """음식 카테고리 자동 분류"""
    name_lower = (name or "").lower()
    cat = (category_name or "").lower()
    maker_str = (maker or "").lower()

    franchise_keywords = [
        "맥도날드", "버거킹", "롯데리아", "kfc", "서브웨이",
        "스타벅스", "파리바게뜨", "뚜레쥬르", "배스킨라빈스",
        "교촌", "bbq", "bhc", "굽네", "네네", "도미노",
        "피자헛", "미스터피자", "이삭토스트", "맘스터치",
    ]
    for kw in franchise_keywords:
        if kw in name_lower or kw in maker_str:
            return "franchise"

    foreign_keywords = [
        "파스타", "피자", "햄버거", "스테이크", "샌드위치",
        "타코", "카레", "스시", "라멘", "우동", "짬뽕",
        "탕수육", "마파두부", "양장피",
    ]
    for kw in foreign_keywords:
        if kw in name_lower:
            return "foreign_popular"

    diet_keywords = [
        "프로틴", "protein", "닭가슴살", "샐러드", "그래놀라",
        "오트밀", "제로", "zero", "다이어트", "diet", "저칼로리",
        "무설탕", "sugar free", "식이섬유",
    ]
    for kw in diet_keywords:
        if kw in name_lower:
            return "diet_fitness"

    snack_keywords = [
        "과자", "초콜릿", "사탕", "젤리", "아이스크림",
        "음료", "주스", "커피", "차", "우유", "요구르트",
        "빵", "케이크", "쿠키", "캔디", "껌", "시리얼",
        "음료류", "과자류", "빙과류", "당류",
    ]
    for kw in snack_keywords:
        if kw in name_lower or kw in cat:
            return "snack_drink"

    return "korean"


def load_processed_csv():
    """가공식품 CSV 로드"""
    csv_path = PROCESSED_CSV if PROCESSED_CSV.exists() else PROCESSED_CSV_ALT
    if not csv_path.exists():
        print(f"  가공식품 CSV 없음: {PROCESSED_CSV}")
        return []

    foods = []
    # cp949 인코딩 시도
    for enc in ['cp949', 'euc-kr', 'utf-8']:
        try:
            with open(csv_path, encoding=enc, errors='replace') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    name = (row.get("식품명") or "").strip()
                    if not name:
                        continue

                    category_name = row.get("식품대분류명", "") or ""
                    maker = row.get("제조사명", "") or ""
                    food_cd = row.get("식품코드", "") or ""

                    foods.append({
                        "food_id": f"PROC_{food_cd}" if food_cd else f"PROC_{len(foods)}",
                        "name_ko": name,
                        "name_en": "",
                        "category": classify_food(name, category_name, maker),
                        "subcategory": row.get("식품소분류명", "") or "",
                        "serving_size_g": safe_float(row.get("영양성분함량기준량", "").replace("g", "").replace("ml", "")),
                        "calories_kcal": safe_float(row.get("에너지(kcal)")),
                        "protein_g": safe_float(row.get("단백질(g)")),
                        "carbs_g": safe_float(row.get("탄수화물(g)")),
                        "fat_g": safe_float(row.get("지방(g)")),
                        "fiber_g": safe_float(row.get("식이섬유(g)")),
                        "sodium_mg": safe_float(row.get("나트륨(mg)")),
                        "sugar_g": safe_float(row.get("당류(g)")),
                        "sat_fat_g": safe_float(row.get("포화지방산(g)")),
                        "cholesterol_mg": safe_float(row.get("콜레스테롤(mg)")),
                        "source": "가공식품CSV",
                        "brand": maker,
                        "cooking_method": "",
                        "tags": f"가공식품,{category_name}" if category_name else "가공식품",
                    })
            print(f"  가공식품 CSV: {len(foods)}종 로드 ({enc})")
            return foods
        except Exception as e:
            continue

    print(f"  가공식품 CSV 로드 실패")
    return []


def load_sr_legacy():
    """USDA SR Legacy JSON 로드"""
    json_path = SR_LEGACY_JSON if SR_LEGACY_JSON.exists() else SR_LEGACY_JSON_ALT
    if not json_path.exists():
        print(f"  SR Legacy JSON 없음: {SR_LEGACY_JSON}")
        return []

    print("  SR Legacy JSON 로딩 중 (200MB+, 시간이 좀 걸립니다)...")
    with open(json_path, encoding='utf-8') as f:
        data = json.load(f)

    raw_foods = data.get("SRLegacyFoods", [])
    foods = []

    for item in raw_foods:
        name = (item.get("description") or "").strip()
        if not name:
            continue

        # 영양소 추출
        nutrients = {}
        for n in item.get("foodNutrients", []):
            nutr = n.get("nutrient", {})
            nutr_name = nutr.get("name", "")
            amount = n.get("amount", 0)
            nutrients[nutr_name] = amount

        foods.append({
            "food_id": f"USDA_{item.get('fdcId', len(foods))}",
            "name_ko": "",
            "name_en": name,
            "category": classify_food(name),
            "subcategory": item.get("foodCategory", {}).get("description", "") if isinstance(item.get("foodCategory"), dict) else "",
            "serving_size_g": 100.0,  # SR Legacy는 100g 기준
            "calories_kcal": safe_float(nutrients.get("Energy", 0)),
            "protein_g": safe_float(nutrients.get("Protein", 0)),
            "carbs_g": safe_float(nutrients.get("Carbohydrate, by difference", 0)),
            "fat_g": safe_float(nutrients.get("Total lipid (fat)", 0)),
            "fiber_g": safe_float(nutrients.get("Fiber, total dietary", 0)),
            "sodium_mg": safe_float(nutrients.get("Sodium, Na", 0)),
            "sugar_g": safe_float(nutrients.get("Sugars, total including NLEA", nutrients.get("Sugars, total", 0))),
            "sat_fat_g": safe_float(nutrients.get("Fatty acids, total saturated", 0)),
            "cholesterol_mg": safe_float(nutrients.get("Cholesterol", 0)),
            "source": "USDA_SR_Legacy",
            "brand": "",
            "cooking_method": "",
            "tags": f"USDA,영어",
        })

    print(f"  SR Legacy: {len(foods)}종 로드")
    return foods


def save_to_excel(all_foods):
    """전체 데이터를 엑셀로 저장"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl 필요: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "음식DB_전체"

    headers = [
        "food_id", "name_ko", "name_en", "category", "subcategory",
        "serving_size_g", "calories_kcal", "protein_g", "carbs_g", "fat_g",
        "fiber_g", "sodium_mg", "sugar_g", "sat_fat_g", "cholesterol_mg",
        "source", "brand", "cooking_method", "tags"
    ]

    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='333333'),
        right=Side(style='thin', color='333333'),
        top=Side(style='thin', color='333333'),
        bottom=Side(style='thin', color='333333')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, food in enumerate(all_foods, 2):
        for col_idx, h in enumerate(headers, 1):
            val = food.get(h, "")
            ws.cell(row=row_idx, column=col_idx, value=val)

    # 열 너비 자동 조정
    col_widths = {
        "food_id": 18, "name_ko": 25, "name_en": 35, "category": 14,
        "subcategory": 18, "serving_size_g": 12, "calories_kcal": 12,
        "protein_g": 10, "carbs_g": 10, "fat_g": 10, "fiber_g": 10,
        "sodium_mg": 10, "sugar_g": 10, "sat_fat_g": 10, "cholesterol_mg": 12,
        "source": 15, "brand": 18, "cooking_method": 12, "tags": 20,
    }
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 12)

    # ── 통계 시트 ──
    ws2 = wb.create_sheet("통계")
    ws2.cell(row=1, column=1, value="소스").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="카테고리").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="수량").font = Font(bold=True)

    source_counts = Counter(f.get("source", "") for f in all_foods)
    cat_counts = Counter(f.get("category", "") for f in all_foods)

    row = 2
    ws2.cell(row=row, column=1, value="=== 소스별 ===").font = Font(bold=True)
    row += 1
    for src, cnt in source_counts.most_common():
        ws2.cell(row=row, column=1, value=src)
        ws2.cell(row=row, column=3, value=cnt)
        row += 1

    row += 1
    ws2.cell(row=row, column=1, value="=== 카테고리별 ===").font = Font(bold=True)
    row += 1
    cat_names = {
        "korean": "한식", "diet_fitness": "다이어트/피트니스",
        "foreign_popular": "외국음식", "franchise": "프랜차이즈",
        "snack_drink": "간식/음료",
    }
    for cat, cnt in cat_counts.most_common():
        ws2.cell(row=row, column=1, value=cat_names.get(cat, cat))
        ws2.cell(row=row, column=3, value=cnt)
        row += 1

    row += 1
    ws2.cell(row=row, column=1, value="총계").font = Font(bold=True)
    ws2.cell(row=row, column=3, value=len(all_foods)).font = Font(bold=True)

    for col in range(1, 4):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    wb.save(DB_PATH)
    print(f"\n  DB 저장 완료: {DB_PATH}")


def main():
    print()
    print("=" * 55)
    print("  NutriLens DB 통합")
    print("=" * 55)

    # 1. 기존 DB 로드
    print("\n[1/4] 기존 DB 로드...")
    existing, existing_names = load_existing_db()
    print(f"  기존 DB: {len(existing)}종 (중복 체크용 이름: {len(existing_names)}개)")

    # 2. 가공식품 CSV 로드
    print("\n[2/4] 가공식품 데이터 로드...")
    processed_foods = load_processed_csv()

    # 3. SR Legacy 로드
    print("\n[3/4] USDA SR Legacy 데이터 로드...")
    sr_foods = load_sr_legacy()

    # 4. 통합 (중복 제거)
    print("\n[4/4] DB 통합 중...")
    new_foods = []
    duplicates = 0

    for food in processed_foods + sr_foods:
        name_ko = str(food.get("name_ko", "")).strip()
        name_en = str(food.get("name_en", "")).strip()
        check_name = name_ko or name_en

        if check_name and check_name in existing_names:
            duplicates += 1
            continue

        if check_name:
            existing_names.add(check_name)
        new_foods.append(food)

    print(f"  신규 추가: {len(new_foods)}종")
    print(f"  중복 제외: {duplicates}종")

    if len(new_foods) == 0:
        print("\n  추가할 새로운 데이터가 없습니다.")
        return

    # 소스별 통계
    src_counts = Counter(f.get("source", "") for f in new_foods)
    print(f"\n  신규 데이터 소스별:")
    for src, cnt in src_counts.most_common():
        print(f"    {src}: {cnt}종")

    # 카테고리 통계
    cat_counts = Counter(f.get("category", "") for f in new_foods)
    cat_names = {
        "korean": "한식", "diet_fitness": "다이어트/피트니스",
        "foreign_popular": "외국음식", "franchise": "프랜차이즈",
        "snack_drink": "간식/음료",
    }
    print(f"\n  카테고리 분포:")
    for cat, cnt in cat_counts.most_common():
        print(f"    {cat_names.get(cat, cat)}: {cnt}종")

    # 합치기
    all_foods = existing + new_foods
    print(f"\n  최종 DB: 기존 {len(existing)} + 신규 {len(new_foods)} = 총 {len(all_foods)}종")

    # 저장
    print("\n  엑셀 저장 중 (대용량이라 시간이 걸립니다)...")
    save_to_excel(all_foods)

    print(f"\n  총 {len(all_foods)}종")
    print("=" * 55)
    print("  통합 완료!")
    print("=" * 55)


if __name__ == "__main__":
    main()
