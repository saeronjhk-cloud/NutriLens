#!/usr/bin/env python3
"""
엑셀 DB → SQLite 변환
서버 시작 속도를 대폭 개선하고 메모리 사용량을 줄입니다.
"""
import json
import sys
import sqlite3
import time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

PROJECT_DIR = Path(__file__).parent.parent
DB_XLSX = PROJECT_DIR / "NutriLens_음식DB.xlsx"
DB_SQLITE = PROJECT_DIR / "nutrilens_db.sqlite"

print(f"엑셀 DB 로딩: {DB_XLSX.name}")
wb = openpyxl.load_workbook(DB_XLSX, read_only=True, data_only=True)
ws = wb["음식DB_전체"]

headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"컬럼: {len(headers)}개")

# SQLite DB 생성
# ★ 2026-07-17 세션32 수정: 기존 코드는 unlink 실패를 `except: pass` 로 삼키고
#   "삭제 실패해도 덮어쓰기 가능" 이라 주석했으나 **사실이 아니다** — 곧바로
#   `sqlite3.OperationalError: table foods already exists` 로 죽는다.
#   실제로 일부 마운트(FUSE 등)에서 unlink 가 막힌다(IP/148 §3.1: "unlink 는 막히고 rename 은 된다").
#   → ① unlink 시도 ② 실패하면 rename 으로 치움 ③ 그래도 안 되면 DROP TABLE 로 이어간다.
if DB_SQLITE.exists():
    try:
        DB_SQLITE.unlink()
    except Exception as e:
        stale = DB_SQLITE.with_suffix(".sqlite.stale_%d" % int(time.time()))
        try:
            DB_SQLITE.rename(stale)
            print(f"  기존 DB 삭제 불가({e.__class__.__name__}) → {stale.name} 으로 옮김")
        except Exception:
            print(f"  기존 DB 삭제·이동 모두 불가({e.__class__.__name__}) → DROP TABLE 로 진행")

conn = sqlite3.connect(str(DB_SQLITE))
cur = conn.cursor()

cur.execute("DROP TABLE IF EXISTS foods")
cur.execute("""
    CREATE TABLE foods (
        food_id TEXT,
        name_ko TEXT NOT NULL,
        category TEXT DEFAULT '',
        subcategory TEXT DEFAULT '',
        serving_size_g REAL DEFAULT 100,
        calories_kcal REAL DEFAULT 0,
        protein_g REAL DEFAULT 0,
        carbs_g REAL DEFAULT 0,
        fat_g REAL DEFAULT 0,
        fiber_g REAL DEFAULT 0,
        sodium_mg REAL DEFAULT 0,
        sugar_g REAL DEFAULT 0
    )
""")

# 컬럼 인덱스 매핑
col_map = {}
for i, h in enumerate(headers):
    col_map[h] = i

FIELDS = ["food_id", "name_ko", "category", "subcategory", "serving_size_g",
           "calories_kcal", "protein_g", "carbs_g", "fat_g", "fiber_g", "sodium_mg", "sugar_g"]

count = 0
batch = []
# ★ 2026-07-17 세션32: food_id 충돌 감지 (IP/152 §3-2)
#   실제 사고 — KR-0416 이 '쌈밥 정식'(엑셀 301행)과 '대하구이'(302행) 두 음식에 붙어 있었다.
#   name_ko 는 유일하므로 영양소 조회(search_food_db)는 멀쩡하지만, GI 표는 food_id 로 조회하므로
#   대하구이가 쌈밥 정식의 GI 를 받는다. resolve_gi 의 T1 이 가용탄수 게이트보다 먼저 걸리기 때문.
#   → 조용히 넘어가면 안 된다. 빌드에서 잡는다.
seen_food_id = {}
collisions = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name_idx = col_map.get("name_ko", 1)
    if not row[name_idx]:
        continue

    _fid_idx = col_map.get("food_id")
    if _fid_idx is not None and _fid_idx < len(row):
        _fid = str(row[_fid_idx] or "").strip()
        _nm = str(row[name_idx] or "").strip()
        if _fid:
            if _fid in seen_food_id and seen_food_id[_fid][1] != _nm:
                collisions.append((_fid, seen_food_id[_fid], (count + 2, _nm)))
            else:
                seen_food_id.setdefault(_fid, (count + 2, _nm))

    values = []
    for field in FIELDS:
        idx = col_map.get(field)
        if idx is not None and idx < len(row):
            val = row[idx]
            if val is None:
                val = "" if field in ("food_id", "name_ko", "category", "subcategory") else 0
            values.append(val)
        else:
            values.append("" if field in ("food_id", "name_ko", "category", "subcategory") else 0)

    batch.append(tuple(values))
    count += 1

    if len(batch) >= 10000:
        cur.executemany(f"INSERT INTO foods ({','.join(FIELDS)}) VALUES ({','.join('?' * len(FIELDS))})", batch)
        conn.commit()
        batch = []
        print(f"  {count:,}종 처리...")

if batch:
    cur.executemany(f"INSERT INTO foods ({','.join(FIELDS)}) VALUES ({','.join('?' * len(FIELDS))})", batch)
    conn.commit()

# 인덱스 생성 (검색 속도 향상)
print("인덱스 생성 중...")
cur.execute("CREATE INDEX idx_name_ko ON foods(name_ko)")
cur.execute("CREATE INDEX idx_category ON foods(category)")
conn.close()
wb.close()

size_mb = DB_SQLITE.stat().st_size / 1024 / 1024
print(f"\n완료! {count:,}종 → {DB_SQLITE.name} ({size_mb:.1f}MB)")

# ── food_id 충돌 보고 (빌드 성공 후, 크게) ──────────────────────────────────
# 실패로 처리하지 않는 이유: 이 스크립트가 죽으면 엔진이 쓸 DB 자체가 없어진다.
# 대신 exit code 를 1 로 주어 CI/자동화에서는 잡히게 한다.
if collisions:
    print("\n" + "=" * 72)
    print("⚠️  food_id 충돌 %d건 — 서로 다른 음식이 같은 ID 를 씁니다" % len(collisions))
    print("=" * 72)
    for fid, (r1, n1), (r2, n2) in collisions:
        print("  %s : 엑셀 %d행 '%s'  ↔  엑셀 %d행 '%s'" % (fid, r1, n1, r2, n2))
    print("\n  영향: 영양소 조회는 name_ko(유일) 기준이라 정상. 그러나 **GI 표는 food_id 로 조회**하므로")
    print("        뒤 음식이 앞 음식의 GI 를 받습니다(resolve_gi 의 T1 이 가용탄수 게이트보다 먼저).")
    print("  조치: 엑셀 원본에서 **뒤 행**에 미사용 food_id 를 부여하십시오.")
    print("        GI 표(IP/content/gi_table_v1.csv)에 등록된 쪽이 기존 ID 를 유지해야 합니다.")
    print("        점검 스크립트: python tools/check_food_id_collision.py")
    print("=" * 72)
    sys.exit(1)
else:
    print("food_id 충돌 점검: 이상 없음")
