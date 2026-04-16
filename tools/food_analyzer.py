#!/usr/bin/env python3
"""
NutriLens AI 음식 분석 엔진 (2단계 핵심 도구)
─────────────────────────────────────────────
사진 한 장 → AI가 음식 인식 → DB에서 영양소 매칭 → 결과 반환

사용법:
  python food_analyzer.py --image 사진경로.jpg
  python food_analyzer.py --image 사진경로.jpg --api-key YOUR_OPENAI_KEY

환경변수:
  OPENAI_API_KEY: OpenAI API 키 (.env 파일에 저장)
"""

import os
import sys
import json
import base64
import argparse
from pathlib import Path

# ── .env 파일 로드 ──
def load_env():
    env_paths = [
        Path(__file__).parent.parent / '.env',
        Path.cwd() / '.env',
    ]
    for env_path in env_paths:
        if env_path.exists():
            with open(env_path) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, val = line.split('=', 1)
                        os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))
            break

load_env()

# ── 음식 DB 로드 ──
# DB 연결을 전역으로 유지 (SQLite 검색용)
_DB_CONN = None
_DB_PATH = None

def load_food_db(db_path=None):
    """SQLite DB 연결을 열고 아이템 수만 반환 (전체를 메모리에 올리지 않음)"""
    global _DB_CONN, _DB_PATH
    import sqlite3

    if db_path is None:
        sqlite_path = Path(__file__).parent.parent / 'nutrilens_db.sqlite'
        xlsx_path = Path(__file__).parent.parent / 'NutriLens_음식DB.xlsx'

        if sqlite_path.exists():
            db_path = sqlite_path
        elif xlsx_path.exists():
            return _load_xlsx(xlsx_path)
        else:
            print("DB 파일을 찾을 수 없습니다.")
            return []

    p = Path(db_path)
    if p.suffix != '.sqlite':
        return _load_xlsx(p)

    # SQLite: 연결만 열고 전체 로드 안 함
    _DB_CONN = sqlite3.connect(str(p))
    _DB_CONN.row_factory = sqlite3.Row
    _DB_PATH = str(p)

    cur = _DB_CONN.cursor()
    cur.execute("SELECT COUNT(*) FROM foods")
    count = cur.fetchone()[0]

    # 호환성: 기존 코드에서 len(FOODS_DB) 쓰므로 count를 가진 리스트처럼 보이는 객체 반환
    return _FoodDBProxy(count)


class _FoodDBProxy:
    """len()과 bool() 호환용 프록시 — 실제 데이터는 SQLite에서 검색"""
    def __init__(self, count):
        self._count = count
    def __len__(self):
        return self._count
    def __bool__(self):
        return self._count > 0


def _is_relevant_match(query, db_name):
    """부분 매칭 결과가 실제로 관련 있는지 판단
    예: "귤" → "레귤러 피자"는 관련 없음 (False)
    예: "김치찌개" → "김치찌개(돼지고기)"는 관련 있음 (True)
    """
    query = query.strip()
    db_name = db_name.strip()

    # DB 이름이 query로 시작하면 → 관련 있음 (예: "귤" → "귤차")
    if db_name.startswith(query):
        return True

    # DB 이름에 query가 독립적 단어로 포함되는지 확인
    # 구분자: 공백, _, (, ), /, 등
    import re
    # query 앞뒤가 단어 경계이거나 구분자인 경우만 허용
    # 예: "귤"이 "레귤러"에 포함 → 앞에 "레"가 붙어있으므로 False
    # 예: "귤"이 "귤차"에 포함 → 귤로 시작하므로 True (위에서 이미 처리)
    # 예: "김치"가 "김치찌개"에 포함 → 김치로 시작하므로 True
    idx = db_name.find(query)
    if idx < 0:
        return False

    # query 앞의 문자가 구분자이면 OK
    if idx > 0:
        prev_char = db_name[idx - 1]
        if prev_char not in ' _(/·\t':
            return False

    return True


def search_food_db(name):
    """SQLite에서 음식 이름으로 빠르게 검색 (인덱스 활용)"""
    global _DB_CONN
    if _DB_CONN is None:
        return None, []

    cur = _DB_CONN.cursor()

    # 1) 정확 매칭 (인덱스 사용 → 즉시)
    cur.execute("SELECT * FROM foods WHERE name_ko = ? LIMIT 1", (name,))
    row = cur.fetchone()
    if row:
        return dict(row), []

    # 2) 부분 매칭 (LIKE) + 관련성 필터
    cur.execute("SELECT * FROM foods WHERE name_ko LIKE ? LIMIT 20", (f"%{name}%",))
    raw_partials = [dict(r) for r in cur.fetchall()]
    # 엉뚱한 매칭 걸러내기 (예: "귤" → "레귤러 피자" 제거)
    partials = [r for r in raw_partials if _is_relevant_match(name, r.get('name_ko', ''))]
    if partials:
        return None, partials[:5]

    # 3) 역방향 부분 매칭: DB 이름이 AI 이름에 포함되는 경우
    # 예: AI가 "김치찌개" → DB에 "김치" 매칭
    if len(name) >= 3:  # 최소 3글자 이상일 때만 역방향 매칭
        cur.execute("SELECT * FROM foods WHERE ? LIKE '%' || name_ko || '%' AND length(name_ko) >= 2 LIMIT 5", (name,))
        reverse_partials = [dict(r) for r in cur.fetchall()]
        if reverse_partials:
            return None, reverse_partials

    return None, []


def _load_sqlite(db_path):
    """SQLite에서 전체 로드 — 하위 호환용 (accuracy_test 등)"""
    import sqlite3
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM foods")
    foods = [dict(row) for row in cur.fetchall()]
    conn.close()
    return foods


def _load_xlsx(db_path):
    """엑셀에서 로드 — SQLite 없을 때 fallback"""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl 필요: pip install openpyxl --break-system-packages")
        sys.exit(1)

    if not Path(db_path).exists():
        print(f"DB 파일을 찾을 수 없습니다: {db_path}")
        return []

    wb = openpyxl.load_workbook(db_path, read_only=True, data_only=True)
    ws = wb['음식DB_전체']

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    foods = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            food = dict(zip(headers, row))
            foods.append(food)
    wb.close()
    return foods


def build_food_list_for_prompt(foods):
    """DB 음식 목록을 프롬프트에 넣을 간결한 텍스트로 변환"""
    categories = {}
    for f in foods:
        cat = f.get('category', 'other')
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(f['name_ko'])

    lines = []
    cat_names = {
        'korean': '한식',
        'diet_fitness': '다이어트/피트니스',
        'foreign_popular': '외국음식',
        'franchise': '프랜차이즈',
        'snack_drink': '간식/음료',
    }
    for cat, names in categories.items():
        label = cat_names.get(cat, cat)
        lines.append(f"[{label}] {', '.join(names[:50])}")  # 카테고리당 최대 50개
    return '\n'.join(lines)


# ── AI 분석 프롬프트 ──
SYSTEM_PROMPT = """당신은 NutriLens의 AI 음식 영양 분석 전문가입니다.

## 역할
사진에 보이는 모든 음식을 정확하게 식별하고, 각 음식의 영양 성분을 분석합니다.

## 규칙
1. 사진에 보이는 음식을 **모두** 개별적으로 식별하세요 (메인, 사이드, 음료 포함)
2. 각 음식의 **예상 양(g)**을 추정하세요 (접시/그릇 크기 참고)
3. 영양소는 추정 양 기준으로 계산하세요
4. 한국 음식이면 category를 "korean", 다이어트 식단이면 "diet_fitness", 외국 음식이면 "foreign_popular", 프랜차이즈면 "franchise", 간식/음료면 "snack_drink"으로 분류
5. confidence는 0.0~1.0 (1.0 = 100% 확신)
6. 프랜차이즈 메뉴가 보이면 brand 필드에 브랜드명 포함
7. 반드시 아래 JSON 형식으로만 답변하세요. 다른 텍스트 없이 JSON만 출력하세요.

## 응답 형식
```json
{
  "foods": [
    {
      "name_ko": "음식 한국어 이름",
      "name_en": "English Name",
      "category": "korean",
      "subcategory": "밥류",
      "estimated_serving_g": 300,
      "calories_kcal": 450,
      "protein_g": 15,
      "carbs_g": 65,
      "fat_g": 12,
      "fiber_g": 2,
      "sodium_mg": 800,
      "sugar_g": 3,
      "confidence": 0.92,
      "brand": "",
      "cooking_method": "볶음",
      "tags": "한끼,매운맛"
    }
  ],
  "meal_summary": {
    "total_calories": 450,
    "total_protein": 15,
    "total_carbs": 65,
    "total_fat": 12,
    "meal_type": "점심",
    "health_score": 7,
    "one_line_comment": "균형 잡힌 한끼입니다. 식이섬유를 위해 채소를 추가하면 더 좋아요."
  }
}
```

## health_score 기준 (1~10)
- 9~10: 영양 균형 우수, 저나트륨, 적정 칼로리
- 7~8: 대체로 양호, 약간의 개선 여지
- 5~6: 보통, 특정 영양소 과다/부족
- 3~4: 영양 불균형, 고칼로리/고나트륨
- 1~2: 매우 불균형, 과도한 열량/지방/나트륨"""


def encode_image(image_path):
    """이미지를 base64로 인코딩"""
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def analyze_food_image(image_path, api_key=None, model="gpt-4o"):
    """
    음식 사진을 GPT-4o Vision으로 분석

    Args:
        image_path: 이미지 파일 경로
        api_key: OpenAI API 키 (없으면 환경변수에서 가져옴)
        model: 사용할 모델 (기본: gpt-4o)

    Returns:
        dict: 분석 결과 (JSON)
    """
    try:
        import httpx
    except ImportError:
        try:
            import requests as httpx
            httpx.Client = None  # fallback marker
        except ImportError:
            print("httpx 또는 requests 필요: pip install httpx --break-system-packages")
            sys.exit(1)

    if api_key is None:
        api_key = os.environ.get('OPENAI_API_KEY')

    if not api_key:
        return {
            "error": "OPENAI_API_KEY가 설정되지 않았습니다.",
            "help": "1) .env 파일에 OPENAI_API_KEY=sk-... 추가하거나\n"
                    "2) --api-key 옵션으로 전달하세요.\n"
                    "3) OpenAI API 키는 https://platform.openai.com/api-keys 에서 발급"
        }

    if not Path(image_path).exists():
        return {"error": f"이미지 파일을 찾을 수 없습니다: {image_path}"}

    # 이미지 인코딩
    base64_image = encode_image(image_path)
    ext = Path(image_path).suffix.lower()
    media_type = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.gif': 'image/gif',
        '.webp': 'image/webp',
    }.get(ext, 'image/jpeg')

    # API 호출
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "이 사진에 있는 음식을 분석해주세요. 모든 음식을 개별적으로 식별하고 영양 성분을 알려주세요."
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{media_type};base64,{base64_image}",
                            "detail": "high"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 2000,
        "temperature": 0.2,
    }

    try:
        import httpx
        response = httpx.post(
            "https://api.openai.com/v1/chat/completions",
            headers=headers,
            json=payload,
            timeout=30.0,
        )
        result = response.json()
    except Exception as e:
        return {"error": f"API 호출 실패: {str(e)}"}

    if "error" in result:
        return {"error": f"OpenAI API 에러: {result['error'].get('message', str(result['error']))}"}

    # 응답 파싱
    content = result["choices"][0]["message"]["content"]

    # JSON 추출 (```json ... ``` 블록이 있을 수 있음)
    if "```json" in content:
        content = content.split("```json")[1].split("```")[0]
    elif "```" in content:
        content = content.split("```")[1].split("```")[0]

    try:
        analysis = json.loads(content.strip())
    except json.JSONDecodeError:
        return {
            "error": "AI 응답을 파싱할 수 없습니다",
            "raw_response": content
        }

    return analysis


def _db_quality_score(match):
    """DB 항목의 데이터 품질을 0~5점으로 평가
    핵심 영양소(칼로리, 단백질, 탄수화물, 지방, serving_size)가 있으면 각 1점
    serving_size 없으면 비율 보정 불가 → 최대 3점까지만 (사실상 통과 불가)
    """
    score = 0
    has_serving = (match.get('serving_size_g') or 0) > 0
    if (match.get('calories_kcal') or 0) > 0: score += 1
    if (match.get('protein_g') or 0) > 0: score += 1
    if (match.get('carbs_g') or 0) > 0: score += 1
    if (match.get('fat_g') or 0) > 0: score += 1
    if has_serving: score += 1

    # serving_size 없으면 비율 보정이 불가능하므로 최대 2점으로 제한
    # → 품질 3점 기준을 통과하지 못해 AI 추정값 사용됨
    if not has_serving and score > 2:
        score = 2

    return score


# ── 핵심 음식 참조 테이블 (식약처 기준, 100g당) ──
# 우선순위 1: 이 테이블에서 먼저 매칭
CORE_FOODS = {
    # ── 과일 ──
    "귤": {"cal":39,"prot":0.7,"carbs":9.3,"fat":0.1,"fiber":0.5,"sodium":1,"sugar":7.8,"serving":100,"category":"fruit"},
    "감귤": {"cal":39,"prot":0.7,"carbs":9.3,"fat":0.1,"fiber":0.5,"sodium":1,"sugar":7.8,"serving":100,"category":"fruit"},
    "오렌지": {"cal":47,"prot":0.9,"carbs":11.8,"fat":0.1,"fiber":2.4,"sodium":0,"sugar":9.4,"serving":200,"category":"fruit"},
    "사과": {"cal":57,"prot":0.2,"carbs":14.1,"fat":0.2,"fiber":2.4,"sodium":1,"sugar":10.4,"serving":200,"category":"fruit"},
    "바나나": {"cal":93,"prot":1.0,"carbs":22.0,"fat":0.1,"fiber":2.6,"sodium":1,"sugar":12.2,"serving":120,"category":"fruit"},
    "포도": {"cal":60,"prot":0.6,"carbs":15.5,"fat":0.1,"fiber":0.9,"sodium":2,"sugar":15.5,"serving":100,"category":"fruit"},
    "딸기": {"cal":32,"prot":0.7,"carbs":7.7,"fat":0.3,"fiber":2.0,"sodium":1,"sugar":4.9,"serving":150,"category":"fruit"},
    "수박": {"cal":30,"prot":0.6,"carbs":7.6,"fat":0.2,"fiber":0.4,"sodium":1,"sugar":6.2,"serving":200,"category":"fruit"},
    "참외": {"cal":31,"prot":1.1,"carbs":6.8,"fat":0.1,"fiber":0.9,"sodium":3,"sugar":5.6,"serving":200,"category":"fruit"},
    "배": {"cal":51,"prot":0.3,"carbs":13.1,"fat":0.1,"fiber":2.1,"sodium":2,"sugar":8.5,"serving":200,"category":"fruit"},
    "복숭아": {"cal":39,"prot":0.9,"carbs":9.5,"fat":0.1,"fiber":1.5,"sodium":0,"sugar":8.4,"serving":200,"category":"fruit"},
    "감": {"cal":64,"prot":0.5,"carbs":15.9,"fat":0.2,"fiber":2.5,"sodium":1,"sugar":12.5,"serving":150,"category":"fruit"},
    "키위": {"cal":61,"prot":1.1,"carbs":14.7,"fat":0.5,"fiber":3.0,"sodium":3,"sugar":9.0,"serving":100,"category":"fruit"},
    "망고": {"cal":60,"prot":0.8,"carbs":15.0,"fat":0.4,"fiber":1.6,"sodium":1,"sugar":13.7,"serving":150,"category":"fruit"},
    "블루베리": {"cal":57,"prot":0.7,"carbs":14.5,"fat":0.3,"fiber":2.4,"sodium":1,"sugar":10.0,"serving":100,"category":"fruit"},
    "체리": {"cal":63,"prot":1.1,"carbs":16.0,"fat":0.2,"fiber":2.1,"sodium":0,"sugar":12.8,"serving":100,"category":"fruit"},
    "자두": {"cal":46,"prot":0.7,"carbs":11.4,"fat":0.3,"fiber":1.4,"sodium":0,"sugar":9.9,"serving":100,"category":"fruit"},
    "파인애플": {"cal":50,"prot":0.5,"carbs":13.1,"fat":0.1,"fiber":1.4,"sodium":1,"sugar":9.9,"serving":150,"category":"fruit"},
    "메론": {"cal":34,"prot":0.8,"carbs":8.2,"fat":0.2,"fiber":0.9,"sodium":16,"sugar":7.9,"serving":200,"category":"fruit"},
    "한라봉": {"cal":42,"prot":0.8,"carbs":10.0,"fat":0.1,"fiber":0.6,"sodium":1,"sugar":8.0,"serving":150,"category":"fruit"},

    # ── 밥/곡류 ──
    "밥": {"cal":150,"prot":2.7,"carbs":33.5,"fat":0.3,"fiber":0.4,"sodium":2,"sugar":0,"serving":210,"category":"korean"},
    "흰밥": {"cal":150,"prot":2.7,"carbs":33.5,"fat":0.3,"fiber":0.4,"sodium":2,"sugar":0,"serving":210,"category":"korean"},
    "현미밥": {"cal":149,"prot":3.2,"carbs":32.0,"fat":1.0,"fiber":1.8,"sodium":2,"sugar":0,"serving":210,"category":"korean"},
    "잡곡밥": {"cal":152,"prot":3.5,"carbs":32.5,"fat":0.8,"fiber":1.5,"sodium":2,"sugar":0,"serving":210,"category":"korean"},
    "볶음밥": {"cal":180,"prot":5.0,"carbs":28.0,"fat":5.5,"fiber":0.8,"sodium":350,"sugar":1.0,"serving":250,"category":"korean"},
    "김밥": {"cal":200,"prot":5.0,"carbs":30.0,"fat":6.5,"fiber":1.2,"sodium":450,"sugar":2.5,"serving":230,"category":"korean"},
    "비빔밥": {"cal":116,"prot":4.0,"carbs":17.3,"fat":3.1,"fiber":1.5,"sodium":300,"sugar":2.0,"serving":450,"category":"korean"},
    "죽": {"cal":46,"prot":1.0,"carbs":10.0,"fat":0.2,"fiber":0.2,"sodium":200,"sugar":0,"serving":300,"category":"korean"},

    # ── 국/찌개 ──
    "김치찌개": {"cal":50,"prot":3.5,"carbs":2.5,"fat":3.0,"fiber":0.5,"sodium":500,"sugar":1.0,"serving":400,"category":"korean"},
    "된장찌개": {"cal":38,"prot":2.5,"carbs":3.0,"fat":1.8,"fiber":0.8,"sodium":480,"sugar":0.5,"serving":400,"category":"korean"},
    "순두부찌개": {"cal":45,"prot":3.0,"carbs":2.5,"fat":2.5,"fiber":0.5,"sodium":450,"sugar":0.5,"serving":400,"category":"korean"},
    "부대찌개": {"cal":65,"prot":4.0,"carbs":4.0,"fat":3.5,"fiber":0.5,"sodium":550,"sugar":1.5,"serving":400,"category":"korean"},
    "미역국": {"cal":20,"prot":1.5,"carbs":1.5,"fat":1.0,"fiber":0.5,"sodium":400,"sugar":0,"serving":300,"category":"korean"},
    "갈비탕": {"cal":55,"prot":4.5,"carbs":1.0,"fat":3.5,"fiber":0,"sodium":400,"sugar":0,"serving":500,"category":"korean"},
    "설렁탕": {"cal":45,"prot":4.0,"carbs":0.5,"fat":2.5,"fiber":0,"sodium":350,"sugar":0,"serving":500,"category":"korean"},
    "떡국": {"cal":70,"prot":3.0,"carbs":11.0,"fat":1.5,"fiber":0.3,"sodium":400,"sugar":0.5,"serving":400,"category":"korean"},
    "삼계탕": {"cal":60,"prot":5.0,"carbs":3.0,"fat":3.0,"fiber":0.2,"sodium":300,"sugar":0.5,"serving":600,"category":"korean"},

    # ── 고기/구이 ──
    "삼겹살": {"cal":331,"prot":17.0,"carbs":0.0,"fat":29.0,"fiber":0,"sodium":60,"sugar":0,"serving":200,"category":"korean"},
    "목살": {"cal":220,"prot":18.0,"carbs":0.0,"fat":16.0,"fiber":0,"sodium":55,"sugar":0,"serving":200,"category":"korean"},
    "소고기": {"cal":250,"prot":20.0,"carbs":0.0,"fat":18.0,"fiber":0,"sodium":60,"sugar":0,"serving":150,"category":"korean"},
    "불고기": {"cal":157,"prot":16.0,"carbs":8.0,"fat":7.0,"fiber":0.5,"sodium":400,"sugar":5.0,"serving":200,"category":"korean"},
    "갈비": {"cal":280,"prot":18.0,"carbs":5.0,"fat":20.0,"fiber":0,"sodium":350,"sugar":3.0,"serving":200,"category":"korean"},
    "닭가슴살": {"cal":109,"prot":23.0,"carbs":0.0,"fat":1.2,"fiber":0,"sodium":65,"sugar":0,"serving":100,"category":"diet_fitness"},
    "닭다리": {"cal":190,"prot":16.0,"carbs":0.0,"fat":14.0,"fiber":0,"sodium":75,"sugar":0,"serving":100,"category":"korean"},
    "돼지갈비": {"cal":240,"prot":15.0,"carbs":6.0,"fat":17.0,"fiber":0,"sodium":400,"sugar":4.0,"serving":200,"category":"korean"},
    "제육볶음": {"cal":150,"prot":12.0,"carbs":8.0,"fat":8.0,"fiber":0.5,"sodium":450,"sugar":4.0,"serving":200,"category":"korean"},
    "오리고기": {"cal":200,"prot":16.0,"carbs":0.0,"fat":15.0,"fiber":0,"sodium":60,"sugar":0,"serving":150,"category":"korean"},

    # ── 해산물 ──
    "연어": {"cal":208,"prot":20.0,"carbs":0.0,"fat":13.0,"fiber":0,"sodium":60,"sugar":0,"serving":100,"category":"foreign_popular"},
    "참치": {"cal":132,"prot":28.0,"carbs":0.0,"fat":1.3,"fiber":0,"sodium":40,"sugar":0,"serving":100,"category":"korean"},
    "고등어": {"cal":202,"prot":19.0,"carbs":0.0,"fat":13.0,"fiber":0,"sodium":70,"sugar":0,"serving":100,"category":"korean"},
    "새우": {"cal":85,"prot":18.0,"carbs":0.8,"fat":0.7,"fiber":0,"sodium":150,"sugar":0,"serving":100,"category":"korean"},
    "오징어": {"cal":84,"prot":17.5,"carbs":0.8,"fat":1.0,"fiber":0,"sodium":200,"sugar":0,"serving":100,"category":"korean"},
    "광어": {"cal":110,"prot":21.0,"carbs":0.0,"fat":2.5,"fiber":0,"sodium":55,"sugar":0,"serving":100,"category":"korean"},
    "조개": {"cal":78,"prot":12.0,"carbs":3.5,"fat":1.5,"fiber":0,"sodium":250,"sugar":0,"serving":100,"category":"korean"},
    "회": {"cal":100,"prot":20.0,"carbs":0.5,"fat":1.5,"fiber":0,"sodium":50,"sugar":0,"serving":150,"category":"korean"},

    # ── 반찬/김치 ──
    "김치": {"cal":17,"prot":1.5,"carbs":2.4,"fat":0.5,"fiber":1.5,"sodium":600,"sugar":1.0,"serving":40,"category":"korean"},
    "배추김치": {"cal":17,"prot":1.5,"carbs":2.4,"fat":0.5,"fiber":1.5,"sodium":600,"sugar":1.0,"serving":40,"category":"korean"},
    "깍두기": {"cal":23,"prot":1.2,"carbs":4.2,"fat":0.3,"fiber":1.2,"sodium":550,"sugar":2.0,"serving":40,"category":"korean"},
    "계란후라이": {"cal":196,"prot":12.0,"carbs":0.8,"fat":15.0,"fiber":0,"sodium":180,"sugar":0.3,"serving":60,"category":"korean"},
    "계란말이": {"cal":160,"prot":11.0,"carbs":2.0,"fat":12.0,"fiber":0,"sodium":300,"sugar":1.0,"serving":80,"category":"korean"},
    "삶은계란": {"cal":155,"prot":13.0,"carbs":1.1,"fat":11.0,"fiber":0,"sodium":125,"sugar":1.1,"serving":50,"category":"korean"},
    "두부": {"cal":79,"prot":8.0,"carbs":1.5,"fat":4.5,"fiber":0.5,"sodium":10,"sugar":0.5,"serving":100,"category":"korean"},
    "콩나물": {"cal":30,"prot":3.5,"carbs":3.0,"fat":0.5,"fiber":1.5,"sodium":5,"sugar":0,"serving":70,"category":"korean"},
    "시금치": {"cal":23,"prot":2.9,"carbs":3.6,"fat":0.4,"fiber":2.2,"sodium":79,"sugar":0.4,"serving":70,"category":"korean"},
    "멸치볶음": {"cal":200,"prot":25.0,"carbs":10.0,"fat":7.0,"fiber":0.5,"sodium":800,"sugar":6.0,"serving":30,"category":"korean"},
    "잡채": {"cal":130,"prot":3.0,"carbs":20.0,"fat":4.5,"fiber":1.0,"sodium":350,"sugar":5.0,"serving":150,"category":"korean"},
    "나물": {"cal":35,"prot":2.5,"carbs":4.0,"fat":1.5,"fiber":2.0,"sodium":250,"sugar":1.0,"serving":70,"category":"korean"},

    # ── 면류 ──
    "라면": {"cal":375,"prot":8.3,"carbs":51.7,"fat":13.3,"fiber":2.0,"sodium":1500,"sugar":3.0,"serving":120,"category":"korean"},
    "신라면": {"cal":375,"prot":8.3,"carbs":51.7,"fat":13.3,"fiber":2.0,"sodium":1500,"sugar":3.0,"serving":120,"category":"franchise"},
    "짜파게티": {"cal":370,"prot":9.0,"carbs":52.0,"fat":13.0,"fiber":2.5,"sodium":1200,"sugar":4.0,"serving":140,"category":"franchise"},
    "잔치국수": {"cal":120,"prot":4.0,"carbs":20.0,"fat":2.0,"fiber":1.0,"sodium":500,"sugar":2.0,"serving":400,"category":"korean"},
    "칼국수": {"cal":110,"prot":4.5,"carbs":18.0,"fat":2.0,"fiber":0.8,"sodium":450,"sugar":1.5,"serving":450,"category":"korean"},
    "냉면": {"cal":115,"prot":4.0,"carbs":22.0,"fat":1.0,"fiber":1.0,"sodium":400,"sugar":3.0,"serving":500,"category":"korean"},
    "자장면": {"cal":130,"prot":4.0,"carbs":22.0,"fat":3.0,"fiber":1.0,"sodium":500,"sugar":4.0,"serving":450,"category":"korean"},
    "짬뽕": {"cal":75,"prot":4.5,"carbs":10.0,"fat":2.0,"fiber":1.0,"sodium":600,"sugar":2.0,"serving":500,"category":"korean"},
    "파스타": {"cal":160,"prot":6.0,"carbs":28.0,"fat":3.0,"fiber":1.5,"sodium":350,"sugar":2.5,"serving":300,"category":"foreign_popular"},
    "우동": {"cal":105,"prot":3.5,"carbs":20.0,"fat":1.0,"fiber":0.8,"sodium":500,"sugar":2.0,"serving":400,"category":"foreign_popular"},

    # ── 분식 ──
    "떡볶이": {"cal":193,"prot":3.5,"carbs":41.0,"fat":1.5,"fiber":0.8,"sodium":600,"sugar":8.0,"serving":300,"category":"korean"},
    "순대": {"cal":180,"prot":10.0,"carbs":18.0,"fat":7.0,"fiber":0.5,"sodium":400,"sugar":1.0,"serving":150,"category":"korean"},
    "튀김": {"cal":220,"prot":5.0,"carbs":25.0,"fat":11.0,"fiber":1.0,"sodium":300,"sugar":2.0,"serving":100,"category":"korean"},
    "만두": {"cal":200,"prot":8.0,"carbs":25.0,"fat":7.0,"fiber":1.0,"sodium":400,"sugar":2.0,"serving":100,"category":"korean"},
    "어묵": {"cal":95,"prot":8.0,"carbs":10.0,"fat":2.5,"fiber":0,"sodium":500,"sugar":2.0,"serving":100,"category":"korean"},
    "붕어빵": {"cal":260,"prot":6.0,"carbs":45.0,"fat":6.0,"fiber":1.0,"sodium":200,"sugar":15.0,"serving":80,"category":"snack_drink"},
    "호떡": {"cal":300,"prot":4.0,"carbs":50.0,"fat":10.0,"fiber":0.5,"sodium":200,"sugar":20.0,"serving":100,"category":"snack_drink"},

    # ── 외식/프랜차이즈 ──
    "치킨": {"cal":266,"prot":17.0,"carbs":10.0,"fat":18.0,"fiber":0.5,"sodium":500,"sugar":1.0,"serving":100,"category":"franchise"},
    "후라이드치킨": {"cal":266,"prot":17.0,"carbs":10.0,"fat":18.0,"fiber":0.5,"sodium":500,"sugar":1.0,"serving":100,"category":"franchise"},
    "양념치킨": {"cal":280,"prot":16.0,"carbs":15.0,"fat":17.0,"fiber":0.5,"sodium":550,"sugar":8.0,"serving":100,"category":"franchise"},
    "피자": {"cal":266,"prot":11.0,"carbs":33.0,"fat":10.0,"fiber":2.0,"sodium":600,"sugar":3.5,"serving":107,"category":"franchise"},
    "햄버거": {"cal":250,"prot":12.0,"carbs":28.0,"fat":10.0,"fiber":1.5,"sodium":500,"sugar":5.0,"serving":200,"category":"franchise"},
    "감자튀김": {"cal":312,"prot":3.4,"carbs":41.0,"fat":15.0,"fiber":3.8,"sodium":210,"sugar":0.3,"serving":115,"category":"franchise"},
    "핫도그": {"cal":290,"prot":10.0,"carbs":24.0,"fat":17.0,"fiber":1.0,"sodium":700,"sugar":4.0,"serving":100,"category":"franchise"},
    "돈까스": {"cal":230,"prot":14.0,"carbs":18.0,"fat":12.0,"fiber":1.0,"sodium":500,"sugar":2.0,"serving":200,"category":"foreign_popular"},
    "초밥": {"cal":150,"prot":6.0,"carbs":25.0,"fat":3.0,"fiber":0.3,"sodium":350,"sugar":4.0,"serving":200,"category":"foreign_popular"},
    "카레": {"cal":120,"prot":4.0,"carbs":16.0,"fat":5.0,"fiber":1.5,"sodium":400,"sugar":3.0,"serving":350,"category":"foreign_popular"},
    "짜장밥": {"cal":140,"prot":4.5,"carbs":22.0,"fat":4.0,"fiber":1.0,"sodium":500,"sugar":4.0,"serving":400,"category":"korean"},
    "덮밥": {"cal":130,"prot":5.0,"carbs":20.0,"fat":3.5,"fiber":1.0,"sodium":400,"sugar":3.0,"serving":400,"category":"korean"},
    "샐러드": {"cal":25,"prot":1.5,"carbs":3.5,"fat":0.5,"fiber":2.0,"sodium":50,"sugar":2.0,"serving":200,"category":"diet_fitness"},
    "샌드위치": {"cal":210,"prot":9.0,"carbs":25.0,"fat":8.0,"fiber":1.5,"sodium":450,"sugar":4.0,"serving":200,"category":"foreign_popular"},

    # ── 빵/디저트 ──
    "식빵": {"cal":265,"prot":9.0,"carbs":49.0,"fat":3.2,"fiber":2.7,"sodium":500,"sugar":5.0,"serving":30,"category":"snack_drink"},
    "크로와상": {"cal":406,"prot":8.0,"carbs":45.0,"fat":21.0,"fiber":2.3,"sodium":350,"sugar":6.0,"serving":60,"category":"snack_drink"},
    "케이크": {"cal":300,"prot":5.0,"carbs":38.0,"fat":15.0,"fiber":0.5,"sodium":200,"sugar":25.0,"serving":100,"category":"snack_drink"},
    "도넛": {"cal":400,"prot":5.0,"carbs":50.0,"fat":20.0,"fiber":1.0,"sodium":350,"sugar":20.0,"serving":75,"category":"snack_drink"},
    "와플": {"cal":310,"prot":6.0,"carbs":40.0,"fat":14.0,"fiber":1.0,"sodium":400,"sugar":12.0,"serving":100,"category":"snack_drink"},
    "떡": {"cal":230,"prot":4.0,"carbs":52.0,"fat":0.5,"fiber":0.5,"sodium":150,"sugar":8.0,"serving":100,"category":"korean"},
    "아이스크림": {"cal":207,"prot":3.5,"carbs":24.0,"fat":11.0,"fiber":0.5,"sodium":80,"sugar":21.0,"serving":100,"category":"snack_drink"},
    "초콜릿": {"cal":546,"prot":5.0,"carbs":60.0,"fat":31.0,"fiber":3.4,"sodium":24,"sugar":48.0,"serving":40,"category":"snack_drink"},

    # ── 음료 ──
    "물": {"cal":0,"prot":0.0,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":0,"sugar":0,"serving":250,"category":"snack_drink"},
    "생수": {"cal":0,"prot":0.0,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":0,"sugar":0,"serving":500,"category":"snack_drink"},
    "탄산수": {"cal":0,"prot":0.0,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":5,"sugar":0,"serving":350,"category":"snack_drink"},
    "얼음물": {"cal":0,"prot":0.0,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":0,"sugar":0,"serving":250,"category":"snack_drink"},
    "보리차": {"cal":1,"prot":0.0,"carbs":0.2,"fat":0.0,"fiber":0,"sodium":2,"sugar":0,"serving":200,"category":"snack_drink"},
    "우유": {"cal":65,"prot":3.2,"carbs":4.7,"fat":3.5,"fiber":0,"sodium":40,"sugar":4.7,"serving":200,"category":"snack_drink"},
    "저지방우유": {"cal":42,"prot":3.4,"carbs":5.0,"fat":1.5,"fiber":0,"sodium":45,"sugar":5.0,"serving":200,"category":"snack_drink"},
    "두유": {"cal":50,"prot":3.5,"carbs":5.0,"fat":1.8,"fiber":0.5,"sodium":40,"sugar":3.5,"serving":200,"category":"snack_drink"},
    "콜라": {"cal":39,"prot":0.0,"carbs":10.6,"fat":0.0,"fiber":0,"sodium":5,"sugar":10.6,"serving":250,"category":"snack_drink"},
    "제로콜라": {"cal":0,"prot":0.0,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":5,"sugar":0,"serving":250,"category":"snack_drink"},
    "사이다": {"cal":40,"prot":0.0,"carbs":10.0,"fat":0.0,"fiber":0,"sodium":10,"sugar":10.0,"serving":250,"category":"snack_drink"},
    "커피": {"cal":2,"prot":0.3,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":2,"sugar":0,"serving":200,"category":"snack_drink"},
    "아메리카노": {"cal":5,"prot":0.3,"carbs":0.5,"fat":0.0,"fiber":0,"sodium":2,"sugar":0,"serving":350,"category":"snack_drink"},
    "카페라떼": {"cal":50,"prot":3.0,"carbs":4.5,"fat":2.5,"fiber":0,"sodium":50,"sugar":4.5,"serving":350,"category":"snack_drink"},
    "카푸치노": {"cal":45,"prot":2.5,"carbs":4.0,"fat":2.0,"fiber":0,"sodium":45,"sugar":4.0,"serving":300,"category":"snack_drink"},
    "카페모카": {"cal":120,"prot":3.0,"carbs":18.0,"fat":4.5,"fiber":0.5,"sodium":55,"sugar":15.0,"serving":350,"category":"snack_drink"},
    "바닐라라떼": {"cal":100,"prot":3.0,"carbs":16.0,"fat":3.0,"fiber":0,"sodium":50,"sugar":14.0,"serving":350,"category":"snack_drink"},
    "녹차": {"cal":1,"prot":0.2,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":1,"sugar":0,"serving":200,"category":"snack_drink"},
    "녹차라떼": {"cal":80,"prot":2.5,"carbs":12.0,"fat":2.5,"fiber":0.3,"sodium":40,"sugar":10.0,"serving":350,"category":"snack_drink"},
    "홍차": {"cal":1,"prot":0.0,"carbs":0.3,"fat":0.0,"fiber":0,"sodium":3,"sugar":0,"serving":200,"category":"snack_drink"},
    "밀크티": {"cal":80,"prot":2.0,"carbs":14.0,"fat":2.0,"fiber":0,"sodium":40,"sugar":12.0,"serving":350,"category":"snack_drink"},
    "스무디": {"cal":60,"prot":1.0,"carbs":14.0,"fat":0.3,"fiber":1.0,"sodium":5,"sugar":12.0,"serving":350,"category":"snack_drink"},
    "오렌지주스": {"cal":45,"prot":0.7,"carbs":10.5,"fat":0.2,"fiber":0.2,"sodium":2,"sugar":8.5,"serving":200,"category":"snack_drink"},
    "맥주": {"cal":43,"prot":0.5,"carbs":3.6,"fat":0.0,"fiber":0,"sodium":4,"sugar":0,"serving":500,"category":"snack_drink"},
    "와인": {"cal":83,"prot":0.1,"carbs":2.6,"fat":0.0,"fiber":0,"sodium":5,"sugar":0.6,"serving":150,"category":"snack_drink"},
    "소주": {"cal":87,"prot":0.0,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":0,"sugar":0,"serving":360,"category":"snack_drink"},
    "막걸리": {"cal":56,"prot":0.8,"carbs":5.0,"fat":0.1,"fiber":0,"sodium":5,"sugar":3.5,"serving":300,"category":"snack_drink"},
    "하이볼": {"cal":70,"prot":0.0,"carbs":5.0,"fat":0.0,"fiber":0,"sodium":5,"sugar":4.0,"serving":350,"category":"snack_drink"},

    # ── 간식 ──
    "라면땅": {"cal":450,"prot":8.0,"carbs":55.0,"fat":22.0,"fiber":1.0,"sodium":500,"sugar":5.0,"serving":90,"category":"snack_drink"},
    "과자": {"cal":480,"prot":5.0,"carbs":62.0,"fat":24.0,"fiber":1.0,"sodium":400,"sugar":10.0,"serving":100,"category":"snack_drink"},
    "팝콘": {"cal":375,"prot":11.0,"carbs":74.0,"fat":4.5,"fiber":15.0,"sodium":5,"sugar":1.0,"serving":50,"category":"snack_drink"},
    "견과류": {"cal":607,"prot":20.0,"carbs":21.0,"fat":54.0,"fiber":7.0,"sodium":5,"sugar":4.0,"serving":30,"category":"snack_drink"},
    "아몬드": {"cal":579,"prot":21.0,"carbs":22.0,"fat":50.0,"fiber":12.0,"sodium":1,"sugar":4.0,"serving":30,"category":"snack_drink"},
    "고구마": {"cal":111,"prot":1.5,"carbs":26.0,"fat":0.1,"fiber":3.0,"sodium":10,"sugar":5.5,"serving":150,"category":"korean"},
    "감자": {"cal":77,"prot":2.0,"carbs":17.5,"fat":0.1,"fiber":2.2,"sodium":6,"sugar":0.8,"serving":150,"category":"korean"},
    "감자탕": {"cal":55,"prot":4.0,"carbs":3.0,"fat":3.0,"fiber":0.5,"sodium":500,"sugar":1.0,"serving":500,"category":"korean"},
    "김치볶음밥": {"cal":170,"prot":5.0,"carbs":26.0,"fat":5.0,"fiber":1.0,"sodium":450,"sugar":2.0,"serving":300,"category":"korean"},
    "순대국": {"cal":60,"prot":5.0,"carbs":3.0,"fat":3.5,"fiber":0.3,"sodium":500,"sugar":0.5,"serving":500,"category":"korean"},
    "닭볶음탕": {"cal":90,"prot":8.0,"carbs":5.0,"fat":4.0,"fiber":0.5,"sodium":450,"sugar":3.0,"serving":400,"category":"korean"},
    "보쌈": {"cal":180,"prot":18.0,"carbs":2.0,"fat":11.0,"fiber":0,"sodium":350,"sugar":1.0,"serving":200,"category":"korean"},
    "족발": {"cal":220,"prot":22.0,"carbs":0.5,"fat":14.0,"fiber":0,"sodium":300,"sugar":0.5,"serving":200,"category":"korean"},
    "해물파전": {"cal":160,"prot":6.0,"carbs":20.0,"fat":6.0,"fiber":1.0,"sodium":400,"sugar":2.0,"serving":200,"category":"korean"},
    "빈대떡": {"cal":200,"prot":8.0,"carbs":20.0,"fat":10.0,"fiber":1.5,"sodium":350,"sugar":1.0,"serving":150,"category":"korean"},
    "옥수수": {"cal":96,"prot":3.2,"carbs":19.0,"fat":1.2,"fiber":2.7,"sodium":15,"sugar":3.2,"serving":150,"category":"korean"},
    "양파": {"cal":40,"prot":1.1,"carbs":9.3,"fat":0.1,"fiber":1.7,"sodium":4,"sugar":4.2,"serving":150,"category":"korean"},
    "튀긴양파": {"cal":300,"prot":3.5,"carbs":30.0,"fat":19.0,"fiber":2.0,"sodium":200,"sugar":5.0,"serving":30,"category":"korean"},
    "튀긴 양파": {"cal":300,"prot":3.5,"carbs":30.0,"fat":19.0,"fiber":2.0,"sodium":200,"sugar":5.0,"serving":30,"category":"korean"},
    "마늘": {"cal":149,"prot":6.4,"carbs":33.1,"fat":0.5,"fiber":2.1,"sodium":17,"sugar":1.0,"serving":10,"category":"korean"},
    "당근": {"cal":41,"prot":0.9,"carbs":9.6,"fat":0.2,"fiber":2.8,"sodium":69,"sugar":4.7,"serving":100,"category":"korean"},
    "오이": {"cal":15,"prot":0.7,"carbs":3.6,"fat":0.1,"fiber":0.5,"sodium":2,"sugar":1.7,"serving":100,"category":"korean"},
    "토마토": {"cal":18,"prot":0.9,"carbs":3.9,"fat":0.2,"fiber":1.2,"sodium":5,"sugar":2.6,"serving":150,"category":"korean"},
    "버섯": {"cal":22,"prot":3.1,"carbs":3.3,"fat":0.3,"fiber":1.0,"sodium":5,"sugar":2.0,"serving":70,"category":"korean"},
    "브로콜리": {"cal":34,"prot":2.8,"carbs":7.0,"fat":0.4,"fiber":2.6,"sodium":33,"sugar":1.7,"serving":100,"category":"diet_fitness"},
    "파프리카": {"cal":31,"prot":1.0,"carbs":6.0,"fat":0.3,"fiber":2.1,"sodium":4,"sugar":4.2,"serving":100,"category":"korean"},
    "상추": {"cal":15,"prot":1.4,"carbs":2.9,"fat":0.2,"fiber":1.3,"sodium":28,"sugar":0.8,"serving":30,"category":"korean"},
    "토스트": {"cal":280,"prot":8.0,"carbs":35.0,"fat":12.0,"fiber":1.5,"sodium":400,"sugar":5.0,"serving":150,"category":"snack_drink"},
    "떡갈비": {"cal":220,"prot":12.0,"carbs":12.0,"fat":14.0,"fiber":0.5,"sodium":400,"sugar":3.0,"serving":150,"category":"korean"},
    "탕수육": {"cal":240,"prot":12.0,"carbs":25.0,"fat":10.0,"fiber":0.5,"sodium":350,"sugar":10.0,"serving":200,"category":"korean"},
    "갈비찜": {"cal":168,"prot":11.2,"carbs":8.0,"fat":9.6,"fiber":0.6,"sodium":392,"sugar":4.0,"serving":250,"category":"korean"},
    "간장게장": {"cal":80,"prot":12.0,"carbs":3.3,"fat":3.3,"fiber":0,"sodium":500,"sugar":1.0,"serving":150,"category":"korean"},
    "곰탕": {"cal":56,"prot":3.6,"carbs":1.6,"fat":3.6,"fiber":0,"sodium":300,"sugar":0,"serving":500,"category":"korean"},
    "곱창구이": {"cal":187,"prot":10.7,"carbs":1.3,"fat":14.7,"fiber":0,"sodium":300,"sugar":0,"serving":150,"category":"korean"},
    "계란찜": {"cal":73,"prot":6.0,"carbs":1.3,"fat":5.0,"fiber":0,"sodium":250,"sugar":0.5,"serving":150,"category":"korean"},
    "닭갈비": {"cal":128,"prot":10.0,"carbs":8.0,"fat":6.0,"fiber":0.5,"sodium":400,"sugar":4.0,"serving":250,"category":"korean"},
    "비빔냉면": {"cal":130,"prot":4.5,"carbs":25.0,"fat":1.5,"fiber":1.5,"sodium":450,"sugar":5.0,"serving":500,"category":"korean"},
    "스테이크": {"cal":271,"prot":26.0,"carbs":0.0,"fat":18.0,"fiber":0,"sodium":65,"sugar":0,"serving":200,"category":"foreign_popular"},

    # ── 양식/브런치 (자주 인식되는 항목 보강) ──
    "까르보나라": {"cal":180,"prot":8.0,"carbs":22.0,"fat":7.0,"fiber":1.0,"sodium":400,"sugar":1.0,"serving":300,"category":"foreign_popular"},
    "까르보나라 파스타": {"cal":180,"prot":8.0,"carbs":22.0,"fat":7.0,"fiber":1.0,"sodium":400,"sugar":1.0,"serving":300,"category":"foreign_popular"},
    "크림 파스타": {"cal":170,"prot":6.5,"carbs":22.0,"fat":6.5,"fiber":1.0,"sodium":380,"sugar":1.5,"serving":300,"category":"foreign_popular"},
    "토마토 파스타": {"cal":130,"prot":5.0,"carbs":22.0,"fat":3.0,"fiber":2.0,"sodium":400,"sugar":4.0,"serving":300,"category":"foreign_popular"},
    "토마토 소스 파스타": {"cal":130,"prot":5.0,"carbs":22.0,"fat":3.0,"fiber":2.0,"sodium":400,"sugar":4.0,"serving":300,"category":"foreign_popular"},
    "알리오올리오": {"cal":160,"prot":5.5,"carbs":24.0,"fat":5.0,"fiber":1.2,"sodium":350,"sugar":1.0,"serving":300,"category":"foreign_popular"},
    "봉골레 파스타": {"cal":140,"prot":7.0,"carbs":21.0,"fat":3.5,"fiber":1.0,"sodium":420,"sugar":1.0,"serving":300,"category":"foreign_popular"},
    "리조또": {"cal":140,"prot":4.5,"carbs":20.0,"fat":5.0,"fiber":0.5,"sodium":350,"sugar":1.0,"serving":300,"category":"foreign_popular"},
    "리코타 치즈 샐러드": {"cal":75,"prot":5.0,"carbs":6.0,"fat":4.0,"fiber":2.0,"sodium":150,"sugar":3.0,"serving":250,"category":"foreign_popular"},
    "시저 샐러드": {"cal":80,"prot":4.0,"carbs":5.0,"fat":5.5,"fiber":1.5,"sodium":300,"sugar":1.5,"serving":200,"category":"foreign_popular"},
    "닭가슴살 샐러드": {"cal":60,"prot":7.0,"carbs":4.0,"fat":2.0,"fiber":2.0,"sodium":200,"sugar":2.0,"serving":250,"category":"diet_fitness"},
    "연어 샐러드": {"cal":70,"prot":6.5,"carbs":4.0,"fat":3.5,"fiber":1.5,"sodium":180,"sugar":2.0,"serving":250,"category":"foreign_popular"},
    "에그 베네딕트": {"cal":180,"prot":10.0,"carbs":15.0,"fat":9.0,"fiber":0.5,"sodium":500,"sugar":1.5,"serving":250,"category":"foreign_popular"},
    "팬케이크": {"cal":227,"prot":6.0,"carbs":35.0,"fat":7.0,"fiber":1.0,"sodium":350,"sugar":10.0,"serving":150,"category":"foreign_popular"},
    "프렌치토스트": {"cal":220,"prot":7.0,"carbs":28.0,"fat":9.0,"fiber":1.0,"sodium":300,"sugar":8.0,"serving":150,"category":"foreign_popular"},
    "오므라이스": {"cal":140,"prot":5.5,"carbs":18.0,"fat":5.0,"fiber":0.5,"sodium":400,"sugar":3.0,"serving":350,"category":"foreign_popular"},
    "그라탕": {"cal":140,"prot":6.0,"carbs":12.0,"fat":7.5,"fiber":0.5,"sodium":400,"sugar":2.0,"serving":250,"category":"foreign_popular"},
    "빵": {"cal":265,"prot":9.0,"carbs":49.0,"fat":3.2,"fiber":2.7,"sodium":500,"sugar":5.0,"serving":60,"category":"snack_drink"},
    "바게트": {"cal":274,"prot":9.0,"carbs":56.0,"fat":1.0,"fiber":2.4,"sodium":600,"sugar":3.0,"serving":60,"category":"snack_drink"},
    "치아바타": {"cal":271,"prot":9.0,"carbs":50.0,"fat":3.5,"fiber":2.0,"sodium":500,"sugar":2.0,"serving":80,"category":"snack_drink"},
    "마카롱": {"cal":400,"prot":5.0,"carbs":60.0,"fat":16.0,"fiber":1.0,"sodium":50,"sugar":50.0,"serving":30,"category":"snack_drink"},
    "타르트": {"cal":320,"prot":4.0,"carbs":40.0,"fat":16.0,"fiber":1.0,"sodium":200,"sugar":20.0,"serving":100,"category":"snack_drink"},
    "머핀": {"cal":340,"prot":5.0,"carbs":48.0,"fat":14.0,"fiber":1.5,"sodium":350,"sugar":22.0,"serving":100,"category":"snack_drink"},
    "브라우니": {"cal":400,"prot":5.0,"carbs":50.0,"fat":20.0,"fiber":2.0,"sodium":200,"sugar":30.0,"serving":60,"category":"snack_drink"},
    "수프": {"cal":40,"prot":1.5,"carbs":5.0,"fat":1.5,"fiber":0.5,"sodium":400,"sugar":1.5,"serving":250,"category":"foreign_popular"},
    "클램차우더": {"cal":60,"prot":2.5,"carbs":7.0,"fat":2.5,"fiber":0.5,"sodium":450,"sugar":1.0,"serving":250,"category":"foreign_popular"},

    # ── 추가: 한식 밥류/덮밥 ──
    "누드김밥": {"cal":180,"prot":5.0,"carbs":25.0,"fat":6.0,"fiber":1.5,"sodium":400,"sugar":2.0,"serving":230,"category":"korean"},
    "제육덮밥": {"cal":150,"prot":8.0,"carbs":20.0,"fat":5.0,"fiber":0.5,"sodium":450,"sugar":3.0,"serving":400,"category":"korean"},

    # ── 추가: 국/찌개/탕 ──
    "시래기국": {"cal":20,"prot":1.5,"carbs":2.5,"fat":0.8,"fiber":1.5,"sodium":400,"sugar":0.5,"serving":300,"category":"korean"},
    "선지국": {"cal":35,"prot":4.0,"carbs":1.5,"fat":1.5,"fiber":0,"sodium":400,"sugar":0,"serving":500,"category":"korean"},
    "떡만둣국": {"cal":60,"prot":3.5,"carbs":9.0,"fat":1.5,"fiber":0.3,"sodium":400,"sugar":0.5,"serving":500,"category":"korean"},
    "오뎅탕": {"cal":25,"prot":2.5,"carbs":2.0,"fat":0.8,"fiber":0,"sodium":400,"sugar":1.0,"serving":400,"category":"korean"},

    # ── 추가: 고기/구이 ──
    "항정살": {"cal":290,"prot":15.0,"carbs":0.0,"fat":25.0,"fiber":0,"sodium":55,"sugar":0,"serving":150,"category":"korean"},
    "가브리살": {"cal":260,"prot":16.0,"carbs":0.0,"fat":22.0,"fiber":0,"sodium":55,"sugar":0,"serving":150,"category":"korean"},
    "LA갈비": {"cal":280,"prot":17.0,"carbs":6.0,"fat":21.0,"fiber":0,"sodium":350,"sugar":4.0,"serving":200,"category":"korean"},
    "오리훈제": {"cal":190,"prot":16.0,"carbs":2.0,"fat":13.0,"fiber":0,"sodium":350,"sugar":1.0,"serving":150,"category":"korean"},
    "양고기": {"cal":230,"prot":18.0,"carbs":0.0,"fat":17.0,"fiber":0,"sodium":60,"sugar":0,"serving":150,"category":"foreign_popular"},

    # ── 추가: 반찬 ──
    "진미채볶음": {"cal":280,"prot":12.0,"carbs":35.0,"fat":10.0,"fiber":0.5,"sodium":600,"sugar":15.0,"serving":30,"category":"korean"},

    # ── 추가: 분식 ──
    "떡꼬치": {"cal":200,"prot":3.0,"carbs":40.0,"fat":3.0,"fiber":0.5,"sodium":350,"sugar":8.0,"serving":100,"category":"korean"},
    "핫바": {"cal":180,"prot":7.0,"carbs":18.0,"fat":8.0,"fiber":0.5,"sodium":500,"sugar":3.0,"serving":100,"category":"korean"},
    "소떡소떡": {"cal":200,"prot":5.0,"carbs":35.0,"fat":4.0,"fiber":0.5,"sodium":400,"sugar":5.0,"serving":120,"category":"korean"},
    "왕만두": {"cal":200,"prot":8.0,"carbs":25.0,"fat":7.0,"fiber":1.0,"sodium":400,"sugar":2.0,"serving":150,"category":"korean"},

    # ── 추가: 면류/라면 ──
    "컵라면": {"cal":300,"prot":6.0,"carbs":42.0,"fat":12.0,"fiber":1.5,"sodium":1300,"sugar":3.0,"serving":65,"category":"korean"},
    "진라면": {"cal":375,"prot":8.5,"carbs":52.0,"fat":13.5,"fiber":2.0,"sodium":1500,"sugar":3.0,"serving":120,"category":"franchise"},
    "불닭볶음면": {"cal":378,"prot":6.4,"carbs":57.1,"fat":11.4,"fiber":1.4,"sodium":1286,"sugar":4.3,"serving":140,"category":"franchise"},
    "밀면": {"cal":100,"prot":3.5,"carbs":20.0,"fat":0.5,"fiber":0.8,"sodium":400,"sugar":2.0,"serving":500,"category":"korean"},
    "볶음우동": {"cal":130,"prot":4.0,"carbs":22.0,"fat":3.0,"fiber":1.0,"sodium":500,"sugar":3.0,"serving":350,"category":"foreign_popular"},
    "소바": {"cal":110,"prot":5.0,"carbs":20.0,"fat":1.0,"fiber":1.5,"sodium":400,"sugar":2.0,"serving":350,"category":"foreign_popular"},
    "냉소바": {"cal":95,"prot":5.0,"carbs":18.0,"fat":0.5,"fiber":1.5,"sodium":350,"sugar":3.0,"serving":350,"category":"foreign_popular"},
    "잔뽕": {"cal":80,"prot":5.0,"carbs":10.0,"fat":2.5,"fiber":1.0,"sodium":550,"sugar":2.0,"serving":500,"category":"korean"},

    # ── 추가: 양식 파스타 ──
    "뇨끼": {"cal":130,"prot":4.0,"carbs":22.0,"fat":3.0,"fiber":1.0,"sodium":300,"sugar":1.0,"serving":300,"category":"foreign_popular"},
    "라자냐": {"cal":135,"prot":7.0,"carbs":14.0,"fat":6.0,"fiber":1.0,"sodium":400,"sugar":2.0,"serving":350,"category":"foreign_popular"},
    "라비올리": {"cal":120,"prot":5.5,"carbs":18.0,"fat":3.0,"fiber":1.0,"sodium":350,"sugar":1.0,"serving":300,"category":"foreign_popular"},
    "펜네": {"cal":155,"prot":5.5,"carbs":27.0,"fat":2.5,"fiber":1.5,"sodium":350,"sugar":2.0,"serving":300,"category":"foreign_popular"},
    "링귀네": {"cal":155,"prot":5.5,"carbs":27.0,"fat":2.5,"fiber":1.5,"sodium":350,"sugar":2.0,"serving":300,"category":"foreign_popular"},
    "오레키에떼": {"cal":145,"prot":5.0,"carbs":25.0,"fat":3.0,"fiber":1.0,"sodium":350,"sugar":1.5,"serving":300,"category":"foreign_popular"},
    "카넬로니": {"cal":120,"prot":5.5,"carbs":14.0,"fat":4.5,"fiber":1.0,"sodium":380,"sugar":2.0,"serving":300,"category":"foreign_popular"},
    "스파게티": {"cal":155,"prot":5.5,"carbs":27.0,"fat":2.5,"fiber":1.5,"sodium":350,"sugar":2.0,"serving":300,"category":"foreign_popular"},

    # ── 추가: 양식 메인 ──
    "치즈돈까스": {"cal":260,"prot":15.0,"carbs":20.0,"fat":14.0,"fiber":1.0,"sodium":550,"sugar":2.0,"serving":220,"category":"foreign_popular"},
    "필렛미뇽": {"cal":200,"prot":26.0,"carbs":0.0,"fat":10.0,"fiber":0,"sodium":55,"sugar":0,"serving":200,"category":"foreign_popular"},
    "비프스튜": {"cal":80,"prot":5.5,"carbs":6.0,"fat":3.5,"fiber":1.0,"sodium":350,"sugar":2.0,"serving":400,"category":"foreign_popular"},
    "피쉬앤칩스": {"cal":220,"prot":10.0,"carbs":22.0,"fat":10.0,"fiber":1.5,"sodium":400,"sugar":1.0,"serving":250,"category":"foreign_popular"},
    "미트로프": {"cal":160,"prot":10.0,"carbs":8.0,"fat":10.0,"fiber":0.5,"sodium":400,"sugar":2.0,"serving":250,"category":"foreign_popular"},
    "미트파이": {"cal":240,"prot":8.0,"carbs":22.0,"fat":13.0,"fiber":1.0,"sodium":450,"sugar":2.0,"serving":200,"category":"foreign_popular"},
    "바비큐립": {"cal":250,"prot":18.0,"carbs":8.0,"fat":16.0,"fiber":0,"sodium":400,"sugar":5.0,"serving":200,"category":"foreign_popular"},
    "비프부르기뇽": {"cal":90,"prot":6.0,"carbs":5.0,"fat":5.0,"fiber":1.0,"sodium":350,"sugar":2.0,"serving":400,"category":"foreign_popular"},

    # ── 추가: 샐러드/수프 ──
    "카프레제": {"cal":70,"prot":4.0,"carbs":3.0,"fat":5.0,"fiber":0.5,"sodium":150,"sugar":2.0,"serving":250,"category":"foreign_popular"},
    "양송이수프": {"cal":40,"prot":1.5,"carbs":4.0,"fat":2.0,"fiber":0.5,"sodium":380,"sugar":1.0,"serving":250,"category":"foreign_popular"},
    "크림수프": {"cal":55,"prot":1.5,"carbs":6.0,"fat":3.0,"fiber":0.5,"sodium":400,"sugar":1.5,"serving":250,"category":"foreign_popular"},
    "미네스트로네": {"cal":35,"prot":1.5,"carbs":5.5,"fat":0.8,"fiber":1.5,"sodium":350,"sugar":2.0,"serving":300,"category":"foreign_popular"},
    "호박수프": {"cal":35,"prot":1.0,"carbs":6.0,"fat":1.0,"fiber":1.0,"sodium":350,"sugar":2.5,"serving":300,"category":"foreign_popular"},
    "비시수아즈": {"cal":50,"prot":1.5,"carbs":6.0,"fat":2.5,"fiber":0.5,"sodium":350,"sugar":1.5,"serving":250,"category":"foreign_popular"},
    "토마토수프": {"cal":35,"prot":1.0,"carbs":5.5,"fat":1.0,"fiber":1.0,"sodium":400,"sugar":3.0,"serving":300,"category":"foreign_popular"},
    "옥수수수프": {"cal":50,"prot":1.5,"carbs":8.0,"fat":1.5,"fiber":0.5,"sodium":380,"sugar":2.0,"serving":250,"category":"foreign_popular"},
    "감자수프": {"cal":45,"prot":1.0,"carbs":7.0,"fat":1.5,"fiber":0.5,"sodium":380,"sugar":1.0,"serving":300,"category":"foreign_popular"},
    "그릭 샐러드": {"cal":60,"prot":3.0,"carbs":5.0,"fat":3.5,"fiber":1.5,"sodium":250,"sugar":2.5,"serving":250,"category":"foreign_popular"},
    "코브 샐러드": {"cal":80,"prot":6.0,"carbs":4.0,"fat":5.0,"fiber":2.0,"sodium":300,"sugar":2.0,"serving":300,"category":"foreign_popular"},
    "월도프 샐러드": {"cal":70,"prot":1.5,"carbs":8.0,"fat":4.0,"fiber":1.5,"sodium":150,"sugar":5.0,"serving":250,"category":"foreign_popular"},

    # ── 추가: 패스트푸드/치킨 ──
    "간장치킨": {"cal":250,"prot":16.0,"carbs":12.0,"fat":16.0,"fiber":0.5,"sodium":550,"sugar":6.0,"serving":100,"category":"franchise"},
    "허니버터치킨": {"cal":280,"prot":15.0,"carbs":16.0,"fat":17.0,"fiber":0.5,"sodium":500,"sugar":10.0,"serving":100,"category":"franchise"},
    "마늘치킨": {"cal":260,"prot":16.0,"carbs":12.0,"fat":17.0,"fiber":0.5,"sodium":500,"sugar":4.0,"serving":100,"category":"franchise"},
    "하와이안피자": {"cal":250,"prot":10.0,"carbs":32.0,"fat":9.0,"fiber":2.0,"sodium":580,"sugar":5.0,"serving":107,"category":"franchise"},
    "불고기피자": {"cal":260,"prot":11.0,"carbs":31.0,"fat":10.0,"fiber":2.0,"sodium":600,"sugar":4.0,"serving":107,"category":"franchise"},
    "너겟": {"cal":280,"prot":14.0,"carbs":18.0,"fat":16.0,"fiber":1.0,"sodium":500,"sugar":1.0,"serving":100,"category":"franchise"},
    "치킨너겟": {"cal":280,"prot":14.0,"carbs":18.0,"fat":16.0,"fiber":1.0,"sodium":500,"sugar":1.0,"serving":100,"category":"franchise"},
    "타코": {"cal":170,"prot":8.0,"carbs":15.0,"fat":8.0,"fiber":2.0,"sodium":400,"sugar":2.0,"serving":150,"category":"foreign_popular"},
    "부리토": {"cal":150,"prot":7.0,"carbs":18.0,"fat":5.5,"fiber":2.5,"sodium":450,"sugar":2.0,"serving":350,"category":"foreign_popular"},
    "퀘사디아": {"cal":220,"prot":10.0,"carbs":20.0,"fat":11.0,"fiber":1.5,"sodium":450,"sugar":1.5,"serving":200,"category":"foreign_popular"},
    "나초": {"cal":350,"prot":5.0,"carbs":40.0,"fat":19.0,"fiber":3.0,"sodium":400,"sugar":2.0,"serving":100,"category":"foreign_popular"},
    "오니언링": {"cal":280,"prot":3.5,"carbs":30.0,"fat":16.0,"fiber":1.5,"sodium":350,"sugar":3.0,"serving":100,"category":"foreign_popular"},
    "콘도그": {"cal":260,"prot":8.0,"carbs":28.0,"fat":13.0,"fiber":1.0,"sodium":500,"sugar":5.0,"serving":100,"category":"franchise"},
    "피시버거": {"cal":240,"prot":12.0,"carbs":28.0,"fat":10.0,"fiber":1.5,"sodium":550,"sugar":4.0,"serving":200,"category":"franchise"},

    # ── 추가: 일식 ──
    "사시미": {"cal":100,"prot":20.0,"carbs":0.5,"fat":1.5,"fiber":0,"sodium":50,"sugar":0,"serving":150,"category":"foreign_popular"},
    "규동": {"cal":130,"prot":6.0,"carbs":18.0,"fat":3.5,"fiber":0.5,"sodium":400,"sugar":4.0,"serving":400,"category":"foreign_popular"},
    "텐동": {"cal":150,"prot":6.0,"carbs":20.0,"fat":5.0,"fiber":1.0,"sodium":450,"sugar":3.0,"serving":400,"category":"foreign_popular"},
    "오야코동": {"cal":120,"prot":7.0,"carbs":16.0,"fat":3.0,"fiber":0.3,"sodium":400,"sugar":3.0,"serving":400,"category":"foreign_popular"},
    "미소시루": {"cal":20,"prot":1.5,"carbs":2.0,"fat":0.5,"fiber":0.3,"sodium":350,"sugar":0.5,"serving":200,"category":"foreign_popular"},
    "야키소바": {"cal":140,"prot":4.5,"carbs":22.0,"fat":4.0,"fiber":1.0,"sodium":500,"sugar":3.0,"serving":350,"category":"foreign_popular"},
    "야키토리": {"cal":180,"prot":16.0,"carbs":5.0,"fat":10.0,"fiber":0,"sodium":400,"sugar":3.0,"serving":120,"category":"foreign_popular"},
    "돈부리": {"cal":130,"prot":6.0,"carbs":18.0,"fat":3.5,"fiber":0.5,"sodium":400,"sugar":3.0,"serving":400,"category":"foreign_popular"},
    "라멘": {"cal":105,"prot":5.0,"carbs":15.0,"fat":3.0,"fiber":0.5,"sodium":600,"sugar":2.0,"serving":500,"category":"foreign_popular"},
    "츠케멘": {"cal":120,"prot":5.5,"carbs":18.0,"fat":3.0,"fiber":0.5,"sodium":550,"sugar":2.0,"serving":450,"category":"foreign_popular"},
    "규카츠": {"cal":230,"prot":18.0,"carbs":12.0,"fat":12.0,"fiber":0.5,"sodium":350,"sugar":1.0,"serving":200,"category":"foreign_popular"},
    "샤부샤부": {"cal":50,"prot":4.0,"carbs":3.0,"fat":2.5,"fiber":0.5,"sodium":300,"sugar":1.0,"serving":500,"category":"foreign_popular"},
    "스키야키": {"cal":80,"prot":5.0,"carbs":8.0,"fat":3.0,"fiber":0.5,"sodium":450,"sugar":5.0,"serving":500,"category":"foreign_popular"},
    "냉모밀": {"cal":95,"prot":5.0,"carbs":18.0,"fat":0.5,"fiber":1.5,"sodium":350,"sugar":3.0,"serving":350,"category":"foreign_popular"},
    "가라아게": {"cal":250,"prot":15.0,"carbs":12.0,"fat":16.0,"fiber":0.5,"sodium":400,"sugar":1.0,"serving":120,"category":"foreign_popular"},
    "모찌": {"cal":230,"prot":4.0,"carbs":52.0,"fat":0.5,"fiber":0.5,"sodium":50,"sugar":15.0,"serving":50,"category":"foreign_popular"},

    # ── 추가: 중식 ──
    "유린기": {"cal":200,"prot":14.0,"carbs":10.0,"fat":12.0,"fiber":0.5,"sodium":450,"sugar":4.0,"serving":200,"category":"korean"},
    "샤오롱바오": {"cal":200,"prot":8.0,"carbs":22.0,"fat":8.0,"fiber":0.5,"sodium":400,"sugar":1.0,"serving":120,"category":"foreign_popular"},
    "딤섬": {"cal":180,"prot":7.0,"carbs":20.0,"fat":7.0,"fiber":0.5,"sodium":400,"sugar":1.5,"serving":120,"category":"foreign_popular"},
    "차슈": {"cal":200,"prot":18.0,"carbs":5.0,"fat":12.0,"fiber":0,"sodium":400,"sugar":3.0,"serving":120,"category":"foreign_popular"},
    "훠궈": {"cal":60,"prot":4.0,"carbs":4.0,"fat":3.0,"fiber":0.5,"sodium":500,"sugar":2.0,"serving":600,"category":"foreign_popular"},
    "고량주": {"cal":220,"prot":0.0,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":0,"sugar":0,"serving":100,"category":"snack_drink"},

    # ── 추가: 음료/차 ──
    "둥굴레차": {"cal":1,"prot":0.0,"carbs":0.2,"fat":0.0,"fiber":0,"sodium":2,"sugar":0,"serving":200,"category":"snack_drink"},
    "현미차": {"cal":1,"prot":0.0,"carbs":0.2,"fat":0.0,"fiber":0,"sodium":2,"sugar":0,"serving":200,"category":"snack_drink"},
    "옥수수차": {"cal":1,"prot":0.0,"carbs":0.2,"fat":0.0,"fiber":0,"sodium":2,"sugar":0,"serving":200,"category":"snack_drink"},
    "캐모마일차": {"cal":1,"prot":0.0,"carbs":0.1,"fat":0.0,"fiber":0,"sodium":1,"sugar":0,"serving":200,"category":"snack_drink"},
    "카라멜마키아또": {"cal":100,"prot":3.0,"carbs":16.0,"fat":3.5,"fiber":0,"sodium":55,"sugar":14.0,"serving":350,"category":"snack_drink"},
    "아이스티": {"cal":30,"prot":0.0,"carbs":8.0,"fat":0.0,"fiber":0,"sodium":5,"sugar":7.0,"serving":350,"category":"snack_drink"},
    "딸기라떼": {"cal":85,"prot":2.5,"carbs":14.0,"fat":2.5,"fiber":0.3,"sodium":40,"sugar":12.0,"serving":350,"category":"snack_drink"},
    "환타": {"cal":42,"prot":0.0,"carbs":11.0,"fat":0.0,"fiber":0,"sodium":10,"sugar":11.0,"serving":250,"category":"snack_drink"},
    "모히또": {"cal":45,"prot":0.0,"carbs":8.0,"fat":0.0,"fiber":0,"sodium":5,"sugar":7.0,"serving":350,"category":"snack_drink"},
    "상그리아": {"cal":50,"prot":0.1,"carbs":8.0,"fat":0.0,"fiber":0.2,"sodium":5,"sugar":7.0,"serving":300,"category":"snack_drink"},
    "그릭요거트": {"cal":60,"prot":5.0,"carbs":4.0,"fat":2.5,"fiber":0,"sodium":30,"sugar":3.5,"serving":200,"category":"snack_drink"},

    # ── 추가: 빵/디저트 ──
    "에클레어": {"cal":260,"prot":5.0,"carbs":25.0,"fat":16.0,"fiber":0.5,"sodium":200,"sugar":12.0,"serving":100,"category":"snack_drink"},
    "스콘": {"cal":350,"prot":7.0,"carbs":45.0,"fat":16.0,"fiber":1.5,"sodium":400,"sugar":8.0,"serving":60,"category":"snack_drink"},
    "시나몬롤": {"cal":350,"prot":5.0,"carbs":48.0,"fat":15.0,"fiber":1.0,"sodium":350,"sugar":22.0,"serving":100,"category":"snack_drink"},
    "팥빵": {"cal":260,"prot":7.0,"carbs":48.0,"fat":4.0,"fiber":2.0,"sodium":200,"sugar":20.0,"serving":90,"category":"snack_drink"},
    "단팥빵": {"cal":270,"prot":7.0,"carbs":50.0,"fat":4.5,"fiber":2.0,"sodium":200,"sugar":22.0,"serving":90,"category":"snack_drink"},
    "젤리": {"cal":330,"prot":6.0,"carbs":78.0,"fat":0.0,"fiber":0,"sodium":15,"sugar":50.0,"serving":30,"category":"snack_drink"},
    "양갱": {"cal":250,"prot":3.0,"carbs":55.0,"fat":0.5,"fiber":2.0,"sodium":10,"sugar":35.0,"serving":50,"category":"snack_drink"},
    "송편": {"cal":200,"prot":3.5,"carbs":40.0,"fat":3.0,"fiber":1.0,"sodium":100,"sugar":10.0,"serving":100,"category":"korean"},
    "찰떡아이스": {"cal":250,"prot":3.0,"carbs":35.0,"fat":10.0,"fiber":0.5,"sodium":60,"sugar":18.0,"serving":70,"category":"snack_drink"},
    "녹차빙수": {"cal":150,"prot":3.0,"carbs":28.0,"fat":3.0,"fiber":0.5,"sodium":30,"sugar":18.0,"serving":250,"category":"snack_drink"},
    "망고빙수": {"cal":160,"prot":2.5,"carbs":32.0,"fat":3.0,"fiber":1.0,"sodium":25,"sugar":22.0,"serving":250,"category":"snack_drink"},
    "팥빙수": {"cal":140,"prot":3.0,"carbs":28.0,"fat":2.0,"fiber":1.5,"sodium":30,"sugar":18.0,"serving":250,"category":"snack_drink"},

    # ── 추가: 과일 ──
    "자몽": {"cal":32,"prot":0.6,"carbs":8.1,"fat":0.1,"fiber":1.1,"sodium":0,"sugar":7.0,"serving":200,"category":"fruit"},
    "레몬": {"cal":29,"prot":1.1,"carbs":9.3,"fat":0.3,"fiber":2.8,"sodium":2,"sugar":2.5,"serving":60,"category":"fruit"},
    "라임": {"cal":30,"prot":0.7,"carbs":10.5,"fat":0.2,"fiber":2.8,"sodium":2,"sugar":1.7,"serving":60,"category":"fruit"},
    "라즈베리": {"cal":52,"prot":1.2,"carbs":12.0,"fat":0.7,"fiber":6.5,"sodium":1,"sugar":4.4,"serving":80,"category":"fruit"},
    "곶감": {"cal":270,"prot":2.0,"carbs":65.0,"fat":0.5,"fiber":8.0,"sodium":10,"sugar":45.0,"serving":40,"category":"fruit"},

    # ── 추가: 해산물 ──
    "전복": {"cal":80,"prot":15.0,"carbs":4.0,"fat":0.5,"fiber":0,"sodium":250,"sugar":0,"serving":100,"category":"korean"},
    "꽃게": {"cal":63,"prot":13.0,"carbs":0.5,"fat":0.8,"fiber":0,"sodium":350,"sugar":0,"serving":150,"category":"korean"},
    "대하": {"cal":85,"prot":18.0,"carbs":0.8,"fat":0.7,"fiber":0,"sodium":200,"sugar":0,"serving":100,"category":"korean"},
    "가리비": {"cal":70,"prot":12.0,"carbs":3.5,"fat":0.5,"fiber":0,"sodium":250,"sugar":0,"serving":100,"category":"korean"},
    "멍게": {"cal":40,"prot":7.0,"carbs":2.5,"fat":0.5,"fiber":0,"sodium":300,"sugar":0,"serving":100,"category":"korean"},
    "낙지": {"cal":55,"prot":10.5,"carbs":1.0,"fat":0.8,"fiber":0,"sodium":200,"sugar":0,"serving":100,"category":"korean"},
    "문어": {"cal":73,"prot":15.0,"carbs":2.0,"fat":0.6,"fiber":0,"sodium":230,"sugar":0,"serving":100,"category":"korean"},
    "장어": {"cal":230,"prot":17.0,"carbs":0.0,"fat":18.0,"fiber":0,"sodium":60,"sugar":0,"serving":120,"category":"korean"},
    "꼬막": {"cal":70,"prot":12.0,"carbs":3.0,"fat":1.0,"fiber":0,"sodium":250,"sugar":0,"serving":100,"category":"korean"},
    "홍합": {"cal":73,"prot":10.0,"carbs":3.5,"fat":2.0,"fiber":0,"sodium":280,"sugar":0,"serving":100,"category":"korean"},

    # ── 추가: 간식/야식/견과 ──
    "호두": {"cal":654,"prot":15.0,"carbs":14.0,"fat":65.0,"fiber":7.0,"sodium":2,"sugar":2.6,"serving":28,"category":"snack_drink"},
    "땅콩": {"cal":567,"prot":26.0,"carbs":16.0,"fat":49.0,"fiber":9.0,"sodium":5,"sugar":4.0,"serving":30,"category":"snack_drink"},
    "군고구마": {"cal":130,"prot":1.5,"carbs":30.0,"fat":0.1,"fiber":3.5,"sodium":15,"sugar":10.0,"serving":150,"category":"korean"},
    "그래놀라바": {"cal":450,"prot":8.0,"carbs":60.0,"fat":20.0,"fiber":3.0,"sodium":150,"sugar":20.0,"serving":40,"category":"snack_drink"},
    "핫윙": {"cal":250,"prot":16.0,"carbs":8.0,"fat":17.0,"fiber":0.5,"sodium":500,"sugar":2.0,"serving":100,"category":"franchise"},
    "닭발": {"cal":180,"prot":15.0,"carbs":5.0,"fat":11.0,"fiber":0,"sodium":500,"sugar":3.0,"serving":120,"category":"korean"},
    "츄러스": {"cal":350,"prot":4.0,"carbs":40.0,"fat":19.0,"fiber":1.0,"sodium":250,"sugar":15.0,"serving":80,"category":"snack_drink"},

    # ── 추가: 미매칭 보완 ──
    "만둣국": {"cal":55,"prot":3.0,"carbs":8.0,"fat":1.5,"fiber":0.3,"sodium":400,"sugar":0.5,"serving":450,"category":"korean"},
    "소세지볶음": {"cal":240,"prot":10.0,"carbs":5.0,"fat":20.0,"fiber":0.5,"sodium":600,"sugar":2.0,"serving":100,"category":"korean"},
    "김전": {"cal":180,"prot":5.0,"carbs":15.0,"fat":11.0,"fiber":1.0,"sodium":300,"sugar":1.0,"serving":80,"category":"korean"},
    "쭈꾸미볶음": {"cal":90,"prot":10.0,"carbs":5.0,"fat":3.5,"fiber":0.5,"sodium":450,"sugar":3.0,"serving":250,"category":"korean"},
    "폭립": {"cal":250,"prot":18.0,"carbs":8.0,"fat":16.0,"fiber":0,"sodium":400,"sugar":5.0,"serving":200,"category":"foreign_popular"},
    "양념반후라이드반": {"cal":270,"prot":16.0,"carbs":12.0,"fat":17.0,"fiber":0.5,"sodium":520,"sugar":4.0,"serving":100,"category":"franchise"},
    "가츠동": {"cal":140,"prot":7.0,"carbs":18.0,"fat":4.5,"fiber":0.5,"sodium":400,"sugar":3.0,"serving":400,"category":"foreign_popular"},
    "텐푸라": {"cal":200,"prot":5.0,"carbs":20.0,"fat":11.0,"fiber":1.0,"sodium":300,"sugar":1.0,"serving":150,"category":"foreign_popular"},
    "스시": {"cal":150,"prot":6.0,"carbs":25.0,"fat":3.0,"fiber":0.3,"sodium":350,"sugar":4.0,"serving":200,"category":"foreign_popular"},
    "꿔바로우": {"cal":230,"prot":10.0,"carbs":25.0,"fat":10.0,"fiber":0.5,"sodium":400,"sugar":10.0,"serving":200,"category":"korean"},
    "경장육사": {"cal":180,"prot":12.0,"carbs":10.0,"fat":10.0,"fiber":1.0,"sodium":400,"sugar":4.0,"serving":200,"category":"korean"},
    "위스키": {"cal":250,"prot":0.0,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":0,"sugar":0,"serving":45,"category":"snack_drink"},
    "보드카": {"cal":231,"prot":0.0,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":0,"sugar":0,"serving":45,"category":"snack_drink"},
    "진토닉": {"cal":70,"prot":0.0,"carbs":7.0,"fat":0.0,"fiber":0,"sodium":5,"sugar":6.0,"serving":250,"category":"snack_drink"},
    "매실차": {"cal":20,"prot":0.0,"carbs":5.0,"fat":0.0,"fiber":0,"sodium":2,"sugar":4.0,"serving":200,"category":"snack_drink"},
    "생강차": {"cal":15,"prot":0.1,"carbs":3.5,"fat":0.0,"fiber":0,"sodium":2,"sugar":2.5,"serving":200,"category":"snack_drink"},
    "대추차": {"cal":25,"prot":0.2,"carbs":6.0,"fat":0.0,"fiber":0.3,"sodium":2,"sugar":4.5,"serving":200,"category":"snack_drink"},
    "인삼차": {"cal":15,"prot":0.3,"carbs":3.0,"fat":0.0,"fiber":0,"sodium":3,"sugar":2.0,"serving":200,"category":"snack_drink"},
    "허브차": {"cal":1,"prot":0.0,"carbs":0.1,"fat":0.0,"fiber":0,"sodium":1,"sugar":0,"serving":200,"category":"snack_drink"},
    "크레페": {"cal":200,"prot":5.0,"carbs":25.0,"fat":9.0,"fiber":0.5,"sodium":200,"sugar":8.0,"serving":120,"category":"foreign_popular"},
    "수플레": {"cal":180,"prot":5.0,"carbs":20.0,"fat":9.0,"fiber":0.3,"sodium":150,"sugar":12.0,"serving":120,"category":"snack_drink"},
    "푸딩": {"cal":130,"prot":3.0,"carbs":20.0,"fat":4.0,"fiber":0,"sodium":80,"sugar":15.0,"serving":120,"category":"snack_drink"},
    "판나코타": {"cal":200,"prot":3.0,"carbs":22.0,"fat":11.0,"fiber":0,"sodium":40,"sugar":18.0,"serving":120,"category":"snack_drink"},
    "크림브륄레": {"cal":280,"prot":4.0,"carbs":25.0,"fat":18.0,"fiber":0,"sodium":50,"sugar":20.0,"serving":120,"category":"snack_drink"},

    # ── Gold DB 데이터 보정 (비정상 칼로리 덮어쓰기) ──
    "대창구이": {"cal":360,"prot":10.0,"carbs":1.0,"fat":35.0,"fiber":0,"sodium":250,"sugar":0,"serving":100,"category":"korean"},
    "막창구이": {"cal":310,"prot":12.0,"carbs":1.0,"fat":29.0,"fiber":0,"sodium":250,"sugar":0,"serving":150,"category":"korean"},
    "어묵볶음": {"cal":130,"prot":8.0,"carbs":15.0,"fat":4.0,"fiber":0.5,"sodium":500,"sugar":5.0,"serving":80,"category":"korean"},
    "오징어채볶음": {"cal":200,"prot":12.0,"carbs":25.0,"fat":5.0,"fiber":0.5,"sodium":500,"sugar":10.0,"serving":40,"category":"korean"},
    "총각김치": {"cal":25,"prot":1.5,"carbs":4.5,"fat":0.3,"fiber":1.5,"sodium":550,"sugar":2.0,"serving":40,"category":"korean"},
    "파김치": {"cal":20,"prot":1.5,"carbs":3.0,"fat":0.5,"fiber":1.0,"sodium":500,"sugar":1.5,"serving":40,"category":"korean"},
    "라볶이": {"cal":170,"prot":5.0,"carbs":33.0,"fat":2.5,"fiber":1.0,"sodium":600,"sugar":8.0,"serving":350,"category":"korean"},
    "연어초밥": {"cal":150,"prot":6.0,"carbs":22.0,"fat":4.0,"fiber":0.2,"sodium":300,"sugar":3.0,"serving":50,"category":"foreign_popular"},
    "참치초밥": {"cal":140,"prot":6.5,"carbs":22.0,"fat":2.5,"fiber":0.2,"sodium":300,"sugar":3.0,"serving":50,"category":"foreign_popular"},
    "새우초밥": {"cal":130,"prot":5.0,"carbs":23.0,"fat":1.5,"fiber":0.2,"sodium":300,"sugar":3.5,"serving":50,"category":"foreign_popular"},
    "유부초밥": {"cal":180,"prot":5.0,"carbs":28.0,"fat":5.0,"fiber":0.5,"sodium":350,"sugar":6.0,"serving":200,"category":"foreign_popular"},
    # ── 누락 보강 (제미나이 테스트 피드백) ──
    "대하구이": {"cal":95,"prot":20.0,"carbs":0.5,"fat":1.5,"fiber":0,"sodium":350,"sugar":0,"serving":150,"category":"korean"},
    "똠얌꿍": {"cal":45,"prot":5.0,"carbs":3.5,"fat":1.5,"fiber":0.5,"sodium":650,"sugar":1.0,"serving":400,"category":"foreign_popular"},
    "단백질보충제": {"cal":120,"prot":25.0,"carbs":3.0,"fat":1.5,"fiber":0,"sodium":100,"sugar":1.5,"serving":30,"category":"diet_fitness"},
    "단백질쉐이크": {"cal":120,"prot":25.0,"carbs":3.0,"fat":1.5,"fiber":0,"sodium":100,"sugar":1.5,"serving":30,"category":"diet_fitness"},
    "프로틴쉐이크": {"cal":120,"prot":25.0,"carbs":3.0,"fat":1.5,"fiber":0,"sodium":100,"sugar":1.5,"serving":30,"category":"diet_fitness"},
    "마라탕": {"cal":85,"prot":5.0,"carbs":6.0,"fat":4.5,"fiber":1.0,"sodium":800,"sugar":1.0,"serving":500,"category":"foreign_popular"},
    "마라샹궈": {"cal":150,"prot":8.0,"carbs":10.0,"fat":9.0,"fiber":2.0,"sodium":900,"sugar":2.0,"serving":350,"category":"foreign_popular"},
    "훠궈": {"cal":70,"prot":6.0,"carbs":4.0,"fat":3.5,"fiber":1.0,"sodium":700,"sugar":1.0,"serving":500,"category":"foreign_popular"},
    "아사이볼": {"cal":200,"prot":3.0,"carbs":42.0,"fat":5.0,"fiber":4.0,"sodium":10,"sugar":28.0,"serving":300,"category":"diet_fitness"},
    "포케": {"cal":130,"prot":12.0,"carbs":15.0,"fat":3.0,"fiber":1.5,"sodium":500,"sugar":3.0,"serving":350,"category":"diet_fitness"},

    # ══════════════════════════════════════════════════════════
    #  확장: diet_fitness — 다이어터/피트니스 특화 (70종 추가)
    # ══════════════════════════════════════════════════════════
    # ── 곡물/탄수화물 소스 ──
    "오트밀": {"cal":379,"prot":13.2,"carbs":67.7,"fat":6.5,"fiber":10.1,"sodium":6,"sugar":1.0,"serving":40,"category":"diet_fitness"},
    "퀴노아": {"cal":368,"prot":14.1,"carbs":64.2,"fat":6.1,"fiber":7.0,"sodium":5,"sugar":0,"serving":50,"category":"diet_fitness"},
    "통밀빵": {"cal":247,"prot":13.0,"carbs":41.3,"fat":3.4,"fiber":6.8,"sodium":450,"sugar":6.0,"serving":60,"category":"diet_fitness"},
    "호밀빵": {"cal":259,"prot":8.5,"carbs":48.3,"fat":3.3,"fiber":5.8,"sodium":603,"sugar":3.9,"serving":60,"category":"diet_fitness"},
    "곤약밥": {"cal":10,"prot":0.1,"carbs":2.3,"fat":0.0,"fiber":2.0,"sodium":5,"sugar":0,"serving":200,"category":"diet_fitness"},
    "곤약면": {"cal":10,"prot":0.1,"carbs":2.3,"fat":0.0,"fiber":2.0,"sodium":5,"sugar":0,"serving":200,"category":"diet_fitness"},
    "곤약젤리": {"cal":12,"prot":0.0,"carbs":3.0,"fat":0.0,"fiber":1.5,"sodium":30,"sugar":1.5,"serving":150,"category":"diet_fitness"},
    "무가당시리얼": {"cal":360,"prot":10.0,"carbs":75.0,"fat":3.0,"fiber":8.0,"sodium":200,"sugar":3.0,"serving":30,"category":"diet_fitness"},
    "그래놀라": {"cal":450,"prot":10.0,"carbs":64.0,"fat":18.0,"fiber":5.0,"sodium":20,"sugar":18.0,"serving":50,"category":"diet_fitness"},
    # ── 단백질 소스 ──
    "닭안심": {"cal":110,"prot":23.5,"carbs":0.0,"fat":1.0,"fiber":0,"sodium":60,"sugar":0,"serving":100,"category":"diet_fitness"},
    "닭가슴살 소시지": {"cal":130,"prot":14.0,"carbs":8.0,"fat":4.0,"fiber":0,"sodium":500,"sugar":2.0,"serving":70,"category":"diet_fitness"},
    "닭가슴살 스테이크": {"cal":115,"prot":22.0,"carbs":2.0,"fat":2.0,"fiber":0,"sodium":400,"sugar":1.0,"serving":100,"category":"diet_fitness"},
    "닭가슴살 도시락": {"cal":95,"prot":12.0,"carbs":12.0,"fat":1.5,"fiber":1.0,"sodium":350,"sugar":2.0,"serving":300,"category":"diet_fitness"},
    "소고기 안심": {"cal":205,"prot":20.0,"carbs":0.0,"fat":13.5,"fiber":0,"sodium":54,"sugar":0,"serving":150,"category":"diet_fitness"},
    "소고기 등심": {"cal":250,"prot":18.5,"carbs":0.0,"fat":19.0,"fiber":0,"sodium":55,"sugar":0,"serving":150,"category":"diet_fitness"},
    "돼지 안심": {"cal":143,"prot":22.0,"carbs":0.0,"fat":5.5,"fiber":0,"sodium":50,"sugar":0,"serving":150,"category":"diet_fitness"},
    "돼지 등심": {"cal":180,"prot":21.0,"carbs":0.0,"fat":10.0,"fiber":0,"sodium":55,"sugar":0,"serving":150,"category":"diet_fitness"},
    "단백질바": {"cal":380,"prot":30.0,"carbs":40.0,"fat":12.0,"fiber":3.0,"sodium":200,"sugar":8.0,"serving":60,"category":"diet_fitness"},
    "프로틴바": {"cal":380,"prot":30.0,"carbs":40.0,"fat":12.0,"fiber":3.0,"sodium":200,"sugar":8.0,"serving":60,"category":"diet_fitness"},
    "삶은달걀": {"cal":155,"prot":13.0,"carbs":1.1,"fat":11.0,"fiber":0,"sodium":124,"sugar":1.1,"serving":60,"category":"diet_fitness"},
    "삶은계란": {"cal":155,"prot":13.0,"carbs":1.1,"fat":11.0,"fiber":0,"sodium":124,"sugar":1.1,"serving":60,"category":"diet_fitness"},
    "훈제연어": {"cal":117,"prot":18.3,"carbs":0.0,"fat":4.3,"fiber":0,"sodium":672,"sugar":0,"serving":50,"category":"diet_fitness"},
    "참치캔": {"cal":128,"prot":26.0,"carbs":0.0,"fat":2.5,"fiber":0,"sodium":350,"sugar":0,"serving":100,"category":"diet_fitness"},
    "새우": {"cal":99,"prot":24.0,"carbs":0.2,"fat":0.3,"fiber":0,"sodium":111,"sugar":0,"serving":100,"category":"diet_fitness"},
    "관자": {"cal":69,"prot":12.1,"carbs":3.2,"fat":0.5,"fiber":0,"sodium":392,"sugar":0,"serving":80,"category":"diet_fitness"},
    "쇠고기육포": {"cal":320,"prot":50.0,"carbs":11.0,"fat":8.0,"fiber":0,"sodium":1800,"sugar":9.0,"serving":30,"category":"diet_fitness"},
    # ── 유제품/대체우유 ──
    "무지방우유": {"cal":35,"prot":3.4,"carbs":5.0,"fat":0.1,"fiber":0,"sodium":52,"sugar":5.0,"serving":200,"category":"diet_fitness"},
    "저지방우유": {"cal":46,"prot":3.3,"carbs":5.0,"fat":1.5,"fiber":0,"sodium":52,"sugar":5.0,"serving":200,"category":"diet_fitness"},
    "두유": {"cal":54,"prot":3.3,"carbs":6.3,"fat":1.8,"fiber":0.6,"sodium":25,"sugar":3.0,"serving":200,"category":"diet_fitness"},
    "아몬드밀크": {"cal":15,"prot":0.5,"carbs":0.6,"fat":1.1,"fiber":0.5,"sodium":65,"sugar":0,"serving":200,"category":"diet_fitness"},
    "오트밀크": {"cal":40,"prot":0.3,"carbs":7.0,"fat":1.5,"fiber":0.8,"sodium":45,"sugar":4.0,"serving":200,"category":"diet_fitness"},
    "코티지치즈": {"cal":98,"prot":11.1,"carbs":3.4,"fat":4.3,"fiber":0,"sodium":364,"sugar":2.7,"serving":100,"category":"diet_fitness"},
    "리코타치즈": {"cal":174,"prot":11.3,"carbs":3.0,"fat":13.0,"fiber":0,"sodium":84,"sugar":0.3,"serving":60,"category":"diet_fitness"},
    "플레인요거트": {"cal":63,"prot":5.3,"carbs":7.0,"fat":1.5,"fiber":0,"sodium":46,"sugar":7.0,"serving":150,"category":"diet_fitness"},
    # ── 견과류/씨앗/건강지방 ──
    "아보카도": {"cal":160,"prot":2.0,"carbs":8.5,"fat":14.7,"fiber":6.7,"sodium":7,"sugar":0.7,"serving":70,"category":"diet_fitness"},
    "호두": {"cal":654,"prot":15.2,"carbs":13.7,"fat":65.2,"fiber":6.7,"sodium":2,"sugar":2.6,"serving":20,"category":"diet_fitness"},
    "캐슈넛": {"cal":553,"prot":18.2,"carbs":30.2,"fat":43.9,"fiber":3.3,"sodium":12,"sugar":5.9,"serving":20,"category":"diet_fitness"},
    "피스타치오": {"cal":560,"prot":20.2,"carbs":27.2,"fat":45.3,"fiber":10.6,"sodium":1,"sugar":7.7,"serving":20,"category":"diet_fitness"},
    "치아시드": {"cal":486,"prot":16.5,"carbs":42.1,"fat":30.7,"fiber":34.4,"sodium":16,"sugar":0,"serving":15,"category":"diet_fitness"},
    "아마씨": {"cal":534,"prot":18.3,"carbs":28.9,"fat":42.2,"fiber":27.3,"sodium":30,"sugar":1.6,"serving":10,"category":"diet_fitness"},
    "땅콩버터": {"cal":588,"prot":25.1,"carbs":20.0,"fat":50.4,"fiber":6.0,"sodium":459,"sugar":9.2,"serving":15,"category":"diet_fitness"},
    "아몬드버터": {"cal":614,"prot":21.0,"carbs":18.8,"fat":55.5,"fiber":10.5,"sodium":7,"sugar":6.3,"serving":15,"category":"diet_fitness"},
    # ── 건강 간식/보충제 ──
    "프로틴요거트": {"cal":80,"prot":15.0,"carbs":6.0,"fat":0.5,"fiber":0,"sodium":50,"sugar":4.0,"serving":150,"category":"diet_fitness"},
    "프로틴음료": {"cal":100,"prot":20.0,"carbs":5.0,"fat":1.0,"fiber":0,"sodium":80,"sugar":2.0,"serving":250,"category":"diet_fitness"},
    "에너지바": {"cal":400,"prot":10.0,"carbs":55.0,"fat":16.0,"fiber":3.0,"sodium":150,"sugar":20.0,"serving":60,"category":"diet_fitness"},
    "다크초콜릿": {"cal":546,"prot":5.5,"carbs":60.0,"fat":32.0,"fiber":7.0,"sodium":6,"sugar":48.0,"serving":20,"category":"diet_fitness"},
    "건크랜베리": {"cal":308,"prot":0.1,"carbs":82.4,"fat":1.4,"fiber":5.7,"sodium":3,"sugar":72.6,"serving":20,"category":"diet_fitness"},
    "건망고": {"cal":319,"prot":2.5,"carbs":78.6,"fat":0.9,"fiber":2.4,"sodium":162,"sugar":73.0,"serving":30,"category":"diet_fitness"},
    "건블루베리": {"cal":317,"prot":3.0,"carbs":76.0,"fat":3.0,"fiber":7.0,"sodium":3,"sugar":62.0,"serving":20,"category":"diet_fitness"},
    # ── 다이어트 대체식/기능식 ──
    "쉐이크대체식": {"cal":200,"prot":20.0,"carbs":20.0,"fat":5.0,"fiber":5.0,"sodium":300,"sugar":5.0,"serving":55,"category":"diet_fitness"},
    "뉴트리바": {"cal":200,"prot":10.0,"carbs":25.0,"fat":8.0,"fiber":3.0,"sodium":150,"sugar":10.0,"serving":50,"category":"diet_fitness"},
    "닭가슴살볼": {"cal":120,"prot":18.0,"carbs":5.0,"fat":3.0,"fiber":0,"sodium":400,"sugar":2.0,"serving":100,"category":"diet_fitness"},
    "닭가슴살큐브": {"cal":115,"prot":22.0,"carbs":2.0,"fat":2.0,"fiber":0,"sodium":380,"sugar":1.0,"serving":100,"category":"diet_fitness"},
    "닭가슴살만두": {"cal":150,"prot":12.0,"carbs":18.0,"fat":4.0,"fiber":1.0,"sodium":500,"sugar":2.0,"serving":120,"category":"diet_fitness"},
    "곤약볶음밥": {"cal":30,"prot":1.5,"carbs":5.0,"fat":0.5,"fiber":2.0,"sodium":300,"sugar":1.0,"serving":200,"category":"diet_fitness"},
    "저칼로리떡볶이": {"cal":80,"prot":2.0,"carbs":16.0,"fat":1.0,"fiber":2.0,"sodium":500,"sugar":5.0,"serving":200,"category":"diet_fitness"},
    "제로칼로리음료": {"cal":0,"prot":0.0,"carbs":0.0,"fat":0.0,"fiber":0,"sodium":10,"sugar":0,"serving":250,"category":"diet_fitness"},
    "스포츠음료": {"cal":26,"prot":0.0,"carbs":6.5,"fat":0.0,"fiber":0,"sodium":50,"sugar":6.0,"serving":250,"category":"diet_fitness"},
    "BCAA음료": {"cal":10,"prot":2.5,"carbs":0.5,"fat":0.0,"fiber":0,"sodium":20,"sugar":0,"serving":300,"category":"diet_fitness"},

    # ══════════════════════════════════════════════════════════
    #  확장: franchise — 프랜차이즈/편의점 특화 (80종 추가)
    # ══════════════════════════════════════════════════════════
    # ── 맥도날드 ──
    "빅맥": {"cal":257,"prot":13.0,"carbs":21.0,"fat":13.5,"fiber":1.5,"sodium":480,"sugar":4.5,"serving":200,"category":"franchise"},
    "맥스파이시상하이버거": {"cal":240,"prot":13.0,"carbs":22.0,"fat":11.0,"fiber":1.0,"sodium":550,"sugar":3.0,"serving":210,"category":"franchise"},
    "맥너겟": {"cal":250,"prot":15.0,"carbs":15.0,"fat":15.0,"fiber":1.0,"sodium":500,"sugar":0.5,"serving":100,"category":"franchise"},
    "에그맥머핀": {"cal":190,"prot":12.0,"carbs":16.5,"fat":8.5,"fiber":1.0,"sodium":480,"sugar":2.0,"serving":140,"category":"franchise"},
    "치즈버거": {"cal":250,"prot":13.0,"carbs":25.0,"fat":11.0,"fiber":1.5,"sodium":560,"sugar":5.0,"serving":120,"category":"franchise"},
    "1955버거": {"cal":260,"prot":14.0,"carbs":19.0,"fat":15.0,"fiber":1.0,"sodium":500,"sugar":4.0,"serving":220,"category":"franchise"},
    # ── 맘스터치 ──
    "싸이버거": {"cal":245,"prot":12.0,"carbs":24.0,"fat":11.5,"fiber":1.0,"sodium":600,"sugar":4.0,"serving":200,"category":"franchise"},
    "양념치킨버거": {"cal":260,"prot":11.0,"carbs":28.0,"fat":12.0,"fiber":1.0,"sodium":650,"sugar":6.0,"serving":210,"category":"franchise"},
    # ── 버거킹 ──
    "와퍼": {"cal":235,"prot":13.0,"carbs":20.0,"fat":12.5,"fiber":1.5,"sodium":510,"sugar":5.0,"serving":270,"category":"franchise"},
    "통새우와퍼": {"cal":230,"prot":11.0,"carbs":22.0,"fat":11.5,"fiber":1.0,"sodium":580,"sugar":4.0,"serving":250,"category":"franchise"},
    # ── 서브웨이 ──
    "이탈리안BMT": {"cal":160,"prot":9.0,"carbs":18.0,"fat":6.0,"fiber":2.0,"sodium":600,"sugar":3.0,"serving":240,"category":"franchise"},
    "로스트치킨": {"cal":140,"prot":11.0,"carbs":17.0,"fat":3.0,"fiber":2.0,"sodium":450,"sugar":3.0,"serving":240,"category":"franchise"},
    "에그마요": {"cal":175,"prot":8.0,"carbs":18.0,"fat":8.0,"fiber":2.0,"sodium":400,"sugar":3.0,"serving":240,"category":"franchise"},
    "터키베이컨아보카도": {"cal":145,"prot":10.0,"carbs":17.0,"fat":4.5,"fiber":2.0,"sodium":500,"sugar":3.0,"serving":250,"category":"franchise"},
    # ── 스타벅스 ──
    "아메리카노": {"cal":5,"prot":0.3,"carbs":0.5,"fat":0.0,"fiber":0,"sodium":5,"sugar":0,"serving":355,"category":"franchise"},
    "카페라떼": {"cal":38,"prot":2.3,"carbs":3.5,"fat":1.5,"fiber":0,"sodium":40,"sugar":3.5,"serving":355,"category":"franchise"},
    "바닐라라떼": {"cal":55,"prot":2.0,"carbs":8.5,"fat":1.5,"fiber":0,"sodium":45,"sugar":8.0,"serving":355,"category":"franchise"},
    "카라멜마끼아또": {"cal":58,"prot":2.5,"carbs":9.0,"fat":1.5,"fiber":0,"sodium":50,"sugar":8.5,"serving":355,"category":"franchise"},
    "자바칩프라푸치노": {"cal":90,"prot":1.5,"carbs":14.0,"fat":3.5,"fiber":0.5,"sodium":60,"sugar":12.0,"serving":355,"category":"franchise"},
    "딸기프라푸치노": {"cal":82,"prot":0.5,"carbs":15.0,"fat":2.5,"fiber":0,"sodium":30,"sugar":14.0,"serving":355,"category":"franchise"},
    # ── 투썸플레이스/카페 ──
    "아이스티": {"cal":35,"prot":0.0,"carbs":9.0,"fat":0.0,"fiber":0,"sodium":5,"sugar":8.5,"serving":355,"category":"franchise"},
    "녹차라떼": {"cal":55,"prot":2.0,"carbs":9.0,"fat":1.5,"fiber":0.5,"sodium":35,"sugar":7.0,"serving":355,"category":"franchise"},
    "초코라떼": {"cal":70,"prot":2.5,"carbs":12.0,"fat":2.0,"fiber":0.5,"sodium":60,"sugar":10.0,"serving":355,"category":"franchise"},
    # ── KFC/치킨 ──
    "후라이드치킨한조각": {"cal":280,"prot":18.0,"carbs":10.0,"fat":19.0,"fiber":0.5,"sodium":600,"sugar":0.5,"serving":100,"category":"franchise"},
    "양념치킨한조각": {"cal":300,"prot":17.0,"carbs":18.0,"fat":17.0,"fiber":0.5,"sodium":650,"sugar":8.0,"serving":100,"category":"franchise"},
    "간장치킨한조각": {"cal":290,"prot":17.0,"carbs":15.0,"fat":18.0,"fiber":0.5,"sodium":700,"sugar":5.0,"serving":100,"category":"franchise"},
    # ── 편의점 도시락/간편식 ──
    "편의점도시락": {"cal":140,"prot":6.0,"carbs":22.0,"fat":3.5,"fiber":1.0,"sodium":500,"sugar":3.0,"serving":350,"category":"franchise"},
    "삼각김밥": {"cal":170,"prot":4.0,"carbs":32.0,"fat":2.5,"fiber":0.5,"sodium":400,"sugar":2.0,"serving":110,"category":"franchise"},
    "참치마요삼각김밥": {"cal":185,"prot":5.5,"carbs":30.0,"fat":5.0,"fiber":0.5,"sodium":450,"sugar":2.0,"serving":115,"category":"franchise"},
    "불고기삼각김밥": {"cal":180,"prot":5.0,"carbs":31.0,"fat":3.5,"fiber":0.5,"sodium":430,"sugar":4.0,"serving":115,"category":"franchise"},
    "컵밥": {"cal":130,"prot":5.0,"carbs":20.0,"fat":3.0,"fiber":1.0,"sodium":500,"sugar":3.0,"serving":300,"category":"franchise"},
    "편의점샌드위치": {"cal":200,"prot":9.0,"carbs":22.0,"fat":8.0,"fiber":1.5,"sodium":500,"sugar":4.0,"serving":150,"category":"franchise"},
    "편의점샐러드": {"cal":40,"prot":3.0,"carbs":5.0,"fat":1.0,"fiber":2.0,"sodium":200,"sugar":3.0,"serving":200,"category":"franchise"},
    "편의점닭가슴살": {"cal":110,"prot":22.0,"carbs":2.0,"fat":1.5,"fiber":0,"sodium":400,"sugar":1.0,"serving":100,"category":"franchise"},
    "편의점프로틴음료": {"cal":50,"prot":10.0,"carbs":2.0,"fat":0.5,"fiber":0,"sodium":60,"sugar":1.0,"serving":250,"category":"franchise"},
    "즉석밥": {"cal":150,"prot":3.0,"carbs":33.0,"fat":0.5,"fiber":0.5,"sodium":5,"sugar":0,"serving":210,"category":"franchise"},
    # ── 분식 프랜차이즈 ──
    "신전떡볶이": {"cal":200,"prot":4.0,"carbs":38.0,"fat":3.5,"fiber":1.5,"sodium":700,"sugar":8.0,"serving":300,"category":"franchise"},
    "엽기떡볶이": {"cal":210,"prot":4.5,"carbs":38.0,"fat":4.0,"fiber":1.5,"sodium":750,"sugar":9.0,"serving":300,"category":"franchise"},
    # ── 교촌/BBQ/BHC 치킨 ──
    "교촌허니콤보": {"cal":310,"prot":16.0,"carbs":20.0,"fat":18.0,"fiber":0.5,"sodium":700,"sugar":10.0,"serving":100,"category":"franchise"},
    "BBQ올리브치킨": {"cal":270,"prot":18.0,"carbs":10.0,"fat":18.0,"fiber":0.5,"sodium":550,"sugar":1.0,"serving":100,"category":"franchise"},
    "BHC뿌링클": {"cal":300,"prot":16.0,"carbs":16.0,"fat":19.0,"fiber":0.5,"sodium":650,"sugar":4.0,"serving":100,"category":"franchise"},
    # ── 피자 프랜차이즈 ──
    "페퍼로니피자": {"cal":280,"prot":12.0,"carbs":30.0,"fat":12.5,"fiber":2.0,"sodium":700,"sugar":3.0,"serving":130,"category":"franchise"},
    "콤비네이션피자": {"cal":260,"prot":11.0,"carbs":28.0,"fat":11.5,"fiber":2.0,"sodium":650,"sugar":3.0,"serving":130,"category":"franchise"},
    "포테이토피자": {"cal":270,"prot":9.0,"carbs":32.0,"fat":12.0,"fiber":1.5,"sodium":600,"sugar":3.0,"serving":130,"category":"franchise"},
    "고르곤졸라피자": {"cal":285,"prot":10.0,"carbs":28.0,"fat":14.0,"fiber":1.0,"sodium":620,"sugar":4.0,"serving":130,"category":"franchise"},
    # ── 도넛/베이커리 ──
    "글레이즈도넛": {"cal":380,"prot":4.5,"carbs":49.0,"fat":18.0,"fiber":1.0,"sodium":340,"sugar":22.0,"serving":60,"category":"franchise"},
    "크림치즈베이글": {"cal":280,"prot":10.0,"carbs":40.0,"fat":9.0,"fiber":2.0,"sodium":450,"sugar":5.0,"serving":100,"category":"franchise"},

    # ══════════════════════════════════════════════════════════
    #  확장: korean — 한식 반찬/국/탕/일상식 보강 (30종 추가)
    # ══════════════════════════════════════════════════════════
    "시금치나물": {"cal":25,"prot":2.5,"carbs":2.5,"fat":0.8,"fiber":2.0,"sodium":300,"sugar":0.5,"serving":60,"category":"korean"},
    "콩나물무침": {"cal":35,"prot":3.0,"carbs":4.0,"fat":1.0,"fiber":1.5,"sodium":350,"sugar":1.0,"serving":60,"category":"korean"},
    "숙주나물": {"cal":20,"prot":2.0,"carbs":2.5,"fat":0.3,"fiber":1.0,"sodium":280,"sugar":0.5,"serving":60,"category":"korean"},
    "미역줄기볶음": {"cal":30,"prot":1.5,"carbs":4.0,"fat":1.0,"fiber":2.5,"sodium":500,"sugar":1.0,"serving":60,"category":"korean"},
    "깻잎장아찌": {"cal":40,"prot":2.0,"carbs":5.0,"fat":1.5,"fiber":2.0,"sodium":700,"sugar":2.0,"serving":30,"category":"korean"},
    "오이소박이": {"cal":18,"prot":1.0,"carbs":3.0,"fat":0.3,"fiber":1.0,"sodium":600,"sugar":1.5,"serving":50,"category":"korean"},
    "무생채": {"cal":25,"prot":0.8,"carbs":5.0,"fat":0.5,"fiber":1.5,"sodium":500,"sugar":2.5,"serving":60,"category":"korean"},
    "김치전": {"cal":180,"prot":5.0,"carbs":22.0,"fat":8.0,"fiber":1.5,"sodium":600,"sugar":2.0,"serving":100,"category":"korean"},
    "파전": {"cal":170,"prot":4.5,"carbs":20.0,"fat":8.0,"fiber":1.0,"sodium":500,"sugar":2.0,"serving":120,"category":"korean"},
    "해물파전": {"cal":160,"prot":7.0,"carbs":18.0,"fat":7.0,"fiber":1.0,"sodium":550,"sugar":2.0,"serving":150,"category":"korean"},
    "감자조림": {"cal":90,"prot":1.5,"carbs":18.0,"fat":1.5,"fiber":1.5,"sodium":400,"sugar":5.0,"serving":80,"category":"korean"},
    "멸치볶음": {"cal":200,"prot":18.0,"carbs":15.0,"fat":8.0,"fiber":0.5,"sodium":800,"sugar":5.0,"serving":30,"category":"korean"},
    "어묵볶음": {"cal":100,"prot":5.0,"carbs":12.0,"fat":3.5,"fiber":0.5,"sodium":600,"sugar":4.0,"serving":80,"category":"korean"},
    "소세지볶음": {"cal":230,"prot":8.0,"carbs":12.0,"fat":17.0,"fiber":0.5,"sodium":700,"sugar":4.0,"serving":80,"category":"korean"},
    "제육볶음": {"cal":190,"prot":16.0,"carbs":8.0,"fat":11.0,"fiber":1.0,"sodium":650,"sugar":4.0,"serving":150,"category":"korean"},
    "오징어볶음": {"cal":130,"prot":14.0,"carbs":10.0,"fat":4.0,"fiber":1.0,"sodium":600,"sugar":4.0,"serving":150,"category":"korean"},
    "갈비찜": {"cal":180,"prot":15.0,"carbs":10.0,"fat":9.0,"fiber":1.0,"sodium":500,"sugar":5.0,"serving":200,"category":"korean"},
    "안동찜닭": {"cal":150,"prot":14.0,"carbs":10.0,"fat":6.0,"fiber":1.0,"sodium":550,"sugar":4.0,"serving":250,"category":"korean"},
    "소고기뭇국": {"cal":30,"prot":3.0,"carbs":2.5,"fat":1.0,"fiber":0.5,"sodium":400,"sugar":0.5,"serving":350,"category":"korean"},
    "콩나물국": {"cal":20,"prot":2.0,"carbs":2.0,"fat":0.5,"fiber":0.5,"sodium":350,"sugar":0.5,"serving":350,"category":"korean"},
    "미역국": {"cal":25,"prot":2.5,"carbs":2.0,"fat":1.0,"fiber":1.0,"sodium":500,"sugar":0.3,"serving":350,"category":"korean"},
    "갈비탕": {"cal":80,"prot":7.0,"carbs":2.0,"fat":5.0,"fiber":0.3,"sodium":500,"sugar":0.5,"serving":500,"category":"korean"},
    "설렁탕": {"cal":55,"prot":6.0,"carbs":1.0,"fat":3.0,"fiber":0,"sodium":400,"sugar":0,"serving":500,"category":"korean"},
    "곰탕": {"cal":60,"prot":6.5,"carbs":1.0,"fat":3.5,"fiber":0,"sodium":400,"sugar":0,"serving":500,"category":"korean"},
    "육개장": {"cal":50,"prot":5.0,"carbs":3.0,"fat":2.5,"fiber":1.0,"sodium":600,"sugar":1.0,"serving":500,"category":"korean"},
    "순두부찌개": {"cal":45,"prot":4.5,"carbs":3.0,"fat":2.0,"fiber":0.5,"sodium":500,"sugar":1.0,"serving":400,"category":"korean"},
    "부대찌개": {"cal":70,"prot":5.0,"carbs":6.0,"fat":3.0,"fiber":1.0,"sodium":700,"sugar":2.0,"serving":400,"category":"korean"},
    "청국장찌개": {"cal":55,"prot":5.5,"carbs":4.0,"fat":2.5,"fiber":1.5,"sodium":550,"sugar":1.0,"serving":400,"category":"korean"},
    "잡채": {"cal":120,"prot":3.0,"carbs":18.0,"fat":4.0,"fiber":1.5,"sodium":450,"sugar":5.0,"serving":150,"category":"korean"},
    "동그랑땡": {"cal":200,"prot":10.0,"carbs":14.0,"fat":12.0,"fiber":1.0,"sodium":400,"sugar":2.0,"serving":80,"category":"korean"},

    # ══════════════════════════════════════════════════════════
    #  확장: vegetable — 채소 신규 카테고리 (40종 추가)
    # ══════════════════════════════════════════════════════════
    "양배추": {"cal":25,"prot":1.3,"carbs":5.8,"fat":0.1,"fiber":2.5,"sodium":18,"sugar":3.2,"serving":70,"category":"vegetable"},
    "양상추": {"cal":14,"prot":0.9,"carbs":2.9,"fat":0.1,"fiber":1.2,"sodium":10,"sugar":1.3,"serving":50,"category":"vegetable"},
    "당근": {"cal":41,"prot":0.9,"carbs":9.6,"fat":0.2,"fiber":2.8,"sodium":69,"sugar":4.7,"serving":70,"category":"vegetable"},
    "오이": {"cal":15,"prot":0.7,"carbs":3.6,"fat":0.1,"fiber":0.5,"sodium":2,"sugar":1.7,"serving":100,"category":"vegetable"},
    "토마토": {"cal":18,"prot":0.9,"carbs":3.9,"fat":0.2,"fiber":1.2,"sodium":5,"sugar":2.6,"serving":150,"category":"vegetable"},
    "방울토마토": {"cal":18,"prot":0.9,"carbs":3.9,"fat":0.2,"fiber":1.2,"sodium":5,"sugar":2.6,"serving":100,"category":"vegetable"},
    "파프리카": {"cal":31,"prot":1.0,"carbs":6.0,"fat":0.3,"fiber":2.1,"sodium":4,"sugar":4.2,"serving":100,"category":"vegetable"},
    "피망": {"cal":20,"prot":0.9,"carbs":4.6,"fat":0.2,"fiber":1.7,"sodium":3,"sugar":2.4,"serving":80,"category":"vegetable"},
    "시금치": {"cal":23,"prot":2.9,"carbs":3.6,"fat":0.4,"fiber":2.2,"sodium":79,"sugar":0.4,"serving":80,"category":"vegetable"},
    "깻잎": {"cal":37,"prot":3.5,"carbs":6.0,"fat":0.6,"fiber":4.2,"sodium":3,"sugar":0,"serving":10,"category":"vegetable"},
    "상추": {"cal":15,"prot":1.4,"carbs":2.8,"fat":0.2,"fiber":1.3,"sodium":4,"sugar":0.8,"serving":30,"category":"vegetable"},
    "청경채": {"cal":13,"prot":1.5,"carbs":2.2,"fat":0.2,"fiber":1.0,"sodium":65,"sugar":1.2,"serving":80,"category":"vegetable"},
    "케일": {"cal":49,"prot":4.3,"carbs":8.8,"fat":0.9,"fiber":3.6,"sodium":38,"sugar":2.3,"serving":50,"category":"vegetable"},
    "셀러리": {"cal":16,"prot":0.7,"carbs":3.0,"fat":0.2,"fiber":1.6,"sodium":80,"sugar":1.3,"serving":60,"category":"vegetable"},
    "비트": {"cal":43,"prot":1.6,"carbs":9.6,"fat":0.2,"fiber":2.8,"sodium":78,"sugar":6.8,"serving":80,"category":"vegetable"},
    "무": {"cal":18,"prot":0.7,"carbs":4.1,"fat":0.1,"fiber":1.6,"sodium":39,"sugar":1.9,"serving":100,"category":"vegetable"},
    "감자": {"cal":77,"prot":2.0,"carbs":17.5,"fat":0.1,"fiber":2.2,"sodium":6,"sugar":0.8,"serving":150,"category":"vegetable"},
    "단호박": {"cal":40,"prot":0.8,"carbs":9.0,"fat":0.1,"fiber":1.5,"sodium":1,"sugar":4.0,"serving":100,"category":"vegetable"},
    "호박": {"cal":17,"prot":0.6,"carbs":4.0,"fat":0.1,"fiber":0.5,"sodium":1,"sugar":2.0,"serving":100,"category":"vegetable"},
    "애호박": {"cal":17,"prot":1.2,"carbs":3.1,"fat":0.3,"fiber":1.0,"sodium":2,"sugar":2.5,"serving":100,"category":"vegetable"},
    "가지": {"cal":25,"prot":1.0,"carbs":5.9,"fat":0.2,"fiber":3.0,"sodium":2,"sugar":3.5,"serving":100,"category":"vegetable"},
    "양파": {"cal":40,"prot":1.1,"carbs":9.3,"fat":0.1,"fiber":1.7,"sodium":4,"sugar":4.2,"serving":80,"category":"vegetable"},
    "대파": {"cal":27,"prot":1.8,"carbs":6.5,"fat":0.3,"fiber":1.8,"sodium":4,"sugar":2.3,"serving":30,"category":"vegetable"},
    "마늘": {"cal":149,"prot":6.4,"carbs":33.1,"fat":0.5,"fiber":2.1,"sodium":17,"sugar":1.0,"serving":10,"category":"vegetable"},
    "생강": {"cal":80,"prot":1.8,"carbs":17.8,"fat":0.8,"fiber":2.0,"sodium":13,"sugar":1.7,"serving":10,"category":"vegetable"},
    "버섯": {"cal":22,"prot":3.1,"carbs":3.3,"fat":0.3,"fiber":1.0,"sodium":5,"sugar":2.0,"serving":50,"category":"vegetable"},
    "새송이버섯": {"cal":35,"prot":2.5,"carbs":6.5,"fat":0.3,"fiber":1.5,"sodium":5,"sugar":1.5,"serving":50,"category":"vegetable"},
    "팽이버섯": {"cal":37,"prot":2.7,"carbs":7.8,"fat":0.3,"fiber":2.7,"sodium":3,"sugar":0.2,"serving":50,"category":"vegetable"},
    "표고버섯": {"cal":34,"prot":2.2,"carbs":6.8,"fat":0.5,"fiber":2.5,"sodium":9,"sugar":2.4,"serving":30,"category":"vegetable"},
    "콩나물": {"cal":34,"prot":3.5,"carbs":4.5,"fat":0.7,"fiber":1.5,"sodium":5,"sugar":1.0,"serving":80,"category":"vegetable"},
    "숙주": {"cal":26,"prot":2.8,"carbs":3.5,"fat":0.2,"fiber":1.2,"sodium":5,"sugar":1.0,"serving":80,"category":"vegetable"},
    "미나리": {"cal":20,"prot":2.0,"carbs":3.5,"fat":0.3,"fiber":2.0,"sodium":60,"sugar":1.0,"serving":30,"category":"vegetable"},
    "부추": {"cal":30,"prot":3.0,"carbs":4.4,"fat":0.3,"fiber":2.6,"sodium":7,"sugar":1.5,"serving":30,"category":"vegetable"},
    "고추": {"cal":29,"prot":1.3,"carbs":6.7,"fat":0.3,"fiber":1.5,"sodium":4,"sugar":3.5,"serving":15,"category":"vegetable"},
    "옥수수": {"cal":86,"prot":3.3,"carbs":19.0,"fat":1.2,"fiber":2.7,"sodium":15,"sugar":3.2,"serving":150,"category":"vegetable"},
    "연근": {"cal":74,"prot":2.6,"carbs":17.2,"fat":0.1,"fiber":4.9,"sodium":40,"sugar":0.5,"serving":60,"category":"vegetable"},
    "우엉": {"cal":72,"prot":1.5,"carbs":17.3,"fat":0.2,"fiber":3.3,"sodium":5,"sugar":2.9,"serving":60,"category":"vegetable"},
    "아스파라거스": {"cal":20,"prot":2.2,"carbs":3.9,"fat":0.1,"fiber":2.1,"sodium":2,"sugar":1.9,"serving":80,"category":"vegetable"},
    "루꼴라": {"cal":25,"prot":2.6,"carbs":3.7,"fat":0.7,"fiber":1.6,"sodium":27,"sugar":2.1,"serving":30,"category":"vegetable"},
    "적양배추": {"cal":31,"prot":1.4,"carbs":7.4,"fat":0.2,"fiber":2.1,"sodium":28,"sugar":3.8,"serving":70,"category":"vegetable"},
}


# ── 접미사 기반 음식 카테고리 매핑 ──
# "옥수수차" → "차", "녹차빙수" → "빙수" 등 접미사로 카테고리 판별
# 긴 접미사 우선 매칭 (순서 중요)
FOOD_SUFFIXES = [
    # 음료/차 (4자 이상 우선)
    ("마키아또", None),  # CORE_FOODS에 직접 있으면 그걸 사용
    ("라떼", None),
    ("스무디", "스무디"),
    ("밀크티", "밀크티"),
    # 수프/찌개
    ("차우더", "클램차우더"),
    ("수프", "수프"),
    ("찌개", None),
    ("전골", None),
    # 빙수/디저트
    ("빙수", None),  # 녹차빙수, 망고빙수 등은 CORE_FOODS에 직접 등록
    ("케이크", "케이크"),
    # 밥/면류
    ("덮밥", "덮밥"),
    ("볶음밥", "볶음밥"),
    ("파스타", "파스타"),
    ("리조또", "리조또"),
    # 고기/요리
    ("치킨", "치킨"),
    ("돈까스", "돈까스"),
    ("피자", "피자"),
    # 기타
    ("샐러드", "샐러드"),
    ("샌드위치", "샌드위치"),
    ("버거", "햄버거"),
    # 짧은 접미사 (마지막에)
    ("차", None),  # "옥수수차" → 차 (0~2kcal)
]


def _search_core_foods(name):
    """핵심 음식 참조 테이블에서 검색 (정확 → 접미사 → 복합이름 뒤에서부터 → 부분 매칭)

    핵심: "토마토 소스 파스타"는 "토마토"가 아니라 "파스타"에 매칭되어야 함.
    핵심: "옥수수차"는 "옥수수"가 아니라 "차" 계열로 매칭되어야 함.
    """
    # 1) 정확 매칭
    if name in CORE_FOODS:
        return name, CORE_FOODS[name]

    # 2) 접미사 기반 매칭 — 붙어있는 복합어 처리
    #    "옥수수차" → suffix "차", "녹차빙수" → suffix "빙수"
    #    단, 접미사만으로 된 이름은 건너뜀 (name == suffix)
    for suffix, fallback_key in FOOD_SUFFIXES:
        if name.endswith(suffix) and len(name) > len(suffix):
            # 전체 이름이 CORE_FOODS에 있으면 우선 (이미 1단계에서 체크됨)
            # 접미사 조합이 CORE_FOODS에 있는지 체크
            if name in CORE_FOODS:
                return name, CORE_FOODS[name]
            # fallback 키 사용
            if fallback_key and fallback_key in CORE_FOODS:
                return fallback_key, CORE_FOODS[fallback_key]
            # suffix 자체가 CORE_FOODS에 있으면 사용
            if suffix in CORE_FOODS:
                return suffix, CORE_FOODS[suffix]

    # 3) 복합 이름 (공백 구분): 뒤에서부터 핵심어 매칭 (긴 조합 우선)
    #    "리코타 치즈 샐러드" → "리코타 치즈 샐러드" 정확 매칭 시도
    #    없으면 → "치즈 샐러드" → "샐러드" (뒤에서부터)
    words = name.split()
    if len(words) > 1:
        # 뒤에서부터 단어를 늘려가며 시도 (긴 조합 우선)
        for start in range(len(words)):
            combo = ' '.join(words[start:])
            if combo in CORE_FOODS:
                return combo, CORE_FOODS[combo]
        # 뒤에서부터 단일 단어 매칭 (최소 2자)
        for word in reversed(words):
            if len(word) >= 2 and word in CORE_FOODS:
                return word, CORE_FOODS[word]

    # 4) 부분 매칭: name이 CORE_FOODS 키로 시작 (가장 긴 매칭 우선)
    #    예: "김치볶음밥" → "김치볶음밥" (정확 먼저, 없으면 "김치")
    #    단, 매칭 키 길이가 원본의 40% 이상이어야 관련성 인정
    best_key = None
    best_len = 0
    for key in CORE_FOODS:
        if name.startswith(key) and len(key) > best_len:
            best_key = key
            best_len = len(key)
        elif key.startswith(name) and len(name) >= 2 and len(name) > best_len:
            best_key = key
            best_len = len(name)

    if best_key and best_len >= len(name) * 0.4:
        return best_key, CORE_FOODS[best_key]

    # 5) 공백/구분자 뒤에 키가 나오는 경우 (긴 키 우선)
    #    예: "매운 김치찌개" → "김치찌개"
    for key in sorted(CORE_FOODS.keys(), key=len, reverse=True):
        if len(key) >= 2:
            idx = name.find(key)
            if idx > 0:
                prev = name[idx - 1]
                if prev in ' _(/·\t':
                    return key, CORE_FOODS[key]

    return None, None


# ── 골드 테이블 검색 (1,771건 검증된 데이터) ──
_GOLD_DATA = None

def _load_gold_db():
    """골드 테이블 JSON 로드 (메모리 캐싱)"""
    global _GOLD_DATA
    if _GOLD_DATA is not None:
        return _GOLD_DATA
    gold_path = Path(__file__).parent.parent / 'gold_foods.json'
    if gold_path.exists():
        with open(gold_path, 'r', encoding='utf-8') as f:
            _GOLD_DATA = json.load(f)
        return _GOLD_DATA
    return None

def _normalize_food_name(name):
    """AI가 반환한 음식명을 검색 가능한 형태로 정규화"""
    import re
    n = name.strip()
    # 괄호 안 수량 제거: "삼겹살 (200g)" → "삼겹살"
    n = re.sub(r'\s*\(\d+[가-힣a-zA-Z]*\)\s*', ' ', n)
    # 앞뒤 공백/특수문자
    n = re.sub(r'^[!@#$%^&*()"\']+', '', n)
    n = n.strip()
    return n

def _search_gold(name):
    """골드 테이블에서 음식 검색 — 정확→CORE→관련성 검증 매칭

    핵심 원칙:
    - 짧은 이름(1~2자)은 정확 매칭만 허용 (부분매칭 금지)
    - 복합 이름은 뒤에서부터 핵심어 매칭 (한식: "토마토 소스 파스타" → "파스타" 우선)
    - 매칭된 이름과 원본의 관련성이 낮으면 기각
    """
    import re
    gold = _load_gold_db()
    if gold is None:
        return None, None

    clean = _normalize_food_name(name)

    # 1) CORE_FOODS 매칭 (최우선 — 수동 검증 + 보정 데이터)
    #    Gold DB에 같은 이름이 있어도 CORE_FOODS가 이김
    core_key, core_data = _search_core_foods(clean)
    if core_data:
        return core_key, {
            'cal': core_data['cal'], 'prot': core_data['prot'],
            'carbs': core_data['carbs'], 'fat': core_data['fat'],
            'fiber': core_data.get('fiber', 0), 'sodium': core_data.get('sodium', 0),
            'sugar': core_data.get('sugar', 0), 'serving': core_data['serving'],
            'category': core_data['category'], 'source': 'core',
        }

    # 2) 골드 테이블 정확 매칭
    if clean in gold:
        return clean, gold[clean]

    # ── 짧은 이름(1~2자)은 여기서 중단 — 부분매칭 위험 너무 높음 ──
    if len(clean) <= 2:
        return None, None

    # 3) 골드 키가 clean으로 시작 (최소 3자, 가장 짧은 것)
    starts_with = [(k, v) for k, v in gold.items() if k.startswith(clean) and len(clean) >= 3]
    if starts_with:
        starts_with.sort(key=lambda x: len(x[0]))
        return starts_with[0][0], starts_with[0][1]

    # 4) clean이 골드 키로 시작 (가장 긴 매칭 우선, 최소 3자, 관련성 체크)
    contained = [(k, v) for k, v in gold.items() if clean.startswith(k) and len(k) >= 3]
    if contained:
        contained.sort(key=lambda x: len(x[0]), reverse=True)
        best_k, best_v = contained[0]
        # 관련성 체크: 매칭된 키 길이가 원본의 40% 이상이어야 함
        if len(best_k) >= len(clean) * 0.4:
            return best_k, best_v

    # 5) 복합 이름: 뒤에서부터 핵심어 매칭
    #    "토마토 소스 파스타" → "파스타" 먼저, "소스" 다음, "토마토" 마지막
    #    "매운 김치찌개" → "김치찌개" 먼저, "매운" 다음
    words = clean.split()
    if len(words) > 1:
        # 뒤에서부터 (핵심어가 보통 뒤에 옴)
        for word in reversed(words):
            if len(word) >= 2 and word in gold:
                return word, gold[word]
        # 뒤에서 2개씩 묶어서 시도: "소스 파스타" → "크림 파스타" 등
        for i in range(len(words)-1, 0, -1):
            combo = ' '.join(words[i-1:i+1])
            if combo in gold:
                return combo, gold[combo]

    # 6) 접미사 기반 골드 매칭 (CORE_FOODS에 없을 때 골드 테이블에서도 접미사로 시도)
    for suffix, fallback_key in FOOD_SUFFIXES:
        if clean.endswith(suffix) and len(clean) > len(suffix):
            if clean in gold:
                return clean, gold[clean]
            if fallback_key and fallback_key in gold:
                return fallback_key, gold[fallback_key]
            if suffix in gold:
                return suffix, gold[suffix]

    # 7) 검색키 기반 퍼지 매칭 (최소 3자, 관련성 체크)
    search_key = re.sub(r'[^가-힣a-zA-Z0-9]', '', clean).lower()
    if len(search_key) >= 3:
        # 정확 검색키 매칭
        for k, v in gold.items():
            if v.get('search_key') == search_key:
                return k, v

        # 검색키 시작 매칭 — 원본 길이의 50% 이상 겹쳐야 함
        matches = [(k, v) for k, v in gold.items()
                   if v.get('search_key', '').startswith(search_key)
                   and len(search_key) >= len(v.get('search_key', '')) * 0.5]
        if matches:
            matches.sort(key=lambda x: len(x[0]))
            return matches[0][0], matches[0][1]

    return None, None


# ── 1인분 서빙 사이즈 추정 (Gold DB 100g 기준 보정용) ──
# Gold DB의 96.8%가 serving=100g이라 실제 1인분과 맞지 않음
# 음식 이름 패턴으로 현실적인 1인분 크기를 추정
#
# 원칙:
#   - CORE_FOODS에 수동 설정된 serving은 그대로 사용 (이미 현실적)
#   - Gold DB에서 serving=100g인 항목만 보정 대상
#   - 음식 이름의 접미사/키워드로 카테고리 판별 → 해당 카테고리의 표준 1인분 적용
#   - 어디에도 해당 안 되면 보정하지 않음 (100g 유지)

def _estimate_realistic_serving(food_name, gold_serving):
    """음식 이름으로 현실적인 1인분 서빙 사이즈(g) 추정

    gold_serving이 이미 100g이 아니면(=수동 설정됨) 그대로 반환.
    100g인 경우에만 음식 카테고리를 추정하여 보정.
    """
    # 이미 현실적 서빙이 설정된 경우 그대로
    if gold_serving != 100:
        return gold_serving

    name = food_name.strip()

    # ── 초밥 (최우선: "밥" 패턴보다 먼저 체크) ──
    # "연어초밥", "참치초밥" → 1개 (30g)
    # "초밥" 단독 → 1인분 세트 (200g)
    if '초밥' in name:
        if name == '초밥' or name == '유부초밥':
            return 200
        return 30  # 개별 초밥 1개

    # ── 스시 ──
    if name == '스시':
        return 200

    # ── 국/탕 (400~500g, 물 포함 요리) ──
    if any(name.endswith(s) for s in ['국', '탕', '곰탕', '설렁탕', '해장국']):
        return 500
    if any(s in name for s in ['국밥', '곰탕', '설렁탕', '갈비탕', '삼계탕',
                                 '매운탕', '알탕', '추어탕', '대구탕', '해물탕',
                                 '선지국', '미역국', '콩나물국', '시래기국',
                                 '육개장', '어묵탕', '조개탕', '감자탕', '순대국',
                                 '뼈해장국', '오뎅탕', '닭곰탕', '꽃게탕']):
        return 500

    # ── 찌개/전골 (350~400g) ──
    if any(name.endswith(s) for s in ['찌개', '전골']):
        return 400
    if any(s in name for s in ['김치찌개', '된장찌개', '순두부찌개', '부대찌개',
                                 '동태찌개', '참치찌개', '갈치찌개',
                                 '곱창전골', '김치전골', '버섯전골']):
        return 400

    # ── 면류 ──
    if any(s in name for s in ['면', '국수', '우동', '라멘', '소바', '냉면',
                                 '짬뽕', '짜장', '칼국수', '수제비', '쫄면',
                                 '울면', '잔뽕', '밀면', '막국수', '쌀국수',
                                 '콩국수', '잔치국수', '비빔국수', '파스타']):
        # 건면/라면은 제외 - CORE_FOODS에서 이미 처리
        if '건면' in name or '건조' in name:
            return 100
        # 국물 없는 비빔류/볶음류는 300g
        if any(s in name for s in ['비빔', '볶음', '쫄면', '비빔국수']):
            return 300
        # 국물면 (칼국수, 짬뽕 등) 450g
        return 450

    # ── 파스타/리조또 (300g) ──
    if any(s in name for s in ['파스타', '스파게티', '리조또', '뇨끼',
                                 '라자냐', '라비올리', '펜네', '링귀네',
                                 '카넬로니', '페투치네']):
        return 300

    # ── 밥류/덮밥 (350~400g) ──
    if any(name.endswith(s) for s in ['밥', '라이스', '덮밥', '비빔밥']):
        return 400
    if any(s in name for s in ['오므라이스', '카레라이스', '하이라이스',
                                 '제육덮밥', '규동', '텐동', '오야코동',
                                 '돈부리', '가츠동', '잡채밥', '볶음밥']):
        return 400

    # ── 김밥 (1줄 = 230~250g) ──
    if '김밥' in name:
        return 230

    # ── 죽 (300g) ──
    if '죽' in name:
        return 300

    # ── 중식 요리 (반찬/공유 요리) ──
    if any(s in name for s in ['탕수육', '깐풍기', '유린기', '마파두부',
                                 '팔보채', '유산슬', '양장피', '깐쇼새우',
                                 '칠리새우', '라조기', '깐풍새우',
                                 '고추잡채', '해파리냉채', '꿔바로우',
                                 '경장육사']):
        return 250
    # 마라탕, 훠궈는 국물 요리 (이미 국/탕에서 처리되지만 안전장치)
    if any(s in name for s in ['마라탕', '훠궈']):
        return 500

    # ── 구이/스테이크 (150~250g) ──
    if any(s in name for s in ['스테이크', '함박스테이크', '폭찹']):
        return 250  # 스테이크류는 250g
    if any(s in name for s in ['구이', '불고기', '갈비',
                                 '삼겹살', '목살', '항정살', '가브리살']):
        return 200

    # ── 찜 (250g) ──
    if name.endswith('찜') or '찜' in name:
        return 250

    # ── 볶음 ──
    if name.endswith('볶음'):
        # 메인 요리급 볶음 (200~250g)
        if any(s in name for s in ['낙지볶음', '쭈꾸미볶음', '오징어볶음', '해물볶음',
                                     '제육볶음', '곱창볶음']):
            return 250
        # 반찬 볶음 (어묵볶음, 오징어채볶음, 멸치볶음 등) — 소량
        return 100  # 반찬 1인분은 50~70g이지만 Gold DB가 이미 100g 기준

    # ── 전/부침 (150~200g) ──
    if any(name.endswith(s) for s in ['전', '부침', '파전', '빈대떡']):
        return 200

    # ── 기본값: 보정하지 않음 ──
    return 100


def match_with_db(analysis, foods_db):
    """
    AI 분석 결과를 DB와 매칭하여 보정
    우선순위: 골드 테이블(1,771건 검증) → 기존 DB(품질 통과만) → AI 추정값 유지

    골드 테이블 = CORE_FOODS(355건 수동 검증) + DB 고품질(1,625건, 5대 영양소+서빙 완비)
    모두 100g당 기준으로 통일되어 있음
    """
    if "error" in analysis or "foods" not in analysis:
        return analysis

    for food in analysis["foods"]:
        ai_name = (food.get("name_ko") or "").strip()
        if not ai_name:
            food['db_matched'] = False
            food['source'] = 'AI_ESTIMATED'
            continue

        # ── 1순위: 골드 테이블 (1,771건, 100g 기준 통일) ──
        gold_key, gold_data = _search_gold(ai_name)
        if gold_data:
            gold_serving = gold_data.get('serving') or 100
            # AI가 서빙 사이즈를 추정했으면 그걸 사용
            ai_serving = food.get('estimated_serving_g')
            if ai_serving:
                effective_serving = ai_serving
            else:
                # AI 추정 없으면 → 현실적 서빙 사이즈로 보정
                effective_serving = _estimate_realistic_serving(ai_name, gold_serving)
            ratio = effective_serving / 100  # 골드 테이블은 100g 기준

            food['db_matched'] = True
            food['db_name'] = gold_key
            food['calories_kcal'] = round(gold_data['cal'] * ratio, 1)
            food['protein_g'] = round(gold_data['prot'] * ratio, 1)
            food['carbs_g'] = round(gold_data['carbs'] * ratio, 1)
            food['fat_g'] = round(gold_data['fat'] * ratio, 1)
            food['fiber_g'] = round(gold_data.get('fiber', 0) * ratio, 1)
            food['sodium_mg'] = round(gold_data.get('sodium', 0) * ratio, 1)
            food['sugar_g'] = round(gold_data.get('sugar', 0) * ratio, 1)
            food['source'] = 'GOLD_REF' if gold_data.get('source') == 'core' else 'GOLD_DB'
            continue

        # ── 2순위: 기존 SQLite DB (품질 필터 + 관련성 검증) ──
        exact, partials = search_food_db(ai_name)
        match = exact or (partials[0] if partials else None)

        # 관련성 체크: 부분매칭된 이름이 원본과 너무 다르면 기각
        if match and not exact:
            db_name = match.get('name_ko', '')
            # 원본 이름이 DB 이름에 포함되어 있거나 그 반대여야 함
            # "물" → "냉면 (물냉면)" 같은 느슨한 매칭 방지
            name_clean = ai_name.replace(' ', '')
            db_clean = db_name.replace(' ', '').replace('(', '').replace(')', '')
            # 최소 3자 이상 겹치거나, 짧은 이름이 긴 이름의 50% 이상이어야 함
            shorter = min(len(name_clean), len(db_clean))
            if shorter < 3 or (shorter / max(len(name_clean), len(db_clean)) < 0.3):
                match = None  # 관련성 부족 → 기각

        if match:
            quality = _db_quality_score(match)

            # 품질 3점 이상만 사용 (칼로리+단백질+탄수+지방+서빙 중 3개 이상)
            if quality >= 3:
                db_serving = match.get('serving_size_g') or 0
                ai_serving = food.get('estimated_serving_g') or 0

                if db_serving > 0 and ai_serving > 0:
                    ratio = ai_serving / db_serving
                else:
                    ratio = 1.0

                food['db_matched'] = True
                food['db_food_id'] = match.get('food_id', '')
                food['db_name'] = match.get('name_ko', '')

                for field_pair in [
                    ('calories_kcal', 'calories_kcal'),
                    ('protein_g', 'protein_g'),
                    ('carbs_g', 'carbs_g'),
                    ('fat_g', 'fat_g'),
                    ('fiber_g', 'fiber_g'),
                    ('sodium_mg', 'sodium_mg'),
                    ('sugar_g', 'sugar_g'),
                ]:
                    ai_field, db_field = field_pair
                    db_val = match.get(db_field)
                    if db_val is not None and isinstance(db_val, (int, float)) and db_val > 0:
                        food[ai_field] = round(db_val * ratio, 1)

                # ── 안전장치: 수분→탄수화물 오염 데이터 차단 ──
                # 탄수화물(g)*4 > 칼로리*2 이면 물리적으로 불가능 (수분이 탄수화물로 잘못 들어간 데이터)
                db_cal = food.get('calories_kcal', 0) or 0
                db_carbs = food.get('carbs_g', 0) or 0
                if db_cal > 0 and db_carbs > 30 and db_carbs * 4 > db_cal * 2:
                    food['db_matched'] = False
                    food['source'] = 'AI_ESTIMATED'
                    # AI 추정값 복원 — DB 값이 오염되었으므로
                    continue

                food['source'] = 'DB_MATCHED'
                continue

        # ── 3순위: AI 추정값 그대로 유지 ──
        food['db_matched'] = False
        food['source'] = 'AI_ESTIMATED'

    # meal_summary 재계산
    if "meal_summary" in analysis:
        analysis["meal_summary"]["total_calories"] = round(sum(f.get("calories_kcal", 0) or 0 for f in analysis["foods"]), 1)
        analysis["meal_summary"]["total_protein"] = round(sum(f.get("protein_g", 0) or 0 for f in analysis["foods"]), 1)
        analysis["meal_summary"]["total_carbs"] = round(sum(f.get("carbs_g", 0) or 0 for f in analysis["foods"]), 1)
        analysis["meal_summary"]["total_fat"] = round(sum(f.get("fat_g", 0) or 0 for f in analysis["foods"]), 1)

    return analysis


def format_result(analysis):
    """분석 결과를 사람이 읽기 좋은 텍스트로 변환"""
    if "error" in analysis:
        return f"오류: {analysis['error']}"

    lines = []
    lines.append("=" * 50)
    lines.append("  NutriLens 음식 분석 결과")
    lines.append("=" * 50)

    for i, food in enumerate(analysis.get("foods", []), 1):
        matched = "DB" if food.get('db_matched') else "AI추정"
        conf = food.get('confidence', 0)
        lines.append(f"\n[{i}] {food['name_ko']} ({food.get('name_en', '')})  [{matched}] 확신도:{conf:.0%}")
        lines.append(f"    양: {food.get('estimated_serving_g', '?')}g")
        lines.append(f"    칼로리: {food.get('calories_kcal', '?')} kcal")
        lines.append(f"    단백질: {food.get('protein_g', '?')}g  |  탄수화물: {food.get('carbs_g', '?')}g  |  지방: {food.get('fat_g', '?')}g")
        lines.append(f"    식이섬유: {food.get('fiber_g', '?')}g  |  나트륨: {food.get('sodium_mg', '?')}mg  |  당류: {food.get('sugar_g', '?')}g")

    summary = analysis.get("meal_summary", {})
    if summary:
        lines.append(f"\n{'─' * 50}")
        lines.append(f"  총 칼로리: {summary.get('total_calories', '?')} kcal")
        lines.append(f"  단백질 {summary.get('total_protein', '?')}g / 탄수 {summary.get('total_carbs', '?')}g / 지방 {summary.get('total_fat', '?')}g")
        lines.append(f"  식사 유형: {summary.get('meal_type', '?')}  |  건강 점수: {summary.get('health_score', '?')}/10")
        lines.append(f"\n  AI 코멘트: {summary.get('one_line_comment', '')}")
    lines.append("=" * 50)

    return '\n'.join(lines)


# ── 메인 실행 ──
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="NutriLens AI 음식 분석기")
    parser.add_argument("--image", "-i", required=True, help="분석할 음식 사진 경로")
    parser.add_argument("--api-key", "-k", default=None, help="OpenAI API 키")
    parser.add_argument("--model", "-m", default="gpt-4o", help="사용할 모델 (기본: gpt-4o)")
    parser.add_argument("--json", action="store_true", help="JSON 형식으로 출력")
    parser.add_argument("--no-db", action="store_true", help="DB 매칭 건너뛰기")
    args = parser.parse_args()

    print(f"이미지 분석 중: {args.image}")
    print("GPT-4o Vision API에 요청 중...")

    # 1. AI 분석
    analysis = analyze_food_image(args.image, api_key=args.api_key, model=args.model)

    # 2. DB 매칭 (선택)
    if not args.no_db and "error" not in analysis:
        print("DB 매칭 중...")
        foods_db = load_food_db()
        if foods_db:
            analysis = match_with_db(analysis, foods_db)
            print(f"DB 매칭 완료 ({len(foods_db)}종 DB 사용)")

    # 3. 출력
    if args.json:
        print(json.dumps(analysis, ensure_ascii=False, indent=2))
    else:
        print(format_result(analysis))
