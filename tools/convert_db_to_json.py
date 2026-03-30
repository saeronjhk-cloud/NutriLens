#!/usr/bin/env python3
"""
엑셀 DB → 경량 JSON 변환
서버 시작 속도를 대폭 개선합니다.
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

PROJECT_DIR = Path(__file__).parent.parent
DB_XLSX = PROJECT_DIR / "NutriLens_음식DB.xlsx"
DB_JSON = PROJECT_DIR / "NutriLens_음식DB.json"

print(f"엑셀 DB 로딩: {DB_XLSX.name}")
wb = openpyxl.load_workbook(DB_XLSX, read_only=True, data_only=True)
ws = wb["음식DB_전체"]

headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"컬럼: {headers}")

# 필요한 컬럼만 추출
KEEP_FIELDS = {
    "food_id", "name_ko", "category", "subcategory",
    "calories_kcal", "protein_g", "carbs_g", "fat_g",
    "fiber_g", "sodium_mg", "sugar_g", "serving_size_g",
}

foods = []
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    rec = {}
    for h, v in zip(headers, row):
        if h in KEEP_FIELDS and v is not None:
            rec[h] = v
    if rec.get("name_ko"):
        foods.append(rec)
        count += 1
        if count % 50000 == 0:
            print(f"  {count:,}종 처리...")

wb.close()
print(f"총 {count:,}종 추출")

# JSON 저장
print(f"JSON 저장 중: {DB_JSON.name}")
with open(DB_JSON, "w", encoding="utf-8") as f:
    json.dump(foods, f, ensure_ascii=False)

size_mb = DB_JSON.stat().st_size / 1024 / 1024
print(f"완료! {size_mb:.1f}MB (엑셀 대비 크게 축소)")
