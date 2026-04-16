#!/usr/bin/env python3
"""
식약안전나라 (MFDS) 식품영양성분 DB API 연동 도구
──────────────────────────────────────────────────
공공 API에서 영양성분 데이터를 가져와 NutriLens DB(엑셀)에 추가합니다.

API: I2790 (식품영양성분DB정보)
URL: https://openapi.foodsafetykorea.go.kr/api/{KEY}/I2790/json/{start}/{end}

사용법:
  python tools/mfds_importer.py                    # 전체 데이터 가져오기
  python tools/mfds_importer.py --search 김치      # 특정 음식 검색
  python tools/mfds_importer.py --max 500          # 최대 500개만
  python tools/mfds_importer.py --dry-run          # DB 저장 없이 미리보기
"""

import os
import sys
import json
import time
import argparse
import urllib.request
import urllib.parse
import urllib.error
from pathlib import Path
from datetime import datetime

# ── .env 로드 ──
def load_env():
    env_paths = [
        Path(__file__).parent.parent / '.env',
        Path.cwd() / '.env',
    ]
    for p in env_paths:
        if p.exists():
            with open(p) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, val = line.split('=', 1)
                        os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))
            break

load_env()

# ── 설정 ──
# 식품안전나라 API
FOODSAFETY_BASE_URL = "http://openapi.foodsafetykorea.go.kr/api"
# 공공데이터포털 API (data.go.kr)
DATA_GO_KR_ENDPOINTS = [
    "http://apis.data.go.kr/1471000/FoodNtrCpntDbInfo02/getFoodNtrCpntDbInq02",
    "https://apis.data.go.kr/1471000/FoodNtrCpntDbInfo02/getFoodNtrCpntDbInq02",
]

SERVICE_ID = "I2790"
PROJECT_DIR = Path(__file__).parent.parent
DB_PATH = PROJECT_DIR / "NutriLens_음식DB.xlsx"
BATCH_SIZE = 100
API_MODE = None  # "foodsafety" 또는 "data_go_kr" (자동 감지)


# ── API 필드 매핑 ──
# I2790 응답 필드 → 우리 DB 필드
FIELD_MAP = {
    "DESC_KOR": "name_ko",           # 식품이름
    "FOOD_CD": "food_id_mfds",       # 식품코드
    "GROUP_NAME": "subcategory",     # 식품대분류
    "SERVING_SIZE": "serving_size_g", # 1회제공량(g)
    "NUTR_CONT1": "calories_kcal",   # 열량(kcal)
    "NUTR_CONT2": "carbs_g",         # 탄수화물(g)
    "NUTR_CONT3": "protein_g",       # 단백질(g)
    "NUTR_CONT4": "fat_g",           # 지방(g)
    "NUTR_CONT5": "sugar_g",         # 당류(g)
    "NUTR_CONT6": "sodium_mg",       # 나트륨(mg)
    "NUTR_CONT7": "cholesterol_mg",  # 콜레스테롤(mg)
    "NUTR_CONT8": "sat_fat_g",       # 포화지방산(g)
    "NUTR_CONT9": "trans_fat_g",     # 트랜스지방(g)
    "MAKER_NAME": "brand",           # 제조사
    "RESEARCH_YEAR": "research_year", # 조사년도
}


def parse_xml_response(raw_xml, service_id):
    """XML 응답을 파싱하여 (rows, total_count) 반환"""
    import xml.etree.ElementTree as ET
    root = ET.fromstring(raw_xml)

    # 에러 체크
    result_code = root.findtext('.//CODE')
    result_msg = root.findtext('.//MSG')
    if result_code and result_code != "INFO-000":
        return [], 0, f"{result_code}: {result_msg}"

    total = int(root.findtext('.//total_count') or 0)

    rows = []
    for item in root.findall('.//row'):
        row_dict = {}
        for child in item:
            row_dict[child.tag] = child.text
        rows.append(row_dict)

    return rows, total, None


def fetch_api(api_key, service_id, start, end, search_name=None, verbose=False):
    """API 호출 — JSON과 XML 모두 지원"""
    for fmt in ["json", "xml"]:
        url = f"{FOODSAFETY_BASE_URL}/{api_key}/{service_id}/{fmt}/{start}/{end}"
        if search_name:
            encoded = urllib.parse.quote(search_name)
            url += f"/DESC_KOR={encoded}"

        try:
            req = urllib.request.Request(url)
            with urllib.request.urlopen(req, timeout=30) as resp:
                raw = resp.read().decode('utf-8')
        except Exception as e:
            if verbose:
                print(f"    [{fmt}] 요청 실패: {e}")
            continue

        if not raw.strip():
            if verbose:
                print(f"    [{fmt}] 빈 응답")
            continue

        if verbose:
            preview = raw.strip()[:300].replace('\n', ' ')
            print(f"    [{fmt}] 응답 ({len(raw)}자): {preview}")

        # JSON 시도
        if fmt == "json" and not raw.strip().startswith('<'):
            try:
                data = json.loads(raw)
                if "RESULT" in data:
                    code = data["RESULT"].get("CODE", "")
                    msg = data["RESULT"].get("MSG", "")
                    return [], 0, f"{code}: {msg}", fmt
                total = int(data.get(service_id, {}).get("total_count", 0))
                rows = data.get(service_id, {}).get("row", [])
                return rows, total, None, fmt
            except json.JSONDecodeError:
                pass

        # XML 파싱
        if raw.strip().startswith('<'):
            try:
                rows, total, err = parse_xml_response(raw, service_id)
                if verbose:
                    print(f"    [xml 파싱] total={total}, rows={len(rows)}, err={err}")
                return rows, total, err, "xml"
            except Exception as e:
                return [], 0, f"XML 파싱 실패: {e}", fmt

    return [], 0, "응답 없음", "none"


def fetch_data_go_kr(api_key, endpoint, page_no=1, num_of_rows=100, food_name=None):
    """공공데이터포털 (data.go.kr) API 호출"""
    params = {
        "serviceKey": api_key,
        "pageNo": str(page_no),
        "numOfRows": str(num_of_rows),
        "type": "json",
    }
    if food_name:
        params["food_name"] = food_name

    query = urllib.parse.urlencode(params, safe='=+/')
    url = f"{endpoint}?{query}"

    try:
        req = urllib.request.Request(url)
        req.add_header("Accept", "application/json")
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read().decode('utf-8')
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace')[:300]
        return [], 0, f"HTTP {e.code}: {body}"
    except Exception as e:
        return [], 0, f"요청 실패: {e}"

    if not raw.strip():
        return [], 0, "빈 응답"

    # XML 응답인 경우
    if raw.strip().startswith('<'):
        try:
            import xml.etree.ElementTree as ET
            root = ET.fromstring(raw)
            # 에러 체크
            err_code = root.findtext('.//returnReasonCode') or root.findtext('.//resultCode')
            err_msg = root.findtext('.//returnAuthMsg') or root.findtext('.//resultMsg')
            if err_code and err_code != "00":
                return [], 0, f"{err_code}: {err_msg}"
            # 데이터 파싱
            total = int(root.findtext('.//totalCount') or root.findtext('.//total_count') or 0)
            rows = []
            for item in root.findall('.//item') or root.findall('.//row'):
                row_dict = {}
                for child in item:
                    row_dict[child.tag] = child.text
                rows.append(row_dict)
            return rows, total, None
        except Exception as e:
            return [], 0, f"XML 파싱 실패: {e}"

    # JSON 응답
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return [], 0, f"JSON 파싱 실패 (앞부분: {raw[:200]})"

    # 에러 체크 (다양한 에러 형식 대응)
    if "response" in data:
        header = data["response"].get("header", {})
        result_code = header.get("resultCode", "")
        result_msg = header.get("resultMsg", "")
        if result_code and result_code != "00":
            return [], 0, f"{result_code}: {result_msg}"

        body = data["response"].get("body", {})
        total = int(body.get("totalCount", 0))
        items = body.get("items", {})
        if isinstance(items, dict):
            rows = items.get("item", [])
        elif isinstance(items, list):
            rows = items
        else:
            rows = []
        if isinstance(rows, dict):
            rows = [rows]
        return rows, total, None

    # 최상위에 header + body가 있는 경우 (공공데이터포털 신규 형식)
    if "header" in data and "body" in data:
        code = data["header"].get("resultCode", "")
        msg = data["header"].get("resultMsg", "")
        if code and code != "00":
            return [], 0, f"{code}: {msg}"

        body = data.get("body", {})
        total = int(body.get("totalCount", 0))
        items = body.get("items", [])
        if isinstance(items, dict):
            rows = items.get("item", [])
        elif isinstance(items, list):
            rows = items
        else:
            rows = []
        if isinstance(rows, dict):
            rows = [rows]
        return rows, total, None

    return [], 0, f"알 수 없는 응답 형식 (키: {list(data.keys())[:5]})"


def fetch_total_count(api_key, api_key_datagokr=None):
    """전체 데이터 수 조회 — 두 API 시스템 자동 시도"""
    global SERVICE_ID, API_MODE

    # 1. 공공데이터포털 (data.go.kr) 시도
    if api_key_datagokr:
        print("\n  [공공데이터포털] 시도 중...")
        for endpoint in DATA_GO_KR_ENDPOINTS:
            short_url = endpoint.split("/")[-1]
            print(f"    엔드포인트: {short_url}...")
            rows, total, err = fetch_data_go_kr(api_key_datagokr, endpoint, page_no=1, num_of_rows=1)
            if err:
                print(f"    → {err}")
                continue
            if total > 0:
                API_MODE = "data_go_kr"
                print(f"    → 성공! 총 데이터: {total:,}종")
                return total, endpoint
        print("    공공데이터포털 모든 엔드포인트 실패")

    # 2. 식품안전나라 API 시도
    if api_key:
        print("\n  [식품안전나라] 시도 중...")
        # ⚠️ I0750, I2710은 NUTR_CONT2가 '수분'이라 탄수화물과 혼동됨
        # 반드시 I2790만 사용 (NUTR_CONT2 = 탄수화물)
        for sid in ["I2790"]:
            print(f"    서비스: {sid}...")
            rows, total, err, fmt = fetch_api(api_key, sid, 1, 1, verbose=True)
            if err:
                print(f"    → {err}")
                continue
            if total > 0:
                API_MODE = "foodsafety"
                SERVICE_ID = sid
                print(f"    → 성공! 총 데이터: {total:,}종")
                return total, None

    print(f"\n  모든 API 시도 실패.")
    return 0, None


def fetch_batch(api_key, start, end, search_name=None, **kwargs):
    """API에서 데이터 배치 가져오기"""
    if API_MODE == "data_go_kr":
        endpoint = kwargs.get("endpoint")
        api_key_dgk = kwargs.get("api_key_datagokr")
        page_no = (start - 1) // BATCH_SIZE + 1
        rows, total, err = fetch_data_go_kr(api_key_dgk, endpoint, page_no=page_no, num_of_rows=BATCH_SIZE, food_name=search_name)
        if err:
            print(f"  배치 에러 (page {page_no}): {err}")
            return [], 0
        return rows, total
    else:
        rows, total, err, fmt = fetch_api(api_key, SERVICE_ID, start, end, search_name)
        if err:
            print(f"  배치 에러 ({start}-{end}): {err}")
            return [], 0
        return rows, total


def safe_float(val):
    """문자열을 float으로 안전하게 변환"""
    if val is None or val == "" or val == "-" or val == "N/A":
        return 0.0
    try:
        return round(float(str(val).replace(",", "")), 1)
    except (ValueError, TypeError):
        return 0.0


def classify_food(name, group_name, maker):
    """음식 이름/분류로 카테고리 자동 분류"""
    name_lower = (name or "").lower()
    group = (group_name or "").lower()
    maker_str = (maker or "").lower()

    # 프랜차이즈 판별
    franchise_keywords = [
        "맥도날드", "버거킹", "롯데리아", "kfc", "서브웨이",
        "스타벅스", "파리바게뜨", "뚜레쥬르", "배스킨라빈스",
        "교촌", "bbq", "bhc", "굽네", "네네", "도미노",
        "피자헛", "미스터피자", "이삭토스트", "맘스터치",
    ]
    for kw in franchise_keywords:
        if kw in name_lower or kw in maker_str:
            return "franchise"

    # 다이어트/피트니스
    diet_keywords = [
        "프로틴", "protein", "닭가슴살", "샐러드", "그래놀라",
        "오트밀", "제로", "다이어트", "저칼로리", "무설탕",
        "아사이", "퀴노아", "리코타", "두부면", "곤약",
    ]
    for kw in diet_keywords:
        if kw in name_lower:
            return "diet_fitness"

    # 외국 음식
    foreign_keywords = [
        "파스타", "피자", "스테이크", "햄버거", "타코",
        "초밥", "라멘", "카레", "난", "팟타이",
        "쌀국수", "리조또", "그라탕", "퀘사디아",
    ]
    for kw in foreign_keywords:
        if kw in name_lower:
            return "foreign_popular"

    # 간식/음료
    snack_keywords = [
        "과자", "쿠키", "초콜릿", "캔디", "아이스크림",
        "음료", "주스", "커피", "차", "우유", "요구르트",
        "젤리", "떡", "빵", "케이크", "탄산",
    ]
    for kw in snack_keywords:
        if kw in name_lower or kw in group:
            return "snack_drink"

    # 기본: 한식
    return "korean"


def convert_row(row, idx_offset=0):
    """API 응답 row를 우리 DB 형식으로 변환 (두 API 형식 모두 지원)"""
    # 음식 이름 (식품안전나라: DESC_KOR, 공공데이터포털: food_name 또는 FOOD_NM_KR)
    name_ko = (row.get("DESC_KOR") or row.get("food_name") or row.get("FOOD_NM_KR")
               or row.get("foodNm") or row.get("FOOD_NAME") or row.get("FOOD_NM") or "").strip()
    if not name_ko:
        return None

    # 분류명 (새 API: FOOD_CAT1_NM)
    group_name = (row.get("GROUP_NAME") or row.get("food_cat") or row.get("FOOD_CAT")
                  or row.get("foodCat") or row.get("FOOD_GROUP") or row.get("FOOD_CAT1_NM") or "")
    # 제조사 (새 API: DB_CLASS_NM = 품목대표/상용제품)
    maker = (row.get("MAKER_NAME") or row.get("maker_name") or row.get("MAKER_NM")
             or row.get("makerNm") or row.get("DB_CLASS_NM") or "")
    # 식품코드
    food_cd = (row.get("FOOD_CD") or row.get("food_cd") or row.get("FOOD_CODE")
               or row.get("foodCd") or row.get("ITEM_REPORT_NO") or str(idx_offset))

    category = classify_food(name_ko, group_name, maker)

    # 영양소 (두 API 형식 대응)
    def get_nutr(keys):
        for k in keys:
            val = row.get(k)
            if val is not None and val != "":
                return val
        return 0

    cal = safe_float(get_nutr(["NUTR_CONT1", "enerc", "ENERC", "AMT_NUM1", "calorie"]))
    prot = safe_float(get_nutr(["NUTR_CONT3", "prot", "PROT", "AMT_NUM3", "protein"]))
    carbs = safe_float(get_nutr(["NUTR_CONT2", "chocdf", "CHOCDF", "AMT_NUM2", "carbohydrate"]))
    fat = safe_float(get_nutr(["NUTR_CONT4", "fatce", "FATCE", "AMT_NUM4", "fat"]))
    serving = safe_float(get_nutr(["SERVING_SIZE", "serving_size", "SERVING_WT", "servingWt", "NUTRI_AMOUNT_SERVING"]))

    # ── 안전장치 1: 수분→탄수화물 오염 차단 ──
    # carbs*4 > cal*2 이면 수분 데이터가 탄수화물로 잘못 들어온 것
    if cal > 0 and carbs > 30 and carbs * 4 > cal * 2:
        carbs = max(0, round((cal - prot * 4 - fat * 9) / 4, 1))

    # ── 안전장치 2: 탄수화물 누락 복원 ──
    # carbs=0인데 칼로리가 단백질+지방만으로 설명 안 되면 역산
    if carbs == 0 and cal > 0:
        remaining_cal = cal - prot * 4 - fat * 9
        if remaining_cal > 8:  # 최소 2g 이상의 탄수화물
            carbs = max(0, round(remaining_cal / 4, 1))

    # ── 안전장치 3: 제공량 0g이면 서브카테고리 기반 기본값 ──
    if serving == 0:
        SERVING_DEFAULTS = {
            '과자류·빵류 또는 떡류': 30, '빵 및 과자류': 60, '빙과류': 100,
            '즉석식품류': 200, '면류': 120, '음료류': 200, '음료 및 차류': 200,
            '식육가공품 및 포장육': 100, '수산가공식품류': 100, '조미식품': 15,
            '잼류': 20, '장류': 15, '당류': 15, '식용유지류': 10,
            '코코아가공품류 또는 초콜릿류': 30, '농산가공식품류': 100,
            '절임류 또는 조림류': 50, '유가공품류': 200, '찌개 및 전골류': 400,
            '국 및 탕류': 400, '밥류': 400, '면 및 만두류': 500,
        }
        serving = SERVING_DEFAULTS.get(group_name, 100)

    return {
        "food_id": f"MFDS_{food_cd}",
        "name_ko": name_ko,
        "name_en": "",
        "category": category,
        "subcategory": group_name or "",
        "serving_size_g": serving,
        "calories_kcal": cal,
        "protein_g": prot,
        "carbs_g": carbs,
        "fat_g": fat,
        "fiber_g": safe_float(get_nutr(["NUTR_CONT10", "fibtg", "FIBTG", "dietary_fiber"])),
        "sodium_mg": safe_float(get_nutr(["NUTR_CONT6", "na", "NA", "AMT_NUM6", "sodium"])),
        "sugar_g": safe_float(get_nutr(["NUTR_CONT5", "sugar", "SUGAR", "AMT_NUM5"])),
        "sat_fat_g": safe_float(get_nutr(["NUTR_CONT8", "fasat", "FASAT", "AMT_NUM8", "saturated_fat"])),
        "cholesterol_mg": safe_float(get_nutr(["NUTR_CONT7", "chole", "CHOLE", "AMT_NUM7", "cholesterol"])),
        "source": "MFDS_API",
        "brand": maker or "",
        "cooking_method": "",
        "tags": f"공공데이터,{group_name}" if group_name else "공공데이터",
    }


def load_existing_db():
    """기존 DB 로드"""
    if not DB_PATH.exists():
        return []

    try:
        import openpyxl
    except ImportError:
        print("openpyxl 필요: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(DB_PATH, read_only=True, data_only=True)
    ws = wb['음식DB_전체']
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    existing = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing.append(dict(zip(headers, row)))
    wb.close()
    return existing


def save_to_excel(all_foods):
    """전체 데이터를 엑셀로 저장"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl 필요: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # ── 메인 시트: 음식DB_전체 ──
    ws = wb.active
    ws.title = "음식DB_전체"

    headers = [
        "food_id", "name_ko", "name_en", "category", "subcategory",
        "serving_size_g", "calories_kcal", "protein_g", "carbs_g", "fat_g",
        "fiber_g", "sodium_mg", "sugar_g", "sat_fat_g", "cholesterol_mg",
        "source", "brand", "cooking_method", "tags"
    ]

    # 헤더 스타일
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='333333'),
        right=Side(style='thin', color='333333'),
        top=Side(style='thin', color='333333'),
        bottom=Side(style='thin', color='333333')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 카테고리별 색상
    cat_colors = {
        "korean": "E8F5E9",
        "diet_fitness": "E3F2FD",
        "foreign_popular": "FFF3E0",
        "franchise": "F3E5F5",
        "snack_drink": "FFF9C4",
    }

    for row_idx, food in enumerate(all_foods, 2):
        cat = food.get("category", "korean")
        fill = PatternFill(start_color=cat_colors.get(cat, "FFFFFF"),
                          end_color=cat_colors.get(cat, "FFFFFF"),
                          fill_type="solid")

        for col_idx, key in enumerate(headers, 1):
            val = food.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(name="맑은 고딕", size=9)

    # 열 너비 조정
    col_widths = {
        1: 18, 2: 25, 3: 20, 4: 15, 5: 15,
        6: 12, 7: 12, 8: 10, 9: 10, 10: 10,
        11: 10, 12: 10, 13: 10, 14: 10, 15: 12,
        16: 12, 17: 15, 18: 12, 19: 20,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # 필터 적용
    ws.auto_filter.ref = f"A1:S{len(all_foods) + 1}"

    # ── 카테고리 요약 시트 ──
    ws2 = wb.create_sheet("카테고리_요약")
    cat_names = {
        "korean": "한식",
        "diet_fitness": "다이어트/피트니스",
        "foreign_popular": "외국음식",
        "franchise": "프랜차이즈",
        "snack_drink": "간식/음료",
    }

    ws2.cell(row=1, column=1, value="카테고리").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="한국어명").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="수량").font = Font(bold=True)
    ws2.cell(row=1, column=4, value="출처: 기존").font = Font(bold=True)
    ws2.cell(row=1, column=5, value="출처: MFDS API").font = Font(bold=True)

    from collections import Counter
    cat_counts = Counter(f.get("category", "korean") for f in all_foods)
    source_counts = {}
    for f in all_foods:
        cat = f.get("category", "korean")
        src = f.get("source", "")
        if cat not in source_counts:
            source_counts[cat] = {"existing": 0, "mfds": 0}
        if src == "MFDS_API":
            source_counts[cat]["mfds"] += 1
        else:
            source_counts[cat]["existing"] += 1

    for i, (cat, count) in enumerate(cat_counts.most_common(), 2):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=cat_names.get(cat, cat))
        ws2.cell(row=i, column=3, value=count)
        ws2.cell(row=i, column=4, value=source_counts.get(cat, {}).get("existing", 0))
        ws2.cell(row=i, column=5, value=source_counts.get(cat, {}).get("mfds", 0))

    total_row = len(cat_counts) + 2
    ws2.cell(row=total_row, column=1, value="합계").font = Font(bold=True)
    ws2.cell(row=total_row, column=3, value=len(all_foods)).font = Font(bold=True)

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # 저장
    wb.save(DB_PATH)
    print(f"\n  DB 저장 완료: {DB_PATH}")
    print(f"  총 {len(all_foods)}종")


def main():
    parser = argparse.ArgumentParser(description="식약안전나라 영양DB API 연동")
    parser.add_argument("--search", "-s", default=None, help="음식 이름으로 검색")
    parser.add_argument("--max", "-m", type=int, default=5000, help="최대 가져올 수 (기본: 5000)")
    parser.add_argument("--dry-run", action="store_true", help="DB 저장 없이 미리보기")
    parser.add_argument("--api-key", default=None, help="MFDS API 키")
    args = parser.parse_args()

    api_key = args.api_key or os.environ.get("MFDS_API_KEY")
    api_key_datagokr = os.environ.get("DATA_GO_KR_API_KEY")

    if not api_key and not api_key_datagokr:
        print("API 키가 없습니다.")
        print("  .env 파일에 MFDS_API_KEY=... 또는 DATA_GO_KR_API_KEY=... 추가하거나")
        print("  --api-key 옵션으로 전달하세요.")
        sys.exit(1)

    print()
    print("=" * 55)
    print("  영양DB API 가져오기")
    print("=" * 55)
    if api_key_datagokr:
        print(f"  공공데이터포털 키: {api_key_datagokr[:8]}...")
    if api_key:
        print(f"  식품안전나라 키: {api_key[:8]}...")

    # 1. 전체 데이터 수 확인
    print("\n[1/4] API 연결 확인 중...")
    total, endpoint = fetch_total_count(api_key, api_key_datagokr)
    if total == 0:
        print("  API에서 데이터를 가져올 수 없습니다.")
        print("  API 키를 확인해주세요.")
        sys.exit(1)
    print(f"  API 전체 데이터: {total:,}종")

    # 가져올 수 결정
    fetch_count = min(args.max, total)
    print(f"  가져올 데이터: {fetch_count:,}종")

    # 2. 기존 DB 로드
    print("\n[2/4] 기존 DB 로드 중...")
    existing = load_existing_db()
    existing_names = set(f.get("name_ko", "") for f in existing)
    print(f"  기존 DB: {len(existing)}종")

    # 3. API에서 데이터 가져오기
    print(f"\n[3/4] API에서 데이터 가져오는 중...")
    new_foods = []
    duplicates = 0
    errors = 0

    for start in range(1, fetch_count + 1, BATCH_SIZE):
        end = min(start + BATCH_SIZE - 1, fetch_count)
        rows, _ = fetch_batch(api_key, start, end, search_name=args.search,
                              endpoint=endpoint, api_key_datagokr=api_key_datagokr)

        for row in rows:
            food = convert_row(row, start)
            if food is None:
                continue

            # 중복 체크 (이름 기준)
            if food["name_ko"] in existing_names:
                duplicates += 1
                continue

            existing_names.add(food["name_ko"])
            new_foods.append(food)

        progress = min(end, fetch_count)
        pct = progress / fetch_count * 100
        print(f"  {progress:,}/{fetch_count:,} ({pct:.0f}%) — 신규: {len(new_foods)}, 중복: {duplicates}", end="\r")

        # API 부하 방지
        time.sleep(0.3)

    print(f"\n  완료! 신규: {len(new_foods)}, 중복 제외: {duplicates}")

    if len(new_foods) == 0:
        print("\n  추가할 새로운 음식이 없습니다.")
        return

    # 카테고리 분포 출력
    from collections import Counter
    cat_dist = Counter(f["category"] for f in new_foods)
    cat_names = {
        "korean": "한식", "diet_fitness": "다이어트/피트니스",
        "foreign_popular": "외국음식", "franchise": "프랜차이즈",
        "snack_drink": "간식/음료",
    }
    print("\n  카테고리 분포:")
    for cat, cnt in cat_dist.most_common():
        print(f"    {cat_names.get(cat, cat)}: {cnt}종")

    # 미리보기: 처음 5개
    print(f"\n  샘플 (처음 5개):")
    for f in new_foods[:5]:
        print(f"    {f['name_ko']} | {f['calories_kcal']}kcal | {f['category']}")

    # 4. DB 저장
    if args.dry_run:
        print(f"\n  [dry-run] 저장하지 않음. 실제 저장하려면 --dry-run 없이 실행하세요.")
        return

    print(f"\n[4/4] DB에 저장 중...")
    all_foods = existing + new_foods
    save_to_excel(all_foods)

    print(f"\n  기존: {len(existing)}종 + 신규: {len(new_foods)}종 = 총: {len(all_foods)}종")
    print("=" * 55)
    print("  완료!")
    print("=" * 55)


if __name__ == "__main__":
    main()
